# -*- coding: utf-8 -*-
import re
import traceback
import pandas as pd
import numpy as np
import holidays
try:
    import pdfplumber as _pdfplumber
except ImportError:
    _pdfplumber = None
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import os
import threading
from datetime import datetime, date, timedelta
from calendar import monthrange
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
# Logo embutido em base64 (PNG 64x64 com transparência)
_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAWyklEQVR4nO1beZRcVZn/3fveq727uju9JJ1OOp3OgpAEAlEQUAOoxB3QBkURxoPjNo6Mg46OOg3OURQkHMFxFA/jwihSQdk3lSQQE0jSHUJS3Z1Ub+m1urq6qrr2esu93/xRVZ3qTm+ROB7P4Tvnnc57dZfv+33L/e69X0BEHK+DiIgRESt9b21tnXXMmW3PwLx8vvHm+e0kH3MxexqM8FIQW1tb+Vygzmx7BuZV5hKyAJAy128+n2/W3/6u6ExY1Glro8T0ePG91IoKWil9/0tMFAAWFGzHjh1FHpT5LHkmzyXf2GkD0N7ernZ3d2u7d++es++OHTtYcZLu7m7bXExhDgUQEff5fAvy1tLSQgWe+NatW3nR5Pv7+x1+v3/avEWeSogDUEFEagEdtTDxafnGTJ8von06YyyWfD6fMtfYra2tU7yXWuGuXbvUUp6IiAUCAftUfJhF4NP2qZmrwEJ+SUTMV2CgtbWV79q1S13k3KXzLArkEvNXS77Nu3qcFi1C2GnxYVcJI/sGB53Ffx8OBt2l7YpAzjV+f3+/YwHWpvoVrEApWFDRPdW5u54G+Xw+JRAI2EsZKgbGErNT9+07KWxbW5srEo+8lQIB++Do4LaBsZEvhf/857K2tjbvfHPNsLQ5LWBkZMRFRLy1tZX7/X5bAVA1GAy6iUgdGRlxLTTGoqm4ppYgay/5kQUmJsp7e3s3haOhr1lEn05b2c8MpyYfHJeCjqfjdxybjPqHcxka1TM3hlKpnROJ5HNGJvP2saGhTf7+/qW+vADTrMDn8ylzxaqCf/MSfpRSFyAiraurq4yINL/fbzsTZsBaWlrkbYXsijGm9/b21pmcV64HjocV9gItW5rqFMIRTRsXdYVjiGSSiOf6qKKq4mumnoUwTVHmcf/CLgm1DidWulxXrqgsf6xJsaWqU6laxtiVu/r7HasrK52NFRWxUvAZY1T6DsACQCX8SeQjPiMidffu3RQOhzOKoqjhcFiw/v5+R1NTU+4vlb5gRlRkpKO/f/NKVQq9tvYTfQZ/38vD4bNHDAEiiQoSYq3XTfUuO6902LhHVS3GwA0peNqyrLFMhp/IpHAimwE8Dl5t03BB2RKsc1cc0OLJB0aG9v86U/UmeXFjYxZE8Pl8SktLCzHGZAkArPhewqMyNDTkJSJqbGyMtbW1aU8++aS4/fbbJYr+8DoAUIjIPjQ2tql/dPS948nYga6c8dp3Xuu3bt5zjL57oEseCk0I07IEnQaNZdLWI0M94otH9lq39h+iR8cHkpGsuX0yHH5zYV51tuSnZEWZlfx+v63ETV5fOkxErJhwDI6PPxIjSf/dHYzdtLOTfnyohybS2WlCW1KSJSUJKUlIIkknH1H4bklJkuQ0MF4KD1m3dr1I3w11ULueOhBNpXceHxx8SxGIuXibGeQCgYB9tlzhLxV+qvOJ0dGd3Xoufesr3eLLfzpCfZGEKBVaTJdnUSQLfUs//XLoqHVzz056Oh6kdMb4effwQAswu9ZpliWUpgfHabnB6QrPAODIwEBlKJH44aF0lj72RBvd39YzxbElZupxkVLPQqLEKjrjIfHZjidNX3aYgql0Z+9gb9ESTtucZwK06E6BQMDe29vrHYlG7x4gopbfHzKe8A/KKWYLgkgpSZZoUZIkUwoypEWGyD964TGERYa0pmEw5RZCkJSSTJE3rKiepps7HrV+Ej2mR1K54aPdR1cAQFtbm0azmHXBGnghP1BKvp82aFNm1T8y8uJR3aSrfQf15zuHiYjIEIKEnF/vSaFTxMpQ1MpS1MxS1MpSxMoUnuxMs59hIHJq/KSRlbccf5p+ORGYTBni8Y6h4xcWQZgLgMIzJUNvb++8idcpVBhA6RsdvTVo5tKfe+ao+NUr3ZKIyJzH0YuaJCIKGRnqyyXohF76JKk/l6ATuTgZwiIiopxlUUckSv7JCTqaGKNJM3vKWOO5FF3vf1jupTRlUvon9+3b56TTMOu2tjZt0UGA8kmHDD5/2O14W9O/PxKIuxok5A0XrmGmIGgKw3AijYRu4OyaSghJ4DzPC2cnecoJgaQwAQ4oAByMIy2sfBsiWKoLGoB7DvbiuEigpsLABJ9EWkxi+5r3YrmjAkQESYQauxu3Nlwi7+/bA2/dxTc1Nzb23XbbbfuKvC4k05YtW8zFLgMMAA4fPuxWLl71rS5D5e1tg/TPl69lkghF+fyhOF7oDYMIUDjLdyLg2VAffMMdAICkpSNmZZEwdUTNHIJ6GpOWgUlTR9zSCykcYWgihyvrV+GqZWtxxZJ1kPDi671/QCA9DjAGxgBBEpu9y5UtahWeMYcu43X1Gz/1qQ97d5/GQc9iGxIAft5556UnNf62p/eHvddtqJUep50RAQVFw0EcbpFnbnd3ENFMDowB93V2IJCeBAAkLAMxU8ekqSMrBBjjiFpZxEUeAEF5CFxZjvQkYSxqQdVt2OBegSx34qnxABgASQTOOCQRrl9+Hj821mcdyYTvrXatbL6MMWuxAW5RLtBKxBljIpFLXLNnXKynUEy++50buSSCwhmkJIABii5RxTQkMjr+82E/3r1lOZQywnXV63Hj+nUAgLSQmDBNqIyjSmWoUBXETBNVDhsSWbMAAKHSUpEKm2gLTCBXLlBew/CWZWth04tZe97CJAhORWPXVK5le1ODynlVm74wODj4TQBBmrFXmNJm4TSKMSYWYwHsNoCCwaB7XNCd/u7Ekqs31hDnCsOMoR2SQw9l8Ytfvoqf3nA+yp0K9ndFcOPGdbBk3iUzlkQkJxHVBeKmQNISSJoCw6ksEpaYsgBHnMCTDG7TCRscIF1FNsVgGtODvMI4CIR31jQrg9GQFbSrN2lu9+cZY3J3PsycKlAeFElEfEEAKK99Enbl7oTqbhoKhKy3vKlWIWAqyE0xkyMoOYZPXrcBaxqX4HPvWI8dN78dAKByDikJCd2CKQHTAiKGRMSyEDMkyjUH0roFS0gQAfYMh5Ym8IREdUrDxRW14GkbZPpUmSQR7FzF+a46/odIQCytqjrm9/ttWwGaa1UogMAWdIHdhQDorSzv3Hs8xTfUuMlu1/JRfsbQNg4M+6PY6+aICQtWOYflYEjCxPIKO649Zy0MRUE2K2GBIyVMuG0a4jkCyhUkJUOZ0wkGwBgzUOmwYUNtBTI2CT3CkM4osDkKOiNMnfkUwi2uqFlFP4t2K8e1+Cc3bNjwYFHWuWRjjIkFAdgKkN/vtyWhbAwOJHDp2qqpUVkh/BdXgfXn1sBm58gJiaUqYCgExaXBVuaAWwOOH+9FW9cgWPNSELNgh4a0CUwYEjFi6B/J4dG+ffjI2zdj2zUr8PKeMFJBA2m7RCShwOICV1xaP21O4OQy2+xZwsxgB/XL7FmUpSY4cAIlfBapNAbMC8AuIpUxZg2Pj3/EgHpzsH9CNG1Zp8w2KABoNhXrNtVN/0gGdu1px4+f24/nhuNIrTkHH3BXwlmtYlwKVBEhmhEIjuZwsDuK5wMHcd+Ox3HdxefjmusuQePKtSWDCRTdms+YXxBBYYwvkTaRsPOGiJ784RJH2Yd25Fc6UdqWMUZEJIEFVoGt+dMUeGvUrpcHJshBjFdVOGkO+adRZ08ffvf8Xjy1twvD5IW6Zi0qt62AZeNIlNkx1J+AmRNYARWdhydQe6ELRm05at90PmLxEO48ehw//vYruLB6Oa6//FJc8bbNsDvt808KULPTy3vCYVy78pztBUFPWQWKICwIAAr+Y8G+FlywapciwTnPr8EnESjOkcnp2PnKYTz05F74A2lw5zLUNF+JjUsrkXNxZE0TDiMDLWeDZqlIRnXExgwgbYNiMShJC5abw1ZVheVXesCtFA70jeGlXQ9jxaNP4H0bN+KjH9yKxvo6cMamaaEYDpY7yuXRTJSN69kbBgcH9wPQ5xNwUYmQJeUHAA0em5Jfy2i64AAgSeKmr96Fj332AVhDm3DRig/jgsa3oUaphj1uAcEwEi+/DGV/D+xBCWdMoCzDoAmgbFLAGSF4upLIPf4qjCNhWBEJwR1YsrEBjTdtAP/CSvw42I7NV92CQx2B/JzyVOVWaDaWzeaQMawPqarqYYzJ+fYH81rAjsJfhdt/PjoWudGm5q/SitOWaoAzju995VOo8DyKQ3vbscXhBZcaJiJ9GI52IRIbgp5NoGnNu+BOqBDJLHgOcOcYvBnAozNUZzRMHBnFxIFOOFYugWtzDco2V0NtdCJ0cBh1YYYffe8WnPumNfk5+SkWACdXRVlZheopK7ulprw8XIxjfxEARdJhfsLjdkJHbC5/AgA0r1yOn93xT/jtszvxH3c+BNMixI1xcK5AUe1QbHZ4uYZkZxqZTBZLltnhyBA8CUI6kIZXZ9DsNphSgxhNINo3gcnnbRBVGt67bi3uuf2LqKmrWoBbxuOTcSTKE98Ohfr/VAuE5soIFwSgpgCsCow6NRUZIz/G7PZEME0BTVORTueguWpQU7cMMtiJdCwEYZoAAzyGDe44ILISZVUMHoOhPMkQO5qCV7NBEQQyBSAI3GtD1ZuXgNcoiA6m4CpzTaXdM1eBYlqgSwlJCjhjMV1RFtwPLLQKEADYof0BLP0fmYw584wtDwjL7wc0TcUvH38ed/zgRVx+wSdgVGmoa9yIUHIAowN+REZ7kI6Owisb4CYPPAkGW4RQp7tgZwpC8SHk0kmwaie8WxrhvbAKjiYFrgog8FgfWr50Bx770Tdgs9tARLMuRZO6Kasrqnh9Wdl2Oysf2UWkXva6XSCXuaa62m11ZcEBSYxzRnTS9KWU4JzD94fduPPO3bh2/U1IGgLZWBKaV0Ft7Xq4VjWjKhNCf48fYwPPYV2mBmsyZyE8bGEy3otjrA/DdTloV2+G99yl0CoZBLLIxrLQ4zrWXb0ePaIHN37zXvzs9i/A5bADLJ8FEtFUXBrJZLlT4ZTIJL9GRL9h+RyAATN3LosDIN8pZz7JRfrzxJkjFdPJU+mctfHBwwEIacI/chQNNWtQafcioeuwklkoEigrqwV/x7sQy4TQPtAH/6s7YYvkkD1rObBuPVyrl0O6LOREGmbEgOZQYa8og6I5ceKVEZihLI4NJBCOxtDUUI/iWQQrHKYAQHcihdq6alZpt99dyAM4K+Qzs9FCyyABQGhEtnk9iu6scmFkKEVAfgMyNQjnICLc9dV/xCMPfg5r3zmJg6kncHT4JZjJSZSRG27LDjVtgUVS8HiXoqJhLRKUQtA5CZurHLX150KJA4hKiJwd0lYBaTqROhjG6M/8aH4hi+9f8D7s//UP0NRQnz+gmLYKMQCQo6bk1YyNKort4WlKnIPmtQDGGPmIlA2MpUJkfLNyVcV9XZ0RWr+pZra2ICJsWNOMu7/WjC9//348/cQwYtkgaMyBpU2bUF6/HNxWhomhEYT2/RFlm86GqSeR1sfAXnwR3rdeiiwzYCQTSLd3IHmiB+56QAvn8D8//C5q6irzEs3wf1k4lOlPJElzlePcSm8vgFzhHGPeo7EFY0ALAAKxSWRGVp3lob17Bpi0BJiiTAuCxb+mKcA5AxMc53kvQf2KFehMH8JAbxvk8KuobFyN8eOHsHTrW2G6HUh1tMP74csRe+BpKPs15FgW6fgA3GtdWPaJRng2lCP0o06k0zlUCQnG8hZXSpIkOON4YSRMTZUVaND4iwCwFeC3z2P+wOIyQcnAKD2Sa6suo6R7eQXr6YjIotvNjMScMygKBxggczpclg3rl7wZF599FVbWnY3JEwEsP2cLaurPgRrLAYaEodvguuoyRJVhWPVp1N5wHiqvOQ+0tAaZGIOlE7iSH3e2yK8wBktK2hOKKWudPOPg8heMMdq6gPDAIiyg6AYNjA0FKfOZNRctfWjXUz207txT3WAaEIzBJhSoOQ5LpMAE0FCxDrUr1iOlGUjHc9AkhypV2DMcaW6H66rLoWkWdJGGGc5BtSnQyhwAVMy1+SruS14cHqfq6mp+frmzFyfsI/MlP9P4XKgBALQAspWIL4XTV79K6YjbHCxwJCx48TxwFmKMwWap4FmCTedw6Bw8ZYDiGShpEw6TwWYAqlDgytngzGqwYhLZqICZVSEtO0xdhWU6wJRZC82mSJDE/3YH5fubl8lKw/wua2K53fl985kBgDFGtxXu3Zcq6tcvvaaZPffMEJMiv82ebceZSKUQmhxHuXDBa7mg6RyawWA3ODQDsOmAzeSwmRz2NIcjo8Bl2KCaDsicCiEcUGxVyE0qiI7Gp2QpnSp/KsXw265BUVNXq65R5WNPPfXsDh+RMl/yc9oAFEgSkWKNTexcVccfqdzUgKd93QUrKAUrD0jLey6FvikA3/hvMBg7Abd0oFx6YNM5NB2wG4AmGBQLcGYYnBkOR4bBI5wot1XDjAOjz/th/Ool3HzR+aipqwLRycuW4ol032SSngmm2PVNVYNLoNxxbUuLbFmE70/xexoAgIgUxpgIJpPXweP57Y++c9j62JV16jlblkEIgqJMH05C4Kk9+/HQ71/B8fY4qu1nY3XjBqiVNqScEqPJXgRHj2HF5m0wVAHTzhGM9CN64ghWcR3XXbQRH3vvpWhoWDqdjzwzsIjwD8+0m9dffoF2iZZ9ssLm+aCPSLmWsWknQGcMgCIIYJBDqeT2lFJ2y++3t1k3fnS12rC6ahoIxfS4SAc6OvDg73bh5QMR2JVGrG6+ACk5jqHRI1i14f3oCXYiFurEWUsM3LDtLfjwtrfD6c5XzgkpoRTGyrtAPg/4txeOiobmeuXjKysCI4OhD0VW1fdsBeRirsVeD7HWwhX0oJ70HZww6VvfesUa7IkQEZFlTb8ktSxBsuTidCA4Snfc/xBd1NJKq664jWou/hdav62VbvjKdtr9cvu0vqYlpl2vl94cf+WPr1nfeHWQxsga7Rkc3ADMWvT51yEqKT85kU0+djCk0/fuareOto1KokLJi5BT9QFSShJCUGmZUDwZpwceeZZuu/fX1NHdO/VdygJoBWHlVI1A/j2e0+mWZ18z7+gYo1ES/V0DA1uA/AHu/4vwJTR1zx40kvd3piXdvv1V+aivU8jCFbcQksSMa3MpJc1WLyVEvhDilLYl/Q8MTdDHHzloPjA4SSFhBI4PjK0uKORvU/s/zRIS0Xv6dNP61TNDtP2edst/eMyikqKXIhil1SOWJci0rCmQisVSVkk7IqJgPC1+8NJx6+ZnXrP+nLIobOVeebWrf9XfVPhSKjIxFA5eHiP684Fxi+76VS/d+5NX6VD7qDR18xSNF81dSKK5ikL6xxP0X3u6rc8+00X3dYYoTETjqcSDP/3pT7XCvK+71PXMVEwjX6l12WWXWcf6R85asar+7EEj++mhE9lt3f5JJCMZLKuy0fpVHrmqoQwVFXZut9tK5yYpBBIZgwajGeYPpmTnREpJg2jzumVs41JHb3O5fUcqnDiyvHbJQ0D+xvr2MxDtzxgAADBzDR41kv/q0DzfOD6eKu8byijxqEAqZULqGajMAlM4JCPJFM5TlglyemCpQE2dByurVFy4sgpaNvV4jTN3A2M1SeDUytTXS2cUAGCKQbYDwLWMiZ6B0S3NK6tZGMZ7GJyXhLK6YeTkmw2BuljGgGa3Qxp6psKpuFQFh1ZUepqknn261u6+Z3hswrtiWc2uwrhT93lnmue/Gs300Va0cgYgm51cnTKz756k1LaIkfgMEa0bm4xeT0RsYGCgflqffNn9GVdUkf5qAxepFITFZmglfdjflcYXosKyyVqJuO9kHb/SWvJ/AP+a2n6D3qA36A16g96g6fR/lE5vtU7zgUgAAAAASUVORK5CYII="


# ============================================================
# FORMATAÇÃO EXCEL — helpers
# ============================================================

def _borda_excel(cor='B8CCE4'):
    s = Side(style='thin', color=cor)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill_excel(hex_color):
    return PatternFill(fill_type='solid', fgColor=hex_color, bgColor=hex_color)

def _fmt_cabecalho(cell, bg='1F4E79', fg='FFFFFF', size=11):
    cell.fill      = _fill_excel(bg)
    cell.font      = Font(bold=True, color=fg, name='Calibri', size=size)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border    = _borda_excel()

def _fmt_celula(cell, bg='FFFFFF', bold=False, center=True, fmt=None):
    cell.fill      = _fill_excel(bg)
    cell.font      = Font(bold=bold, name='Calibri', size=10)
    cell.alignment = Alignment(
        horizontal='center' if center else 'left',
        vertical='center'
    )
    cell.border    = _borda_excel()
    if fmt:
        cell.number_format = fmt

def _fmt_total(cell, bg='1F4E79', fg='FFFFFF'):
    cell.fill      = _fill_excel(bg)
    cell.font      = Font(bold=True, color=fg, name='Calibri', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = _borda_excel()

def _autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except (TypeError, ValueError):
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)

def _formatar_aba_generica(ws, cor_header='1F4E79'):
    """Formata qualquer aba com cabeçalho colorido, linhas alternadas e fórmulas de soma."""
    for cell in ws[1]:
        _fmt_cabecalho(cell, bg=cor_header)
    ws.row_dimensions[1].height = 26

    max_row    = ws.max_row
    max_col    = ws.max_column
    cores      = ['F7F9FC', 'EEF2FA']

    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        is_total_row = False
        # Detecta linha de total (primeira célula = 'TOTAL GERAL' ou marcador '__SOMA__')
        first_val = str(row[0].value) if row[0].value is not None else ''
        if first_val in ('TOTAL GERAL', 'TOTAL MÊS'):
            is_total_row = True

        for cell in row:
            v = str(cell.value) if cell.value is not None else ''

            if is_total_row:
                if v == '__SOMA__':
                    # Substitui marcador por fórmula de soma
                    col_letter = get_column_letter(cell.column)
                    cell.value = f'=SUM({col_letter}2:{col_letter}{max_row - 1})'
                    _fmt_total(cell)
                    cell.number_format = 'R$ #,##0.00'
                else:
                    _fmt_total(cell)
            else:
                bg     = cores[i % 2]
                num_fmt = None
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    header       = ws.cell(row=1, column=cell.column).value or ''
                    header_lower = str(header).lower()
                    nao_monetario = any(x in header_lower for x in [
                        'volume', 'qtd', 'quantidade', 'mes', 'mês', 'ano',
                        'hora', 'h ', 'dias', 'eventos', 'saída', 'saídas',
                        'utilização', 'ativo', 'ocioso', '%',
                    ])
                    if not nao_monetario:
                        num_fmt = 'R$ #,##0.00'
                _fmt_celula(cell, bg=bg, fmt=num_fmt)
        ws.row_dimensions[i + 2].height = 18
    _autofit(ws)



# ============================================================
# MÓDULO 1 — PEDIDOS E RECEBIMENTOS
# ============================================================
NOME_SAIDA_PEDIDOS     = 'planilha_de_resultados_detalhada.xlsx'
COL_DEPOSITANTE        = 'Depositante'
COL_DATA_INICIO        = 'Data'
COL_DATA_FIM           = 'Expedição'
COL_ORDEM_SAIDA        = 'Ordem de Saída'

def _detectar_aba(caminho, colunas_obrigatorias, log=None):
    """
    Varre as abas do Excel e retorna a primeira que contém todas as colunas_obrigatorias.
    Retorna (nome_aba, df) ou (None, None) se não encontrar.
    """
    try:
        xl = pd.ExcelFile(caminho)
        for aba in xl.sheet_names:
            try:
                df = xl.parse(aba, nrows=3)
                cols = [str(c).strip() for c in df.columns]
                if all(col in cols for col in colunas_obrigatorias):
                    if log:
                        log(f"🔍 Aba detectada automaticamente: '{aba}'\n")
                    return aba, xl.parse(aba)
            except Exception:
                continue
    except Exception as e:
        if log:
            log(f"❌ Erro ao abrir arquivo: {e}\n")
    return None, None

def _classificar_prazo(d):
    if d <= 0:   return 'D+0'
    elif d == 1: return 'D+1'
    else:        return 'Excedido D+1'

def processar_pedidos(caminho, log, _saida_override=None):
    try:
        log(f"📂 Carregando: {os.path.basename(caminho)}\n")
        nome_aba, df = _detectar_aba(caminho,
            [COL_DATA_INICIO, COL_DATA_FIM, COL_DEPOSITANTE, COL_ORDEM_SAIDA], log)
        if df is None:
            log(f"❌ Nenhuma aba encontrada com as colunas esperadas "
                f"({COL_DATA_INICIO}, {COL_DATA_FIM}, {COL_DEPOSITANTE}, {COL_ORDEM_SAIDA}).\n")
            return False
        log(f"✅ {len(df)} linhas encontradas.\n")
    except Exception as e:
        log(f"❌ Erro ao carregar: {e}\n"); return False

    df[COL_DATA_INICIO] = pd.to_datetime(df[COL_DATA_INICIO], errors='coerce')
    df[COL_DATA_FIM]    = pd.to_datetime(df[COL_DATA_FIM],    errors='coerce')
    df.dropna(subset=[COL_DATA_INICIO, COL_DATA_FIM, COL_DEPOSITANTE, COL_ORDEM_SAIDA], inplace=True)
    log(f"🧹 {len(df)} linhas válidas após limpeza.\n")

    anos = list(set(df[COL_DATA_INICIO].dt.year.tolist() + df[COL_DATA_FIM].dt.year.tolist()))
    feriados = list(holidays.BR(years=anos).keys())
    log(f"📅 Feriados calculados para: {anos}\n")

    ini = df[COL_DATA_INICIO].dt.normalize().values.astype('datetime64[D]')
    fim = df[COL_DATA_FIM].dt.normalize().values.astype('datetime64[D]')
    df['Diferenca_Dias'] = np.busday_count(ini, fim, holidays=feriados)
    df['Status_Prazo']   = df['Diferenca_Dias'].apply(_classificar_prazo)

    contagem = df.groupby([COL_DEPOSITANTE, 'Status_Prazo']).size().reset_index(name='Total')
    resumo = contagem.pivot(index=COL_DEPOSITANTE, columns='Status_Prazo', values='Total').fillna(0).astype(int)
    resumo = resumo.reindex(columns=['D+0', 'D+1', 'Excedido D+1'], fill_value=0)
    resumo['Total Geral'] = resumo.sum(axis=1)
    resumo['SLA %'] = ((resumo['D+0'] + resumo['D+1']) / resumo['Total Geral'] * 100).round(4)

    df_exc = df[df['Status_Prazo'] == 'Excedido D+1'].copy()
    df_det = df_exc[[COL_DEPOSITANTE, COL_ORDEM_SAIDA, COL_DATA_INICIO, COL_DATA_FIM, 'Diferenca_Dias']]\
               .sort_values([COL_DEPOSITANTE, COL_DATA_INICIO])
    df_det = df_det.rename(columns={'Diferenca_Dias': 'Dias_de_Atraso'})
    df_det[COL_DATA_INICIO] = df_det[COL_DATA_INICIO].dt.strftime('%d/%m/%Y')
    df_det[COL_DATA_FIM]    = df_det[COL_DATA_FIM].dt.strftime('%d/%m/%Y')

    pasta = os.path.dirname(caminho)
    saida = _saida_override
    if not saida:
        log("❌ Caminho de saída não definido. Verifique as Configurações.\n")
        return False
    log(f"💾 Gerando Excel em: {os.path.basename(saida)}...\n")

    with pd.ExcelWriter(saida, engine='openpyxl') as writer:
        resumo.to_excel(writer, sheet_name='Resumo Por Depositante')
        df_det.to_excel(writer, sheet_name='Ordens Excedidas D+1', index=False)

    # Formatação
    wb = load_workbook(saida)
    _formatar_aba_pedidos_resumo(wb, resumo)
    _formatar_aba_pedidos_det(wb)
    wb.save(saida)

    log(f"✅ Salvo em: {saida}\n")
    log(f"📋 Ordens excedidas: {len(df_det)}\n")
    log("\n🎉 Concluído!\n")
    return True

def _formatar_aba_pedidos_resumo(wb, resumo_df):
    ws = wb['Resumo Por Depositante']
    VERDE_ESC='375623'; AMAR_ESC='7F6000'; VERM_ESC='9C0006'
    AZUL_ESC='1F4E79'; AZUL_MED='2F4F74'
    VERDE_SUV='C6EFCE'; AMAR_SUV='FFEB9C'; VERM_SUV='FFC7CE'; CINZA='D6DCE4'
    LINHA_PAR='F2F7FB'

    cores_h = {'A': AZUL_ESC,'B': VERDE_ESC,'C': AMAR_ESC,'D': VERM_ESC,'E': AZUL_MED,'F': AZUL_ESC}
    for c, cor in cores_h.items():
        _fmt_cabecalho(ws[f'{c}1'], bg=cor)
    ws.row_dimensions[1].height = 30

    n = resumo_df.shape[0]
    for i in range(2, n + 2):
        fundo = LINHA_PAR if i % 2 == 0 else 'FFFFFF'
        ws[f'A{i}'].fill = _fill_excel(fundo)
        ws[f'A{i}'].font = Font(bold=True, name='Calibri', size=10)
        ws[f'A{i}'].alignment = Alignment(vertical='center')
        ws[f'A{i}'].border = _borda_excel()
        for c, cor in [('B', VERDE_SUV),('C', AMAR_SUV),('D', VERM_SUV),('E', CINZA)]:
            _fmt_celula(ws[f'{c}{i}'], bg=cor, fmt='#,##0')
        cf = ws[f'F{i}']
        v  = cf.value or 0
        cs, ct = (VERDE_SUV, VERDE_ESC) if v >= 90 else (AMAR_SUV, AMAR_ESC) if v >= 70 else (VERM_SUV, VERM_ESC)
        cf.value = v / 100
        cf.fill = _fill_excel(cs)
        cf.font = Font(bold=True, name='Calibri', size=10, color=ct)
        cf.alignment = Alignment(horizontal='center', vertical='center')
        cf.border = _borda_excel()
        cf.number_format = '0.00%'

    tot = n + 2
    ws[f'A{tot}'] = 'TOTAL GERAL'
    _fmt_total(ws[f'A{tot}'])
    for c in ['B','C','D','E']:
        ws[f'{c}{tot}'] = f'=SUM({c}2:{c}{n+1})'
        _fmt_total(ws[f'{c}{tot}'])
        ws[f'{c}{tot}'].number_format = '#,##0'
    ws[f'F{tot}'] = f'=(B{tot}+C{tot})/E{tot}'
    _fmt_total(ws[f'F{tot}'])
    ws[f'F{tot}'].number_format = '0.00%'
    ws.row_dimensions[tot].height = 22
    _autofit(ws)

def _formatar_aba_pedidos_det(wb):
    ws = wb['Ordens Excedidas D+1']
    for cell in ws[1]:
        _fmt_cabecalho(cell, bg='9C0006')
    ws.row_dimensions[1].height = 28
    for i, row in enumerate(ws.iter_rows(min_row=2), start=0):
        bg = 'FFC7CE' if i % 2 == 0 else 'FFF2F2'
        for cell in row:
            _fmt_celula(cell, bg=bg)
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell.font = Font(bold=True, color='9C0006', name='Calibri', size=10)
    _autofit(ws)

# ============================================================
# MÓDULO 2 — FRETES
# ============================================================
REMETENTES_IGNORAR = [
    'BAIA 4 LOGISTICA E TRANSPORTES LTDA',
    'FUNCIONAL HEALTH TECH SOLUCOES EM SAUDE LTDA'
]

def _calcular_custo_insumo(pv):
    if 0 < pv <= 0.5:    return 0.35
    elif pv <= 1:        return 2.55
    elif pv <= 2:        return 3.40
    elif pv <= 3:        return 2.18
    elif pv < 4:         return 4.49
    elif pv <= 4.5:      return 4.43
    elif pv <= 13:       return 9.40
    elif pv <= 18.5:     return 12.50
    elif pv > 18.5:      return 16.45
    else:                return 0.0

def _processar_aba_frete(nome_arquivo, nome_aba, col_remetente, col_valor, nome_col_soma, extras=None):
    df = pd.read_excel(nome_arquivo, sheet_name=nome_aba)
    df.columns = df.columns.str.strip().str.lower()

    cols = [col_valor] + (extras or [])
    for col in cols:
        if col not in df.columns:
            df[col] = 0.0
        df[col] = df[col].astype(str).str.strip().str.replace(',', '.', regex=False)
        df[col] = df[col].str.findall(r'-?\d+\.?\d*').str[0]
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)

    agg = {col_valor: 'sum'}
    if extras:
        for c in extras: agg[c] = 'sum'

    agrupado = df.dropna(subset=[col_remetente]).groupby(col_remetente).agg(agg)
    agrupado.rename(columns={col_valor: nome_col_soma}, inplace=True)
    agrupado.reset_index(inplace=True)
    agrupado.rename(columns={agrupado.columns[0]: 'Remetente'}, inplace=True)
    return agrupado

def processar_fretes(caminho, log, _saida_override=None):
    try:
        log(f"📂 Carregando: {os.path.basename(caminho)}\n")

        # Detecta aba de embarques — contém a coluna 'valor frete'
        nome_aba_emb, _ = _detectar_aba(caminho, ['valor frete'], log)
        if nome_aba_emb is None:
            # Tenta busca case-insensitive
            xl = pd.ExcelFile(caminho)
            for aba in xl.sheet_names:
                try:
                    df_test = xl.parse(aba, nrows=3)
                    cols_lower = [str(c).strip().lower() for c in df_test.columns]
                    if 'valor frete' in cols_lower:
                        nome_aba_emb = aba
                        log(f"🔍 Aba detectada automaticamente: '{aba}'\n")
                        break
                except Exception:
                    continue
        if nome_aba_emb is None:
            log("❌ Nenhuma aba com coluna 'valor frete' encontrada.\n"); return False

        # --- Embarques: calcula peso x volume automaticamente ---
        log(f"🔄 Processando aba '{nome_aba_emb}'...\n")
        df_emb_raw = pd.read_excel(caminho, sheet_name=nome_aba_emb)
        df_emb_raw.columns = df_emb_raw.columns.str.strip()

        # Calcula peso x volume = Peso Taxado ÷ Volume
        if 'Peso Taxado' in df_emb_raw.columns and 'Volume' in df_emb_raw.columns:
            df_emb_raw['peso x volume'] = df_emb_raw.apply(
                lambda r: r['Peso Taxado'] / r['Volume'] if r['Volume'] and r['Volume'] != 0 else 0,
                axis=1
            )
            log("✅ Coluna 'peso x volume' calculada automaticamente (Peso Taxado ÷ Volume).\n")
        else:
            cols_disp = list(df_emb_raw.columns)
            log(f"⚠️ Colunas 'Peso Taxado' ou 'Volume' não encontradas.\n   Colunas disponíveis: {cols_disp}\n")
            df_emb_raw['peso x volume'] = 0

        # Custo insumos
        df_custo = df_emb_raw[~df_emb_raw['Remetente'].isin(REMETENTES_IGNORAR)].copy()
        df_custo['custo_unitario'] = df_custo['peso x volume'].apply(_calcular_custo_insumo)
        df_custo['custo_total']    = df_custo['custo_unitario'] * df_custo['Volume']
        custo_agrup = df_custo.groupby('Remetente')['custo_total'].sum().reset_index()
        custo_agrup.rename(columns={'custo_total': 'Custo Insumos (Caixas)', 'Remetente': 'Remetente'}, inplace=True)

        # Embarques normais (sem 'custo insumos (caixas)' pois não existe no Excel original)
        df_emb = _processar_aba_frete(
            caminho, nome_aba_emb,
            col_remetente='remetente',
            col_valor='valor frete',
            nome_col_soma='Valor Frete Total',
            extras=['frete peso', 'itr', 'outros']
        )
        log("✅ Embarques processados.\n")

        # Injeta custo insumos calculado na tabela de embarques
        custo_merge = custo_agrup.rename(columns={'Custo Insumos (Caixas)': 'Custo Insumos (Caixas)_calc'})
        custo_merge['Remetente'] = custo_merge['Remetente'].str.strip()
        df_emb['Remetente_key']  = df_emb['Remetente'].str.strip()
        custo_merge = custo_merge.rename(columns={'Remetente': 'Remetente_key'})
        df_emb = pd.merge(df_emb, custo_merge, on='Remetente_key', how='left')
        df_emb.drop(columns=['Remetente_key'], inplace=True)
        df_emb.rename(columns={'Custo Insumos (Caixas)_calc': 'Custo Insumos (Caixas)'}, inplace=True)
        df_emb['Custo Insumos (Caixas)'] = df_emb['Custo Insumos (Caixas)'].fillna(0)
        # Reordena colunas: Remetente, Valor Frete Total, frete peso, Custo Insumos (Caixas), itr, outros
        cols_emb = ['Remetente', 'Valor Frete Total', 'frete peso', 'Custo Insumos (Caixas)', 'itr', 'outros']
        cols_emb = [c for c in cols_emb if c in df_emb.columns]
        df_emb = df_emb[cols_emb]
        log("✅ Custo Insumos (Caixas) injetado na aba Embarques.\n")

        log("🔄 Processando aba 'RESCOM'...\n")
        df_rescom = _processar_aba_frete(caminho, 'RESCOM', 'remetente', 'valor', 'Valor RESCOM')
        log("✅ RESCOM processado.\n")

        log("🔄 Processando aba 'PORTADORES'...\n")
        df_port = _processar_aba_frete(caminho, 'PORTADORES', 'remetente', 'valor parceiro', 'Valor PORTADORES')
        log("✅ PORTADORES processado.\n")

        # Consolidado — usa df_emb que já tem Custo Insumos (Caixas) correto
        df_cons = df_emb.copy()
        for df_tmp in [df_rescom, df_port]:
            df_cons = pd.merge(df_cons, df_tmp, on='Remetente', how='outer')
        df_cons = df_cons.fillna(0)

        # Linha de total
        def _add_total(df):
            """Adiciona linha de TOTAL GERAL com marcador para fórmulas Excel depois."""
            tot = {col: 'TOTAL GERAL' if col == 'Remetente' else '__SOMA__' for col in df.columns}
            return pd.concat([df, pd.DataFrame([tot])], ignore_index=True)

        df_emb_f   = _add_total(df_emb)
        df_res_f   = _add_total(df_rescom)
        df_port_f  = _add_total(df_port)
        df_custo_f = _add_total(custo_agrup)
        df_cons_f  = _add_total(df_cons)

        saida = _saida_override
        if not saida:
            log("❌ Caminho de saída não definido. Verifique as Configurações.\n")
            return False
        log(f"💾 Gerando Excel em: {os.path.basename(saida)}...\n")

        with pd.ExcelWriter(saida, engine='openpyxl') as writer:
            df_emb_f.to_excel(writer,   sheet_name='Embarques',       index=False)
            df_res_f.to_excel(writer,   sheet_name='RESCOM',           index=False)
            df_port_f.to_excel(writer,  sheet_name='PORTADORES',       index=False)
            df_custo_f.to_excel(writer, sheet_name='Custo Insumos',    index=False)
            df_cons_f.to_excel(writer,  sheet_name='Consolidado',      index=False)

        wb = load_workbook(saida)
        cores = {'Embarques': '1F4E79', 'RESCOM': '375623',
                 'PORTADORES': '7F3F00', 'Custo Insumos': '4B0082', 'Consolidado': '0D3B6E'}
        for aba, cor in cores.items():
            if aba in wb.sheetnames:
                _formatar_aba_generica(wb[aba], cor_header=cor)
        wb.save(saida)

        log(f"✅ Salvo em: {saida}\n")
        log("\n🎉 Concluído!\n")
        return True

    except Exception as e:
        log(f"❌ Erro: {e}\n")
        log(traceback.format_exc() + '\n')
        return False

# ============================================================
# MÓDULO 3 — ARMAZENAGEM
# ============================================================

def processar_armazenagem(caminho, mes_filtro, log, _saida_override=None):
    try:
        log(f"📂 Carregando: {os.path.basename(caminho)}\n")
        df = pd.read_excel(caminho)
        df['Emissão'] = pd.to_datetime(df['Emissão'], format='%d/%m/%Y', errors='coerce')
        df.dropna(subset=['Emissão'], inplace=True)
        log(f"✅ {len(df)} linhas carregadas.\n")

        meses = df['Emissão'].dt.to_period('M').unique()

        if mes_filtro:
            try:
                periodo = pd.Period(mes_filtro, freq='M')
                meses = [m for m in meses if m == periodo]
                if not meses:
                    log(f"⚠️ Nenhum dado encontrado para o período '{mes_filtro}'.\n")
                    log(f"   Períodos disponíveis: {[str(m) for m in df['Emissão'].dt.to_period('M').unique()]}\n")
                    return False
                log(f"🔍 Filtrando período: {mes_filtro}\n")
            except Exception:
                log(f"⚠️ Formato de período inválido. Use YYYY-MM (ex: 2026-01)\n")
                return False

        # Consolida clientes que aparecem em meses diferentes (pagamento fracionado)
        # → soma todos os valores do cliente e atribui ao mês mais recente
        mes_mais_recente = df['Emissão'].dt.to_period('M').max()

        # Identifica clientes que aparecem em mais de um mês
        clientes_fracionados = set(
            df.groupby('Cliente')['Emissão'].apply(
                lambda x: x.dt.to_period('M').nunique()
            ).pipe(lambda s: s[s > 1].index)
        )

        # Atribui mês consolidado: fracionados vão para o mês mais recente, outros ficam no próprio mês
        df['Mês Consolidado'] = df.apply(
            lambda r: mes_mais_recente
            if r['Cliente'] in clientes_fracionados
            else pd.Period(r['Emissão'], freq='M'),
            axis=1
        )

        df_resultados = pd.DataFrame(columns=['Cliente', 'Soma Armazenagem'])

        for mes in meses:
            # Clientes cujo mês consolidado é este mês
            df_mes = df[df['Mês Consolidado'] == mes]
            if df_mes.empty:
                continue
            soma = df_mes.groupby('Cliente')['Valor Principal'].sum().reset_index()
            soma.columns = ['Cliente', 'Soma Armazenagem']
            total = pd.DataFrame([{
                'Cliente': 'TOTAL MÊS',
                'Soma Armazenagem': '__SOMA__'
            }])
            df_resultados = pd.concat([df_resultados, soma, total], ignore_index=True)
            log(f"📅 {mes}: {len(soma)} clientes | Total: R$ {df_mes['Valor Principal'].sum():,.2f}\n")

        saida = _saida_override
        if not saida:
            log("❌ Caminho de saída não definido. Verifique as Configurações.\n")
            return False
        log(f"💾 Gerando Excel em: {os.path.basename(saida)}...\n")

        with pd.ExcelWriter(saida, engine='openpyxl') as writer:
            df_resultados.to_excel(writer, sheet_name='Armazenagem', index=False)

        wb = load_workbook(saida)
        _formatar_aba_generica(wb['Armazenagem'], cor_header='0D4F3C')
        wb.save(saida)

        log(f"✅ Salvo em: {saida}\n")
        log("\n🎉 Concluído!\n")
        return True

    except Exception as e:
        log(f"❌ Erro: {e}\n")
        return False

# ============================================================
# CONFIGURAÇÃO DE PASTAS E ESTADO PERSISTENTE
# ============================================================
import json, subprocess, shutil

PASTA_PEDIDOS     = r'Z:\GUSTAVO\App\Pedidos'
PASTA_FRETES      = r'Z:\GUSTAVO\App\Fretes'
PASTA_ARMAZENAGEM = r'Z:\GUSTAVO\App\Armazenagem'

# Arquivo JSON para guardar estado do app (últimas extrações, dia fixo, etc.)
ESTADO_PATH = os.path.join(os.path.expanduser('~'), '.central_relatorios_state.json')

PASTA_ESTOQUE        = r'Z:\GUSTAVO\App\Estoque'
PASTA_PRODUTIVIDADE  = r'Z:\GUSTAVO\App\Produtividade'
PASTA_CAP_OPERACIONAL  = r'Z:\GUSTAVO\App\CapacidadeOperacional'
PASTA_RECEBIMENTOS     = r'Z:\GUSTAVO\App\Recebimentos'
PASTA_FINANCEIRO       = r'Z:\GUSTAVO\App\Financeiro'
COR_ACCENT8            = '#0891b2'  # azul-petróleo — recebimentos
COR_PAGAR              = '#dc2626'  # vermelho — contas a pagar
COR_RECEBER            = '#16a34a'  # verde — contas a receber
COR_FAT_ARM            = '#7c3aed'  # violeta — faturamento armazenagem
PASTA_CONSOLIDADOS   = r'Z:\GUSTAVO\App\Consolidados'
DE_PARA_PATH        = os.path.join(os.path.expanduser('~'), 'central_relatorios_de_para.json')
DB_ESTOQUE_PATH     = os.path.join(os.path.expanduser('~'), 'central_relatorios_estoque_db.json')
DB_FAMILIAS_PATH    = os.path.join(os.path.expanduser('~'), 'central_relatorios_familias_db.json')
DB_PRECOS_ARM_PATH  = os.path.join(os.path.expanduser('~'), 'central_relatorios_precos_arm.json')
COR_ACCENT4         = '#f59e0b'  # âmbar — estoque
COR_ACCENT5         = '#06b6d4'  # cyan  — produtividade
COR_ACCENT7         = '#e11d48'  # rosa  — capacidade operacional

RELATORIOS_CONFIG = {
    'Pedidos e Recebimentos': {
        'icone': '📦', 'cor': '#4f8ef7',
        'arquivo_base': 'planilha_de_resultados_detalhada',
        'pasta': PASTA_PEDIDOS,
    },
    'Fretes': {
        'icone': '🚚', 'cor': '#7c3aed',
        'arquivo_base': 'relatorio_fretes',
        'pasta': PASTA_FRETES,
    },
    'Armazenagem': {
        'icone': '🏭', 'cor': '#10b981',
        'arquivo_base': 'relatorio_armazenagem',
        'pasta': PASTA_ARMAZENAGEM,
    },
    'Estoque': {
        'icone': '📋', 'cor': '#f59e0b',
        'arquivo_base': 'relatorio_estoque',
        'pasta': PASTA_ESTOQUE,
    },
    'Produtividade': {
        'icone': '👥', 'cor': '#06b6d4',
        'arquivo_base': 'relatorio_produtividade',
        'pasta': PASTA_PRODUTIVIDADE,
    },
    'Capacidade Operacional': {
        'icone': '⚙️', 'cor': '#e11d48',
        'arquivo_base': 'relatorio_cap_operacional',
        'pasta': PASTA_CAP_OPERACIONAL,
    },
    'Recebimentos e Devoluções': {
        'icone': '📥', 'cor': '#0891b2',
        'arquivo_base': 'relatorio_recebimentos',
        'pasta': PASTA_RECEBIMENTOS,
    },
    'Financeiro': {
        'icone': '💰', 'cor': '#16a34a',
        'arquivo_base': 'relatorio_financeiro',
        'pasta': PASTA_FINANCEIRO,
    },
    'Dashboard': {
        'icone': '📈', 'cor': '#8b5cf6',
        'arquivo_base': 'dashboard',
        'pasta': r'Z:\GUSTAVO\App\Dashboards',
    },
}

def _carregar_estado():
    try:
        with open(ESTADO_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {'dia_extracao': 1, 'extrações': {}}

def _salvar_estado(estado):
    try:
        with open(ESTADO_PATH, 'w', encoding='utf-8') as f:
            json.dump(estado, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

def _registrar_extracao(nome_relatorio):
    estado = _carregar_estado()
    estado.setdefault('extrações', {})[nome_relatorio] = date.today().strftime('%d/%m/%Y')
    _salvar_estado(estado)

def _proxima_extracao(dia_fixo):
    hoje = date.today()
    try:
        proximo = date(hoje.year, hoje.month, dia_fixo)
        if proximo <= hoje:
            # Avança para o próximo mês
            if hoje.month == 12:
                proximo = date(hoje.year + 1, 1, dia_fixo)
            else:
                ultimo = monthrange(hoje.year, hoje.month + 1)[1]
                d = min(dia_fixo, ultimo)
                proximo = date(hoje.year, hoje.month + 1, d)
    except ValueError:
        ultimo = monthrange(hoje.year, hoje.month)[1]
        proximo = date(hoje.year, hoje.month, ultimo)
    return proximo.strftime('%d/%m/%Y')

def _pasta_inicial(pasta):
    return pasta if os.path.isdir(pasta) else ''

def _abrir_excel(caminho):
    try:
        os.startfile(caminho)
    except Exception as e:
        messagebox.showerror('Erro', f'Não foi possível abrir o arquivo:\n{e}')

def _notificar_windows(titulo, mensagem, icone_path=None):
    """Envia notificação nativa do Windows via win10toast ou powershell como fallback."""
    try:
        from win10toast import ToastNotifier
        toaster = ToastNotifier()
        toaster.show_toast(
            titulo, mensagem,
            icon_path=icone_path,
            duration=6,
            threaded=True
        )
        return
    except ImportError:
        pass
    # Fallback: PowerShell notification
    try:
        ps_script = f"""
Add-Type -AssemblyName System.Windows.Forms
$notify = New-Object System.Windows.Forms.NotifyIcon
$notify.Icon = [System.Drawing.SystemIcons]::Information
$notify.Visible = $true
$notify.ShowBalloonTip(6000, '{titulo}', '{mensagem}', [System.Windows.Forms.ToolTipIcon]::Info)
Start-Sleep -Seconds 7
$notify.Dispose()
"""
        subprocess.Popen(
            ['powershell', '-WindowStyle', 'Hidden', '-Command', ps_script],
            creationflags=subprocess.CREATE_NO_WINDOW if hasattr(subprocess, 'CREATE_NO_WINDOW') else 0
        )
    except Exception:
        pass

def _exportar_pdf(caminho_xlsx, callback_status):
    """Converte xlsx para PDF via Excel COM (Windows)."""
    pasta   = os.path.dirname(caminho_xlsx)
    nome    = os.path.splitext(os.path.basename(caminho_xlsx))[0]
    pdf_out = os.path.join(pasta, nome + '.pdf')

    # Tenta LibreOffice primeiro
    for lo in [r'C:\Program Files\LibreOffice\program\soffice.exe',
               r'C:\Program Files (x86)\LibreOffice\program\soffice.exe']:
        if os.path.exists(lo):
            try:
                subprocess.run([lo, '--headless', '--convert-to', 'pdf',
                                '--outdir', pasta, caminho_xlsx],
                               check=True, capture_output=True, timeout=30)
                callback_status(f'✅ PDF gerado com sucesso!')
                os.startfile(pdf_out)
                return
            except Exception as e:
                callback_status(f'⚠️ LibreOffice falhou: {e}')

    # Tenta via Excel COM
    try:
        import win32com.client
        xl = win32com.client.DispatchEx('Excel.Application')
        xl.DisplayAlerts = False
        xl.EnableEvents  = False
        wb = xl.Workbooks.Open(os.path.abspath(caminho_xlsx))
        # Configura todas as abas para paisagem e ajuste automático
        for ws in wb.Worksheets:
            ws.PageSetup.Orientation = 2          # 2 = paisagem
            ws.PageSetup.Zoom = False
            ws.PageSetup.FitToPagesWide = 1       # cabe em 1 página de largura
            ws.PageSetup.FitToPagesTall = False   # altura livre
        wb.ExportAsFixedFormat(0, os.path.abspath(pdf_out))
        wb.Close(SaveChanges=False)
        xl.Quit()
        del xl
        callback_status(f'✅ PDF gerado com sucesso!')
        os.startfile(pdf_out)
    except ImportError:
        callback_status('⚠️ pywin32 não instalado. Execute: pip install pywin32')
    except Exception as e:
        callback_status(f'❌ Erro ao gerar PDF: {e}')

# ============================================================
# PALETA E ESTILOS
# ============================================================
COR_BG        = '#0f1117'
COR_SIDEBAR   = '#13161f'
COR_CARD      = '#1a1d27'
COR_CARD2     = '#22263a'
COR_ACCENT    = '#4f8ef7'
COR_ACCENT2   = '#7c3aed'
COR_ACCENT3   = '#10b981'

COR_TEXTO     = '#e2e8f0'
COR_TEXTO_SUB = '#8892a4'
COR_BORDA     = '#2d3148'
COR_SEL       = '#1e2235'
COR_HOVER     = '#2a2f45'
COR_ERRO      = '#ef4444'
COR_AVISO     = '#f59e0b'

FONTE_TITULO  = ('Segoe UI', 20, 'bold')
FONTE_LABEL   = ('Segoe UI', 9)
FONTE_BTN     = ('Segoe UI', 10, 'bold')
FONTE_LOG     = ('Consolas', 9)
FONTE_ENTRY   = ('Segoe UI', 10)
FONTE_SUBTITULO = ('Segoe UI', 10)

def _escurecer(hex_color):
    r = max(0, int(hex_color[1:3], 16) - 25)
    g = max(0, int(hex_color[3:5], 16) - 25)
    b = max(0, int(hex_color[5:7], 16) - 25)
    return f'#{r:02x}{g:02x}{b:02x}'

def estilo_entry(parent, width=46, **kw):
    return tk.Entry(parent, font=FONTE_ENTRY, width=width,
                    bg=COR_CARD2, fg=COR_TEXTO, insertbackground=COR_TEXTO,
                    relief='flat', bd=0, highlightthickness=1,
                    highlightbackground=COR_BORDA, highlightcolor=COR_ACCENT, **kw)

def estilo_btn(parent, text, color, command, width=20):
    btn = tk.Button(parent, text=text, font=FONTE_BTN,
                    bg=color, fg='white', relief='flat',
                    padx=14, pady=8, cursor='hand2',
                    activebackground=_escurecer(color), activeforeground='white',
                    width=width, command=command)
    btn.bind('<Enter>', lambda e: btn.config(bg=_escurecer(color)))
    btn.bind('<Leave>', lambda e: btn.config(bg=color))
    return btn

def criar_log(parent):
    frame = tk.Frame(parent, bg=COR_BG)
    txt   = tk.Text(frame, font=FONTE_LOG, bg='#090b10', fg='#a9b7c6',
                    relief='flat', state='disabled', wrap='word', padx=10, pady=8)
    sb    = ttk.Scrollbar(frame, command=txt.yview)
    txt.configure(yscrollcommand=sb.set)
    txt.pack(side='left', fill='both', expand=True)
    sb.pack(side='right', fill='y')
    return frame, txt

def log_write(w, msg):
    w.configure(state='normal')
    w.insert(tk.END, msg)
    w.see(tk.END)
    w.configure(state='disabled')

def _dropdown_mes(parent, label='Mês de referência:', cor_label=None):
    """
    Dropdown dark mode com os 12 meses anteriores ao mês atual em ordem cronológica.
    Retorna StringVar com formato 'MM-AAAA'. Padrão = mês imediatamente anterior.
    """
    _MESES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                 'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']

    hoje = date.today()
    # Gera os 12 meses anteriores ao mês atual em ordem decrescente (mais recente → mais antigo)
    opcoes = []
    for i in range(0, 12):               # 0 → 11  (do mais recente ao mais antigo)
        mes_ref = hoje.month - 1 - i
        ano_ref = hoje.year
        while mes_ref <= 0:
            mes_ref += 12
            ano_ref -= 1
        opcoes.append((f'{mes_ref:02d}-{ano_ref}',
                       f'{_MESES_PT[mes_ref-1]} de {ano_ref}'))

    # Padrão = mês anterior = primeira opção (mais recente)
    var = tk.StringVar(value=opcoes[0][0])

    frm = tk.Frame(parent, bg=COR_BG)
    frm.pack(anchor='w', padx=32, pady=(4, 10))

    tk.Label(frm, text=label,
             font=('Segoe UI', 9),
             bg=COR_BG,
             fg=cor_label or COR_TEXTO_SUB).pack(side='left', padx=(0, 10))

    # Frame que simula o combobox no estilo dark
    display_opcoes = [f'{disp}  ({cod})' for cod, disp in opcoes]
    display_var    = tk.StringVar(value=display_opcoes[0])

    # Aplica estilo dark via ttk.Style
    style_name = 'DarkCombo.TCombobox'
    try:
        s = ttk.Style()
        s.theme_use('default')
        s.configure(style_name,
                    fieldbackground=COR_CARD2,
                    background=COR_CARD2,
                    foreground=COR_TEXTO,
                    selectbackground=COR_SEL,
                    selectforeground=COR_TEXTO,
                    arrowcolor=COR_TEXTO_SUB,
                    bordercolor=COR_BORDA,
                    lightcolor=COR_BORDA,
                    darkcolor=COR_BORDA,
                    relief='flat')
        s.map(style_name,
              fieldbackground=[('readonly', COR_CARD2)],
              foreground=[('readonly', COR_TEXTO)],
              background=[('active', COR_SEL), ('pressed', COR_SEL)])
    except Exception:
        style_name = 'TCombobox'

    combo = ttk.Combobox(frm, textvariable=display_var,
                         values=display_opcoes,
                         state='readonly', width=30,
                         font=('Segoe UI', 9),
                         style=style_name)
    combo.pack(side='left', ipady=4)

    # Sincroniza display_var → var (MM-AAAA)
    def _on_change(e=None):
        idx = combo.current()
        var.set(opcoes[idx][0] if idx >= 0 else opcoes[0][0])
    combo.bind('<<ComboboxSelected>>', _on_change)

    return var

def separador(parent, pady=8):
    tk.Frame(parent, bg=COR_BORDA, height=1).pack(fill='x', pady=pady)

# ============================================================
# HELPERS DE CAMINHO E MÊS DE REFERÊNCIA
# ============================================================

def _mes_ano_referencia():
    """Retorna (mes, ano) do mês anterior — mês de fechamento."""
    hoje = date.today()
    if hoje.month == 1:
        return 12, hoje.year - 1
    return hoje.month - 1, hoje.year

# ============================================================
# BANCO DE DADOS INTERNO DE ESTOQUE
# ============================================================

def _carregar_db_estoque():
    """Carrega o DB interno: {cliente: {codigo: {desc, saldo, reservado, bloqueado, atualizado}}}"""
    try:
        with open(DB_ESTOQUE_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _salvar_db_estoque(db):
    try:
        with open(DB_ESTOQUE_PATH, 'w', encoding='utf-8') as f:
            json.dump(db, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

def _db_estoque_info():
    """Retorna (total_skus, data_ultima_carga, clientes) do DB."""
    db = _carregar_db_estoque()
    if not db:
        return 0, None, []
    total = sum(len(skus) for skus in db.values())
    datas = []
    for skus in db.values():
        for sku in skus.values():
            if sku.get('atualizado'):
                datas.append(sku['atualizado'])
    ultima = max(datas) if datas else None
    return total, ultima, list(db.keys())

def _carregar_estoque_xlsx(arquivo_estoque, log):
    """Lê arquivo Excel de estoque (abas por cliente) e retorna dict estruturado."""
    xl = pd.ExcelFile(arquivo_estoque)
    db = {}
    hoje = date.today().strftime('%d/%m/%Y')
    for aba in xl.sheet_names:
        try:
            df = xl.parse(aba, header=None)
            # Dados começam na linha 4: col0=Código, col2=Descrição, col6=Saldo, col9=Reservado, col11=Bloqueado
            dados = df.iloc[4:, [0, 2, 6, 9, 11]].copy()
            dados.columns = ['Código','Descrição','Saldo','Reservado','Bloqueado']
            dados['Código'] = dados['Código'].astype(str).str.strip()
            dados = dados[dados['Código'].notna() & (dados['Código'] != 'nan') & (dados['Código'] != '')]
            for col in ['Saldo','Reservado','Bloqueado']:
                dados[col] = pd.to_numeric(dados[col], errors='coerce').fillna(0)
            cliente = aba.strip()
            db[cliente] = {}
            for _, row in dados.iterrows():
                db[cliente][row['Código']] = {
                    'desc':       str(row['Descrição']).strip() if pd.notna(row['Descrição']) else '',
                    'saldo':      int(row['Saldo']),
                    'reservado':  int(row['Reservado']),
                    'bloqueado':  int(row['Bloqueado']),
                    'atualizado': hoje,
                }
            log(f"   📋 {cliente}: {len(db[cliente])} SKUs\n")
        except Exception as e:
            log(f"   ⚠️ Aba '{aba}': {e}\n")
    return db


# ── DB DE FAMÍLIAS ────────────────────────────────────────────────────────────
# Clientes sem divisão de família (faturados como um bloco único)
_CLIENTES_SEM_DIVISAO = {'GSK CASA', 'GSK BLENREP', 'OPORTUTEK BRASIL', 'BANCO SANTANDER'}
# Clientes que usam coluna "Grupo" em vez de "Família"
_CLIENTES_USA_GRUPO   = {'YELUM SEGURADORA'}

def _carregar_db_familias():
    """Carrega o DB de famílias do JSON. Retorna {} se não existir."""
    try:
        with open(DB_FAMILIAS_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _salvar_db_familias(db):
    """Salva o DB de famílias no JSON."""
    try:
        with open(DB_FAMILIAS_PATH, 'w', encoding='utf-8') as f:
            json.dump(db, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

def _detectar_formato_familias(xl):
    """
    Detecta o formato do arquivo:
    - 'correcao' : planilha de retorno com coluna 'Família Sugerida' (pode estar até L4)
    - 'sistema'  : exportação do sistema com col5=Código, col13=Família
    """
    # Ignora aba de Resumo se existir
    abas = [a for a in xl.sheet_names if a.lower() != 'resumo']
    for aba in abas[:3]:
        try:
            df = xl.parse(aba, header=None)
            if df.empty:
                continue
            # Verifica as primeiras 5 linhas em busca de 'Família Sugerida'
            for i in range(min(5, len(df))):
                valores = [str(v).strip() for v in df.iloc[i] if pd.notna(v)]
                if any('sugerida' in v.lower() for v in valores):
                    return 'correcao'
        except Exception:
            continue
    return 'sistema'


def _carregar_familias_xlsx(arquivo, log):
    """
    Detecta automaticamente o formato do arquivo e processa:

    FORMATO SISTEMA  — exportação do cadastro de produtos (abas por cliente)
      col5=Código, col6=Descrição, col13=Família, col14=Grupo
      → carga completa, substitui todo o DB

    FORMATO CORREÇÃO — planilha de retorno preenchida pela gerência
      colunas: 'Código SKU', 'Descrição', 'Saldo em Estoque',
               'Família Sugerida', 'Observação'  (abas por cliente)
      → merge seguro: só preenche SKUs que estão SEM FAMÍLIA no DB atual
      → nunca sobrescreve famílias já corretas
      → ignora linhas com 'Família Sugerida' em branco
    """
    try:
        xl = pd.ExcelFile(arquivo)
    except Exception as e:
        log(f"❌ Erro ao abrir arquivo: {e}\n")
        return None

    formato = _detectar_formato_familias(xl)
    hoje = date.today().strftime('%d/%m/%Y')

    # Detecta se é parcial: arquivo sistema com menos abas que o DB atual
    db_atual = _carregar_db_familias()
    clientes_no_arquivo = [a.strip() for a in xl.sheet_names]
    is_parcial = (formato == 'sistema' and db_atual and
                  len(clientes_no_arquivo) < len(db_atual))

    if formato == 'sistema':
        if is_parcial:
            log(f"📋 Formato detectado: Exportação Parcial do Sistema\n")
            log(f"   → Merge: adiciona/atualiza apenas os {len(clientes_no_arquivo)} "
                f"cliente(s) do arquivo, mantém os demais intactos\n\n")
            db = dict(db_atual)  # começa com DB existente
        else:
            log(f"📋 Formato detectado: Exportação do Sistema\n")
            log(f"   → Carga completa do cadastro de produtos\n\n")
            db = {}

        for aba in xl.sheet_names:
            try:
                df = xl.parse(aba, header=None)
                cliente = aba.strip()
                db[cliente] = {}
                n_skus = 0
                for _, row in df.iloc[1:].iterrows():
                    codigo = str(row[5]).strip() if pd.notna(row[5]) else ''
                    desc   = str(row[6]).strip() if pd.notna(row[6]) else ''
                    if not codigo or codigo in ('nan', 'None', 'Código'):
                        continue
                    grupo = str(row[14]).strip() if pd.notna(row[14]) else ''
                    if grupo in ('nan', 'None', ''): grupo = ''
                    if cliente in _CLIENTES_SEM_DIVISAO:
                        familia = cliente
                    elif cliente in _CLIENTES_USA_GRUPO:
                        familia = grupo if grupo else 'SEM GRUPO'
                    else:
                        familia = str(row[13]).strip() if pd.notna(row[13]) else 'SEM FAMÍLIA'
                        if familia in ('nan', 'None', ''): familia = 'SEM FAMÍLIA'
                    db[cliente][codigo] = {
                        'desc': desc, 'familia': familia,
                        'grupo': grupo, 'atualizado': hoje}
                    n_skus += 1
                n_fam = len(set(v['familia'] for v in db[cliente].values()))
                regra = '(sem divisão)' if cliente in _CLIENTES_SEM_DIVISAO \
                        else '(usa Grupo)' if cliente in _CLIENTES_USA_GRUPO \
                        else '(usa Família)'
                log(f"   ✅ {cliente}: {n_skus} SKUs · {n_fam} famílias {regra}\n")
            except Exception as e:
                log(f"   ⚠️ Aba '{aba}': {e}\n")
        return db

    else:  # formato == 'correcao'
        log(f"🔧 Formato detectado: Planilha de Correções\n")
        log(f"   → Merge seguro: só preenche SKUs que estão SEM FAMÍLIA\n\n")

        # Carrega DB atual para fazer merge
        db = _carregar_db_familias()
        if not db:
            log("❌ DB de famílias vazio — faça a carga inicial com o arquivo do sistema primeiro.\n")
            return None

        total_atualizados = 0
        total_ignorados   = 0
        total_em_branco   = 0

        for aba in xl.sheet_names:
            try:
                # Encontra a linha do cabeçalho (tem 'Família Sugerida' como célula própria)
                df_raw = xl.parse(aba, header=None)
                header_row = None
                for i in range(min(6, len(df_raw))):
                    vals = [str(v).strip() for v in df_raw.iloc[i] if pd.notna(v)]
                    # Deve ter 'Família Sugerida' como célula separada (não dentro de frase longa)
                    if any(v.lower() in ('família sugerida', 'familia sugerida', 'família sugerida')
                           for v in vals):
                        header_row = i
                        break
                if header_row is None:
                    log(f"   ⚠️ {aba}: cabeçalho não encontrado — pulando\n")
                    continue
                df = xl.parse(aba, header=header_row)
                aba_strip = aba.strip()

                # Resolve nome do cliente no DB — a aba pode ser sigla (ex: "IPSEN")
                # mas o DB tem o nome completo (ex: "IPSEN FARMACEUTICA LTDA")
                cliente = aba_strip
                if aba_strip not in db:
                    # Tenta match parcial: nome do DB que contém a sigla ou vice-versa
                    norm = aba_strip.upper()
                    for k in db:
                        if norm in k.upper() or k.upper().startswith(norm):
                            cliente = k
                            break

                # Normaliza nomes de colunas
                df.columns = [str(c).strip() for c in df.columns]

                # Aceita variações de nome da coluna
                col_cod = next((c for c in df.columns
                                if 'digo' in c or c.lower() == 'código sku'), None)
                col_fam = next((c for c in df.columns
                                if 'sugerida' in c.lower() or c.lower() == 'família sugerida'), None)
                col_desc = next((c for c in df.columns
                                 if 'descri' in c.lower()), None)

                if not col_cod or not col_fam:
                    log(f"   ⚠️ {aba}: colunas esperadas não encontradas — pulando\n")
                    continue

                atualizados_aba = 0
                if cliente not in db:
                    db[cliente] = {}

                for _, row in df.iterrows():
                    codigo = str(row[col_cod]).strip() if pd.notna(row[col_cod]) else ''
                    if not codigo or codigo in ('nan', 'None', ''): continue

                    nova_fam = str(row[col_fam]).strip() if pd.notna(row[col_fam]) else ''
                    if nova_fam in ('nan', 'None', ''):
                        total_em_branco += 1
                        continue  # sem sugestão → ignora, não apaga nada

                    fam_atual = db[cliente].get(codigo, {}).get('familia', 'SEM FAMÍLIA')

                    if fam_atual != 'SEM FAMÍLIA':
                        # Já tem família correta → não toca
                        total_ignorados += 1
                        continue

                    # Só atualiza quem estava SEM FAMÍLIA
                    desc = str(row[col_desc]).strip() if col_desc and pd.notna(row.get(col_desc)) else \
                           db[cliente].get(codigo, {}).get('desc', '')
                    db[cliente][codigo] = {
                        'desc': desc, 'familia': nova_fam, 'atualizado': hoje}
                    atualizados_aba += 1
                    total_atualizados += 1

                if atualizados_aba > 0:
                    log(f"   ✅ {cliente}: {atualizados_aba} SKUs classificados\n")
                else:
                    log(f"   ─  {cliente}: nenhuma atualização necessária\n")

            except Exception as e:
                log(f"   ⚠️ Aba '{aba}': {e}\n")

        log(f"\n📊 Resumo do merge:\n")
        log(f"   ✅ {total_atualizados} SKUs classificados\n")
        log(f"   ─  {total_ignorados} SKUs ignorados (já tinham família)\n")
        log(f"   ⬜ {total_em_branco} SKUs sem sugestão preenchida (mantidos como SEM FAMÍLIA)\n")
        return db

def _db_familias_info():
    """Retorna resumo do DB de famílias para exibir na UI."""
    db = _carregar_db_familias()
    if not db:
        return 0, 0, '-'
    total_skus = sum(len(skus) for skus in db.values())
    total_cli  = len(db)
    # Data da última atualização
    ultima = '-'
    for skus in db.values():
        for info in skus.values():
            if info.get('atualizado'):
                ultima = info['atualizado']
            break
        break
    return total_skus, total_cli, ultima

def _get_familia_sku(cliente, codigo):
    """Retorna a família de um SKU. Fallback: 'SEM FAMÍLIA'."""
    db = _carregar_db_familias()
    return db.get(cliente, {}).get(codigo, {}).get('familia', 'SEM FAMÍLIA')

def _get_mapa_familias(cliente):
    """
    Retorna {familia: [lista de codigos]} para um cliente.
    Útil para o Fat. Armazenagem agrupar SKUs por família.
    """
    db = _carregar_db_familias()
    skus_cliente = db.get(cliente, {})
    mapa = {}
    for codigo, info in skus_cliente.items():
        fam = info.get('familia', 'SEM FAMÍLIA')
        mapa.setdefault(fam, []).append(codigo)
    return mapa


# ── DB DE CONFIGURAÇÃO DE FAT. ARMAZENAGEM ───────────────────────────────────
# Lido dinamicamente do arquivo — sem hardcode de nomes

def _carregar_db_precos_arm():
    """Carrega o DB de preços e configuração de fat. armazenagem."""
    try:
        with open(DB_PRECOS_ARM_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return {}

def _salvar_db_precos_arm(db):
    try:
        with open(DB_PRECOS_ARM_PATH, 'w', encoding='utf-8') as f:
            json.dump(db, f, ensure_ascii=False, indent=2)
        return True
    except Exception:
        return False

def _carregar_config_fat_arm_xlsx(arquivo, log):
    """
    Lê o arquivo de configuração de Fat. Armazenagem (duas abas):
      - 'Grupo-Familia': define o tipo de cada cliente (familia/grupo/unico)
                         e lista suas subdivisões
      - 'Valor de armaz.': preço/m³, seguro% e ISS% por cliente

    Regras especiais (sem hardcode — derivadas da configuração):
      - CM HOSPITALAR, GSK CASA, GSK BLENREP, PROJETO ALNYLAM → herdam preços da HUMANIA
      - IPSEN grupo RH → usa tabela IPSEN RH (Mensal); demais grupos → IPSEN (QUINZENAL)

    Retorna DB estruturado:
    {
      'clientes': {
        'ADITUS': {
          'tipo': 'familia',
          'subdivisoes': ['COMPASS', ...],
          'preco_m3': 100.05, 'seguro_pct': 0.1,
          'seguro_base': 1000, 'iss_pct': 2.0,
        },
        ...
      },
      'precos_raw': {          # tabelas de preço brutas (inclui IPSEN variantes)
        'IPSEN (QUINZENAL)': {...},
        'IPSEN RH (Mensal)': {...},
        'HUMANIA': {...},
        ...
      },
      'heranca': {             # depositantes que herdam preço de outro
        'CM HOSPITALAR S.A': 'HUMANIA',
        'GSK CASA': 'HUMANIA', ...
      },
      'ipsen_rh_grupos': [...],  # grupos do IPSEN que usam tabela RH
      'atualizado': 'DD/MM/AAAA'
    }
    """
    import re as _re
    try:
        xl = pd.ExcelFile(arquivo)
    except Exception as e:
        log(f"❌ Erro ao abrir arquivo: {e}\n")
        return None

    hoje = date.today().strftime('%d/%m/%Y')
    db = {
        'clientes': {},
        'precos_raw': {},
        'heranca': {},
        'ipsen_rh_grupos': [],
        'atualizado': hoje,
    }

    # ── Aba Grupo-Familia ─────────────────────────────────────────────
    if 'Grupo-Familia' not in xl.sheet_names:
        log("❌ Aba 'Grupo-Familia' não encontrada.\n"); return None

    try:
        df = xl.parse('Grupo-Familia', header=None)
        cli_fam = None
        cli_grp = None
        # Coleta depositantes únicos para detectar grupo HUMANIA
        depositantes_unicos = []

        for _, row in df.iterrows():
            c1 = str(row[1]).strip() if pd.notna(row[1]) else ''
            c2 = str(row[2]).strip() if pd.notna(row[2]) else ''
            c3 = str(row[3]).strip() if pd.notna(row[3]) else ''
            c5 = str(row[5]).strip() if pd.notna(row[5]) else ''
            c6 = str(row[6]).strip() if pd.notna(row[6]) else ''
            c8 = str(row[8]).strip() if pd.notna(row[8]) else ''

            if 'DIVIDIDOS' in c1 or 'DIVIDIDOS' in c5:
                continue

            # Depositante único
            if c8 and c8 not in ('nan', ''):
                db['clientes'][c8] = {
                    'tipo': 'unico', 'subdivisoes': [],
                    'preco_m3': 0, 'seguro_pct': 0,
                    'seguro_base': 1000, 'iss_pct': 0,
                }
                depositantes_unicos.append(c8)

            # Cliente por família
            if c1 and c1 not in ('nan', ''):
                cli_fam = c1
                if cli_fam not in db['clientes']:
                    db['clientes'][cli_fam] = {
                        'tipo': 'familia', 'subdivisoes': [],
                        'preco_m3': 0, 'seguro_pct': 0,
                        'seguro_base': 1000, 'iss_pct': 0,
                    }
            if c2 and c2 not in ('nan', '') and cli_fam:
                obs = f' ({c3})' if c3 and c3 not in ('nan', '') else ''
                db['clientes'][cli_fam]['subdivisoes'].append(c2 + obs)

            # Cliente por grupo
            if c5 and c5 not in ('nan', ''):
                cli_grp = c5
                if cli_grp not in db['clientes']:
                    db['clientes'][cli_grp] = {
                        'tipo': 'grupo', 'subdivisoes': [],
                        'preco_m3': 0, 'seguro_pct': 0,
                        'seguro_base': 1000, 'iss_pct': 0,
                    }
            if c6 and c6 not in ('nan', '') and cli_grp:
                db['clientes'][cli_grp]['subdivisoes'].append(c6)

        log(f"✅ {len(db['clientes'])} clientes configurados:\n")
        for cli, info in db['clientes'].items():
            log(f"   [{info['tipo'].upper()}] {cli}"
                + (f" — {len(info['subdivisoes'])} subdivisões"
                   if info['subdivisoes'] else '') + "\n")

    except Exception as e:
        log(f"❌ Erro ao ler aba Grupo-Familia: {e}\n"); return None

    # ── Aba Valor de armaz. ───────────────────────────────────────────
    if 'Valor de armaz.' not in xl.sheet_names:
        log("⚠️  Aba 'Valor de armaz.' não encontrada.\n"); return db

    try:
        df2 = xl.parse('Valor de armaz.', header=None)
        cli_preco = None

        for _, row in df2.iterrows():
            c1  = str(row[1]).strip() if pd.notna(row[1]) else ''
            c2  = row[2] if pd.notna(row[2]) else None
            c3  = row[3] if pd.notna(row[3]) else None
            c2s = str(c2).strip().lower() if c2 is not None else ''

            # Nova tabela de cliente
            if c2s in ('unid/m³', 'vol / m³', ' vol / m³', 'descrição', ' vol / m³'):
                cli_preco = c1
                # Registra também em precos_raw (inclui IPSEN variantes e HUMANIA)
                if cli_preco not in db['precos_raw']:
                    db['precos_raw'][cli_preco] = {
                        'preco_m3': 0, 'seguro_pct': 0,
                        'seguro_base': 1000, 'iss_pct': 0,
                    }
                continue

            if not cli_preco or not c1:
                continue

            # Preço por m³ (unid = 1)
            if c2 is not None:
                try:
                    if abs(float(str(c2).strip()) - 1.0) < 0.01 and c3:
                        preco_val = float(c3)
                        db['precos_raw'][cli_preco]['preco_m3'] = preco_val
                        matched = _match_cliente_preco(cli_preco, db['clientes'])
                        if matched:
                            db['clientes'][matched]['preco_m3'] = preco_val
                        continue
                except (ValueError, TypeError):
                    pass

            # Seguro
            if 'seguro' in c1.lower() and c3:
                m = _re.search(r'([\d,\.]+)\s*%', c1)
                if m:
                    pct = float(m.group(1).replace(',', '.'))
                    try:
                        base = float(str(c2).strip()) if c2 else 1000
                    except (ValueError, TypeError):
                        base = 1000
                    db['precos_raw'][cli_preco]['seguro_pct']  = pct
                    db['precos_raw'][cli_preco]['seguro_base'] = base
                    matched = _match_cliente_preco(cli_preco, db['clientes'])
                    if matched:
                        db['clientes'][matched]['seguro_pct']  = pct
                        db['clientes'][matched]['seguro_base'] = base
                continue

            # ISS
            if 'iss' in c1.lower() and c3:
                m = _re.search(r'([\d,\.]+)\s*%', c1)
                if m:
                    pct = float(m.group(1).replace(',', '.'))
                    db['precos_raw'][cli_preco]['iss_pct'] = pct
                    matched = _match_cliente_preco(cli_preco, db['clientes'])
                    if matched:
                        db['clientes'][matched]['iss_pct'] = pct
                continue

    except Exception as e:
        log(f"⚠️  Erro ao ler preços: {e}\n")

    # ── Regras especiais ──────────────────────────────────────────────
    # 1. HUMANIA: detecta qual tabela de preço é a "HUMANIA" e aplica
    #    aos depositantes que não têm preço próprio na tabela
    preco_humania = None
    for nome_raw, preco_raw in db['precos_raw'].items():
        if 'HUMANIA' in nome_raw.upper() and preco_raw.get('preco_m3', 0) > 0:
            preco_humania = preco_raw
            break

    if preco_humania:
        # Depositantes que herdam HUMANIA: aqueles sem preço próprio na tabela
        for cli, info in db['clientes'].items():
            if info['tipo'] == 'unico' and info.get('preco_m3', 0) == 0:
                # Verifica se não tem tabela própria com esse nome
                tem_tabela = any(
                    cli.upper() in k.upper() or k.upper() in cli.upper()
                    for k in db['precos_raw']
                )
                if not tem_tabela:
                    db['clientes'][cli].update({
                        'preco_m3':    preco_humania['preco_m3'],
                        'seguro_pct':  preco_humania['seguro_pct'],
                        'seguro_base': preco_humania['seguro_base'],
                        'iss_pct':     preco_humania['iss_pct'],
                    })
                    db['heranca'][cli] = 'HUMANIA'
                    log(f"   💰 {cli} → herda preços da HUMANIA\n")

    # 2. IPSEN: detecta tabela RH e tabela quinzenal
    #    O grupo 'RH' usa a tabela com 'RH' no nome; demais usam a quinzenal
    ipsen_rh_preco    = None
    ipsen_pad_preco   = None
    ipsen_rh_grupos   = []

    for nome_raw, preco_raw in db['precos_raw'].items():
        nu = nome_raw.upper()
        if 'IPSEN' in nu and 'RH' in nu:
            ipsen_rh_preco = preco_raw
        elif 'IPSEN' in nu and preco_raw.get('preco_m3', 0) > 0:
            ipsen_pad_preco = preco_raw

    # Detecta grupos do IPSEN que usam RH
    for cli, info in db['clientes'].items():
        if info['tipo'] == 'grupo' and 'IPSEN' in cli.upper():
            for sub in info['subdivisoes']:
                if 'RH' in sub.upper():
                    ipsen_rh_grupos.append(sub)
            # Aplica preço padrão (quinzenal) ao cliente IPSEN principal
            if ipsen_pad_preco and info.get('preco_m3', 0) == 0:
                db['clientes'][cli].update({
                    'preco_m3':    ipsen_pad_preco['preco_m3'],
                    'seguro_pct':  ipsen_pad_preco['seguro_pct'],
                    'seguro_base': ipsen_pad_preco['seguro_base'],
                    'iss_pct':     ipsen_pad_preco['iss_pct'],
                })

    db['ipsen_rh_grupos']  = ipsen_rh_grupos
    db['ipsen_rh_preco']   = ipsen_rh_preco
    db['ipsen_pad_preco']  = ipsen_pad_preco

    if ipsen_rh_grupos:
        log(f"   💰 IPSEN grupos RH (preço mensal): {ipsen_rh_grupos}\n")

    log(f"\n💰 Preços finais:\n")
    for cli, info in db['clientes'].items():
        p = info.get('preco_m3', 0)
        if p > 0:
            her = f" [herda {db['heranca'][cli]}]" if cli in db['heranca'] else ''
            log(f"   {cli}: R$ {p:.2f}/m³"
                f" + seguro {info['seguro_pct']}%"
                f" + ISS {info['iss_pct']}%{her}\n")
        else:
            log(f"   ⚠️  {cli}: sem preço configurado\n")

    return db


def _match_cliente_preco(nome_preco, clientes_db):
    """
    Casa o nome do cliente da aba de preços com a chave do DB.
    Ex: 'ADITUS HEALTH' → 'ADITUS', 'CSL' → 'CSL BEHRING'
    """
    n = nome_preco.upper().strip()
    # Tenta match exato primeiro
    for k in clientes_db:
        if k.upper() == n:
            return k
    # Match parcial — um contém o outro
    for k in clientes_db:
        ku = k.upper()
        if n in ku or ku in n or ku.split()[0] == n.split()[0]:
            return k
    return None


def _calcular_total_m3(config_cliente, volume_m3):
    """
    Calcula o valor total a faturar para um dado volume m³.
    Fórmula: (preco_m3 × vol) + seguro + ISS
    O seguro é calculado sobre a base (por 1000 unid de NF) × volume
    """
    preco    = config_cliente.get('preco_m3', 0) * volume_m3
    seg_base = config_cliente.get('seguro_base', 1000)
    seg_pct  = config_cliente.get('seguro_pct', 0) / 100
    seguro   = seg_pct * seg_base * volume_m3
    subtotal = preco + seguro
    iss      = config_cliente.get('iss_pct', 0) / 100 * subtotal
    return round(subtotal + iss, 4)


def _db_precos_arm_info():
    """Retorna resumo do DB de preços para exibir na UI."""
    db = _carregar_db_precos_arm()
    if not db or not db.get('clientes'):
        return 0, '-'
    total = sum(1 for v in db['clientes'].values() if v.get('preco_m3', 0) > 0)
    return total, db.get('atualizado', '-')


def _atualizar_db_com_movimentacao(arquivo_movimentacao, log):
    """Atualiza saldos e acumula histórico de última movimentação por SKU no DB."""
    db = _carregar_db_estoque()
    if not db:
        log("❌ DB vazio — faça a carga inicial primeiro.\n")
        return False

    xl = pd.ExcelFile(arquivo_movimentacao)
    hoje = date.today().strftime('%d/%m/%Y')
    total_saldos = 0
    total_mov    = 0
    total_novos  = 0

    for aba in xl.sheet_names:
        try:
            df = xl.parse(aba, header=None)
            cliente = aba.strip()
            if cliente not in db:
                db[cliente] = {}

            # ── Passo 1: última data de movimentação real por SKU ──────────
            # Linhas de movimentação têm data válida na col 0, código entre [] na col 2
            ultimas_mov = {}  # {codigo: date mais recente}
            for _, row in df.iterrows():
                data_str    = str(row[0]).strip()
                produto_str = str(row[2]) if pd.notna(row[2]) else ''
                match = re.search(r'\[(.+?)\]', produto_str)
                if not match:
                    continue
                try:
                    dt = pd.to_datetime(data_str, dayfirst=True, errors='raise')
                    cod = match.group(1).strip()
                    if cod not in ultimas_mov or dt > ultimas_mov[cod]:
                        ultimas_mov[cod] = dt
                except (ValueError, TypeError):
                    pass

            # ── Passo 2: Saldo Final por SKU ──────────────────────────────
            mask_final = df[0].astype(str).str.strip() == 'Saldo Final'
            df_final   = df[mask_final].copy()

            for _, row in df_final.iterrows():
                produto_str = str(row[2]) if pd.notna(row[2]) else ''
                match = re.search(r'\[(.+?)\]', produto_str)
                if not match:
                    continue
                codigo = match.group(1).strip()
                saldo  = int(pd.to_numeric(row[11], errors='coerce') or 0)

                if codigo in db[cliente]:
                    db[cliente][codigo]['saldo'] = saldo
                    # Acumula última movimentação: só substitui se nova data for mais recente
                    if codigo in ultimas_mov:
                        nova_dt   = ultimas_mov[codigo].strftime('%d/%m/%Y')
                        atual_str = db[cliente][codigo].get('ultima_mov', '')
                        if atual_str:
                            try:
                                atual_dt = pd.to_datetime(atual_str, dayfirst=True)
                                if ultimas_mov[codigo] > atual_dt:
                                    db[cliente][codigo]['ultima_mov'] = nova_dt
                                    total_mov += 1
                            except (ValueError, TypeError):
                                db[cliente][codigo]['ultima_mov'] = nova_dt
                                total_mov += 1
                        else:
                            db[cliente][codigo]['ultima_mov'] = nova_dt
                            total_mov += 1
                    db[cliente][codigo]['atualizado'] = hoje
                    total_saldos += 1
                else:
                    desc   = produto_str.split(']', 1)[-1].strip() if ']' in produto_str else produto_str
                    ultima = ultimas_mov.get(codigo)
                    db[cliente][codigo] = {
                        'desc':      desc,
                        'saldo':     saldo,
                        'reservado': 0,
                        'bloqueado': 0,
                        'atualizado': hoje,
                        'ultima_mov': ultima.strftime('%d/%m/%Y') if ultima else '',
                    }
                    total_novos += 1

            n_mov = len([c for c in ultimas_mov if c in db[cliente]])
            log(f"   ✅ {cliente}: {len(df_final)} saldos | {n_mov} movimentações registradas\n")
        except Exception as e:
            log(f"   ⚠️ Aba '{aba}': {e}\n")

    _salvar_db_estoque(db)
    log(f"\n💾 DB atualizado: {total_saldos} saldos | {total_mov} datas de mov. atualizadas | {total_novos} SKUs novos.\n")
    return True

def _caminho_saida(pasta_modulo, arquivo_base, mes_ref=None):
    """Cria subpasta AAAA e retorna caminho completo do arquivo.
    mes_ref: 'MM-AAAA' opcional — sobrescreve o mês automático (mês anterior).
    Retorna None se a pasta não puder ser criada (ex: drive de rede inacessível)."""
    if mes_ref:
        try:
            mm, aaaa = mes_ref.split('-')
            mes, ano = int(mm), int(aaaa)
        except Exception:
            mes, ano = _mes_ano_referencia()
    else:
        mes, ano = _mes_ano_referencia()
    subpasta = os.path.join(pasta_modulo, str(ano))
    try:
        os.makedirs(subpasta, exist_ok=True)
    except Exception:
        return None
    nome = f'{arquivo_base}_{mes:02d}-{ano}.xlsx'
    caminho = os.path.join(subpasta, nome)
    contador = 1
    while os.path.exists(caminho):
        nome = f'{arquivo_base}_{mes:02d}-{ano}_{contador}.xlsx'
        caminho = os.path.join(subpasta, nome)
        contador += 1
    return caminho

# ============================================================
# HISTÓRICO CONSOLIDADO
# ============================================================

def _atualizar_historico(modulo_key, caminho_relatorio, mes_ano_str, log):
    """
    Copia todas as abas do relatório gerado para o consolidado anual do módulo.
    Cada mês vira um conjunto de abas (ex: 'Mar-2026 · Resumo').
    Se as abas do mês já existirem, são substituídas.

    modulo_key:        ex: 'pedidos', 'fretes', 'armazenagem', 'estoque', 'produtividade'
    caminho_relatorio: caminho do .xlsx recém-gerado
    mes_ano_str:       ex: '03-2026'
    """
    _NOMES_MODULO = {
        'pedidos':            'Pedidos_e_Recebimentos',
        'fretes':             'Fretes',
        'armazenagem':        'Armazenagem',
        'estoque':            'Estoque',
        'produtividade':      'Produtividade',
        'cap_operacional':    'Capacidade_Operacional',
        'recebimentos':       'Recebimentos_e_Devolucoes',
        'financeiro':         'Financeiro',
    }
    _MESES_PT = ['Jan','Fev','Mar','Abr','Mai','Jun',
                 'Jul','Ago','Set','Out','Nov','Dez']

    try:
        mm, aaaa = mes_ano_str.split('-')
        prefixo  = f'{_MESES_PT[int(mm)-1]}-{aaaa}'  # ex: 'Mar-2026'
    except Exception:
        prefixo  = mes_ano_str

    try:
        os.makedirs(PASTA_CONSOLIDADOS, exist_ok=True)
    except Exception as e:
        log(f"⚠️  Consolidado: não foi possível criar pasta '{PASTA_CONSOLIDADOS}': {e}\n")
        return

    nome_base  = _NOMES_MODULO.get(modulo_key, modulo_key.capitalize())
    nome_arq   = f'{nome_base}_{aaaa}.xlsx'
    cons_path  = os.path.join(PASTA_CONSOLIDADOS, nome_arq)

    try:
        # Abre ou cria o arquivo consolidado anual
        if os.path.isfile(cons_path):
            wb_cons = load_workbook(cons_path)
        else:
            wb_cons = Workbook()
            wb_cons.remove(wb_cons.active)

        # Remove abas deste mês se já existirem (substituição)
        for aba in list(wb_cons.sheetnames):
            if aba.startswith(prefixo):
                del wb_cons[aba]
                log(f"🔄 Consolidado: abas '{prefixo}' substituídas.\n")

        # Copia todas as abas do relatório mensal com prefixo do mês
        wb_rel = load_workbook(caminho_relatorio)
        for sheet_name in wb_rel.sheetnames:
            ws_src = wb_rel[sheet_name]
            if len(wb_rel.sheetnames) == 1:
                aba_dest = prefixo[:31]
            else:
                sufixo   = sheet_name[:18]
                aba_dest = f'{prefixo} · {sufixo}'[:31]
                base, cnt = aba_dest, 1
                while aba_dest in wb_cons.sheetnames:
                    aba_dest = f'{base[:28]}_{cnt}'
                    cnt += 1

            ws_dst = wb_cons.create_sheet(aba_dest)

            # Copia células com estilos
            for row in ws_src.iter_rows():
                for cell in row:
                    new_cell = ws_dst.cell(row=cell.row, column=cell.column, value=cell.value)
                    if cell.has_style:
                        new_cell.font          = cell.font.copy()
                        new_cell.fill          = cell.fill.copy()
                        new_cell.border        = cell.border.copy()
                        new_cell.alignment     = cell.alignment.copy()
                        new_cell.number_format = cell.number_format

            for col, dim in ws_src.column_dimensions.items():
                ws_dst.column_dimensions[col].width = dim.width
            for row, dim in ws_src.row_dimensions.items():
                ws_dst.row_dimensions[row].height = dim.height

        # Ordena abas cronologicamente por mês
        def _chave_aba(nome):
            for i, abrev in enumerate(_MESES_PT):
                if nome.startswith(abrev + '-'):
                    return i
            return 99
        wb_cons._sheets = sorted(wb_cons._sheets, key=lambda ws: _chave_aba(ws.title))

        wb_cons.save(cons_path)
        log(f"📚 Consolidado anual atualizado: {nome_arq}\n")
        _gerar_pbids(log)

    except Exception as e:
        log(f"⚠️  Consolidado: erro ao atualizar — {e}\n")


def _gerar_pbids(log=None):
    """
    Gera um .pbids por módulo na pasta Consolidados.
    Protocolo correto para Excel local: "file" com "path".
    Um arquivo por fonte (limitação do Power BI).

    Resultado:
      Consolidados/PowerBI_Pedidos.pbids
      Consolidados/PowerBI_Fretes.pbids
      ...
    """
    import json as _json
    import glob  as _glob

    pasta = PASTA_CONSOLIDADOS
    try:
        os.makedirs(pasta, exist_ok=True)
    except Exception:
        return

    # Mapa: prefixo do arquivo xlsx → nome do .pbids
    _MODULOS_MAP = {
        'Pedidos_e_Recebimentos': 'PowerBI_Pedidos',
        'Fretes':                 'PowerBI_Fretes',
        'Armazenagem':            'PowerBI_Armazenagem',
        'Estoque':                'PowerBI_Estoque',
        'Produtividade':          'PowerBI_Produtividade',
        'Capacidade_Operacional':    'PowerBI_CapOperacional',
        'Recebimentos_e_Devolucoes': 'PowerBI_Recebimentos',
        'consolidado':            'PowerBI_Dashboard',
    }

    gerados = 0
    for prefixo, nome_pbids in _MODULOS_MAP.items():
        # Pega o arquivo mais recente do módulo (ex: Fretes_2026.xlsx)
        matches = sorted(
            _glob.glob(os.path.join(pasta, f'{prefixo}*.xlsx')),
            key=os.path.getmtime, reverse=True
        )
        if not matches:
            continue

        caminho_xlsx = os.path.abspath(matches[0]).replace('/', '\\')
        pbids = {
            "version": "0.1",
            "connections": [
                {
                    "details": {
                        "protocol": "file",
                        "address": {
                            "path": caminho_xlsx
                        }
                    },
                    "mode": "Import"
                }
            ]
        }

        pbids_path = os.path.join(pasta, f'{nome_pbids}.pbids')
        try:
            with open(pbids_path, 'w', encoding='utf-8') as f:
                _json.dump(pbids, f, ensure_ascii=False, indent=2)
            gerados += 1
        except Exception as e:
            if log:
                log(f"⚠️  Power BI ({nome_pbids}): {e}\n")

    if log and gerados:
        log(f"🔗 Power BI: {gerados} arquivo(s) .pbids atualizados em Consolidados\\\n")


# ============================================================
# WRAPPERS — run_* chamam processamento + registram extração
# ============================================================

def run_pedidos(caminho, log, mes_ref=None):
    saida = _caminho_saida(PASTA_PEDIDOS, 'planilha_de_resultados_detalhada',
                           mes_ref=mes_ref)
    if saida is None:
        log(f"❌ Não foi possível criar a pasta de saída em:\n   {PASTA_PEDIDOS}\n"
            f"   Verifique se o drive está acessível.\n")
        return False
    ok = processar_pedidos(caminho, log, _saida_override=saida)
    if ok:
        _registrar_extracao('Pedidos e Recebimentos')
        mes_hist = mes_ref or f'{_mes_ano_referencia()[0]:02d}-{_mes_ano_referencia()[1]}'
        _atualizar_historico('pedidos', saida, mes_hist, log)
        return saida
    return False

def run_fretes(caminho, log, mes_ref=None):
    saida = _caminho_saida(PASTA_FRETES, 'relatorio_fretes', mes_ref=mes_ref)
    if saida is None:
        log(f"❌ Não foi possível criar a pasta de saída em:\n   {PASTA_FRETES}\n"
            f"   Verifique se o drive está acessível.\n")
        return False
    ok = processar_fretes(caminho, log, _saida_override=saida)
    if ok:
        _registrar_extracao('Fretes')
        mes_hist = mes_ref or f'{_mes_ano_referencia()[0]:02d}-{_mes_ano_referencia()[1]}'
        _atualizar_historico('fretes', saida, mes_hist, log)
        return saida
    return False

def run_armazenagem(caminho, mes_filtro, log, mes_ref=None):
    saida = _caminho_saida(PASTA_ARMAZENAGEM, 'relatorio_armazenagem',
                           mes_ref=mes_ref or mes_filtro)
    if saida is None:
        log(f"❌ Não foi possível criar a pasta de saída em:\n   {PASTA_ARMAZENAGEM}\n"
            f"   Verifique se o drive está acessível.\n")
        return False
    ok = processar_armazenagem(caminho, mes_filtro, log, _saida_override=saida)
    if ok:
        _registrar_extracao('Armazenagem')
        mes_hist = mes_ref or mes_filtro or f'{_mes_ano_referencia()[0]:02d}-{_mes_ano_referencia()[1]}'
        _atualizar_historico('armazenagem', saida, mes_hist, log)
        return saida
    return False

# ============================================================
# DE-PARA SKU → CLIENTE
# ============================================================

def _carregar_de_para():
    try:
        with open(DE_PARA_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _salvar_de_para(de_para):
    with open(DE_PARA_PATH, 'w', encoding='utf-8') as f:
        json.dump(de_para, f, ensure_ascii=False, indent=2)

# ============================================================
# MÓDULO 4 — PROCESSAMENTO DE ESTOQUE
# ============================================================

def processar_estoque(arquivo_estoque, pasta_volumes, arquivo_movimentacao,
                      dias_ocioso, log, _saida_override=None):
    try:
        hoje = date.today()

        def _col(df, candidatos):
            for c in candidatos:
                match = [x for x in df.columns if x.lower().strip() == c.lower()]
                if match: return match[0]
            return None

        # ── Carrega dados do DB interno ──────────────────────────────────
        db = _carregar_db_estoque()
        if not db:
            log("❌ Banco de dados interno vazio.\n")
            log("   Use o botão 'Carga Inicial do DB' para carregar o cadastro de SKUs.\n")
            return False

        log(f"📂 Carregando dados do banco de dados interno...\n")
        rows = []
        for cliente, skus in db.items():
            for cod, info in skus.items():
                rows.append({
                    'Cliente':    cliente,
                    'Código':     cod,
                    'Descrição':  info.get('desc',''),
                    'Saldo':      info.get('saldo', 0),
                    'Reservado':  info.get('reservado', 0),
                    'Bloqueado':  info.get('bloqueado', 0),
                })
        df_est = pd.DataFrame(rows)
        df_est = df_est[df_est['Saldo'] > 0].copy()
        log(f"✅ {len(df_est)} SKUs com saldo > 0 ({df_est['Cliente'].nunique()} clientes).\n")

        # ── Pico de Estoque: 1 arquivo com abas por cliente ─────────────
        # Estrutura: linha 0-2 = cabeçalho, linha 3+ = dados diários, IGNORAR última linha (bug sistema)
        # Col 0=Data, Col 10=Qtd, Col 13=Área m², Col 14=Volume m³
        log("📦 Carregando pico de estoque por cliente...\n")
        pico_por_cliente = {}  # {cliente: {pico_vol, saldo_vol, pico_area, saldo_area}}
        arq_vol = pasta_volumes  # agora é direto um arquivo

        if arq_vol and os.path.isfile(arq_vol):
            xl_vol = pd.ExcelFile(arq_vol)
            for aba in xl_vol.sheet_names:
                try:
                    df_v = xl_vol.parse(aba, header=None)
                    # Dados: linhas 3 até penúltima (ignora última = bug sistema)
                    dados = df_v.iloc[3:-1].copy()
                    if dados.empty:
                        continue
                    vol_s  = pd.to_numeric(dados[14], errors='coerce').dropna()
                    area_s = pd.to_numeric(dados[13], errors='coerce').dropna()
                    if vol_s.empty:
                        continue
                    pico_por_cliente[aba.strip()] = {
                        'pico_vol':   round(float(vol_s.max()), 4),
                        'saldo_vol':  round(float(vol_s.iloc[-1]), 4),
                        'pico_area':  round(float(area_s.max()), 4) if not area_s.empty else 0,
                        'saldo_area': round(float(area_s.iloc[-1]), 4) if not area_s.empty else 0,
                    }
                    log(f"   📋 {aba}: pico {vol_s.max():.2f} m³ | saldo {vol_s.iloc[-1]:.2f} m³\n")
                except Exception as e:
                    log(f"   ⚠️ Pico aba '{aba}': {e}\n")
            log(f"✅ Pico de estoque carregado ({len(pico_por_cliente)} clientes).\n")
        else:
            log("   ⚠️ Arquivo de pico de estoque não informado — volumes ficarão zerados.\n")

        # ── Última Movimentação: lida direto do DB (acumulada mês a mês) ──
        log("🔄 Carregando histórico de movimentações do DB...\n")
        mov_map = {}  # {(cliente, codigo): datetime}
        for cliente_db, skus in db.items():
            for cod, info in skus.items():
                ultima_str = info.get('ultima_mov', '')
                if ultima_str:
                    try:
                        mov_map[(cliente_db, cod)] = pd.to_datetime(ultima_str, dayfirst=True)
                    except (ValueError, TypeError):
                        pass

        df_est['Última Movimentação'] = df_est.apply(
            lambda r: mov_map.get((r['Cliente'], r['Código']), pd.NaT), axis=1)
        n_com_mov = df_est['Última Movimentação'].notna().sum()
        log(f"✅ {n_com_mov} SKUs com histórico de movimentação no DB.\n")
        sem_mov = len(df_est) - n_com_mov
        if sem_mov > 0:
            log(f"   ⚠️ {sem_mov} SKUs sem data de movimentação registrada (serão marcados como ociosos).\n")

        # Status: Zerado → Ocioso → Ativo
        log(f"📊 Calculando ociosidade ({dias_ocioso} dias)...\n")
        hoje_ts = pd.Timestamp(hoje)
        df_est['Dias Sem Mov.'] = (hoje_ts - df_est['Última Movimentação']).dt.days

        def _classificar_status(r):
            if r['Saldo'] <= 0:
                return 'Zerado'
            if pd.isna(r['Última Movimentação']) or r['Dias Sem Mov.'] >= dias_ocioso:
                return 'Ocioso'
            return 'Ativo'

        df_est['Status'] = df_est.apply(_classificar_status, axis=1)
        df_est['Última Movimentação'] = df_est['Última Movimentação'].dt.strftime('%d/%m/%Y').fillna('Sem registro')

        # Resumo por cliente — usa pico_por_cliente para volumes
        rows = []
        for cliente, grp in df_est.groupby('Cliente'):
            pico  = pico_por_cliente.get(cliente, {})
            pico_vol   = pico.get('pico_vol',   0)
            saldo_vol  = pico.get('saldo_vol',  0)
            pico_area  = pico.get('pico_area',  0)
            saldo_area = pico.get('saldo_area', 0)
            st     = len(grp)
            so     = (grp['Status'] == 'Ocioso').sum()   # zerados excluídos
            sz     = (grp['Status'] == 'Zerado').sum()
            saldo_t = int(grp['Saldo'].sum())
            pct_oc = so / st if st else 0
            vol_ocioso = round(saldo_vol * pct_oc, 4)
            rows.append({
                'Cliente':               cliente,
                'SKUs em Estoque':       st,
                'SKUs Ociosos':          so,
                'SKUs Zerados':          sz,
                '% SKUs Ociosos':        f'{round(pct_oc*100,1)}%',
                'Qtd. Total em Estoque': saldo_t,
                'Pico Volume m³':        pico_vol,
                'Saldo Final m³':        saldo_vol,
                'Pico Área m²':          pico_area,
                'Saldo Final m²':        saldo_area,
                'Vol. Ocioso Est. m³':   vol_ocioso,
            })
        df_res = pd.DataFrame(rows).sort_values('Pico Volume m³', ascending=False)
        tot = {
            'Cliente':               'TOTAL GERAL',
            'SKUs em Estoque':       int(df_res['SKUs em Estoque'].sum()),
            'SKUs Ociosos':          int(df_res['SKUs Ociosos'].sum()),
            'SKUs Zerados':          int(df_res['SKUs Zerados'].sum()),
            '% SKUs Ociosos':        '',
            'Qtd. Total em Estoque': int(df_res['Qtd. Total em Estoque'].sum()),
            'Pico Volume m³':        round(df_res['Pico Volume m³'].sum(), 4),
            'Saldo Final m³':        round(df_res['Saldo Final m³'].sum(), 4),
            'Pico Área m²':          round(df_res['Pico Área m²'].sum(), 4),
            'Saldo Final m²':        round(df_res['Saldo Final m²'].sum(), 4),
            'Vol. Ocioso Est. m³':   round(df_res['Vol. Ocioso Est. m³'].sum(), 4),
        }
        df_res = pd.concat([df_res, pd.DataFrame([tot])], ignore_index=True)

        col_order = ['Cliente','Código','Descrição','Saldo','Reservado','Bloqueado',
                     'Última Movimentação','Dias Sem Mov.','Status']
        col_order = [c for c in col_order if c in df_est.columns]
        df_det = df_est[col_order].copy()
        # Ordem: Zerado → Ocioso → Ativo
        status_ordem = {'Zerado': 0, 'Ocioso': 1, 'Ativo': 2}
        df_det['_ord'] = df_det['Status'].map(status_ordem)
        df_det.sort_values(['Cliente', '_ord', 'Dias Sem Mov.'], ascending=[True, True, False], inplace=True)
        df_det.drop(columns=['_ord'], inplace=True)

        saida = _saida_override if _saida_override else \
            _caminho_saida(PASTA_ESTOQUE, 'relatorio_estoque')
        log(f"💾 Salvando: {saida}\n")
        with pd.ExcelWriter(saida, engine='openpyxl') as writer:
            df_res.to_excel(writer, sheet_name='Resumo por Cliente', index=False)
            df_det.to_excel(writer, sheet_name='Detalhe por SKU', index=False)
        wb = load_workbook(saida)
        for aba, cor in [('Resumo por Cliente','B45309'),('Detalhe por SKU','92400E')]:
            if aba in wb.sheetnames:
                ws = wb[aba]
                _formatar_aba_generica(ws, cor_header=cor)
                hdr = [ws.cell(1, c).value for c in range(1, ws.max_column+1)]
                # Localiza coluna Status para colorir linhas
                col_status = next((i+1 for i, h in enumerate(hdr) if h == 'Status'), None)
                for ci, h in enumerate(hdr, 1):
                    if h in ('SKUs em Estoque','SKUs Ociosos','SKUs Zerados','Qtd. Total em Estoque',
                              'Reservado','Bloqueado','Dias Sem Mov.'):
                        for ri in range(2, ws.max_row+1):
                            cell = ws.cell(ri, ci)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0'
                    elif h in ('Pico Volume m³','Saldo Final m³','Pico Área m²',
                               'Saldo Final m²','Vol. Ocioso Est. m³','Saldo'):
                        for ri in range(2, ws.max_row+1):
                            cell = ws.cell(ri, ci)
                            if isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.0000'
                # Colorir linhas por status na aba Detalhe
                if aba == 'Detalhe por SKU' and col_status:
                    fill_zerado = PatternFill(fill_type='solid', fgColor='D1D5DB')  # cinza claro
                    fill_ocioso = PatternFill(fill_type='solid', fgColor='FEF3C7')  # âmbar claro
                    font_zerado = Font(color='6B7280')                     # texto cinza
                    font_ocioso = Font(color='92400E')                     # texto âmbar
                    for ri in range(2, ws.max_row+1):
                        status_val = ws.cell(ri, col_status).value
                        if status_val == 'Zerado':
                            for ci in range(1, ws.max_column+1):
                                ws.cell(ri, ci).fill = fill_zerado
                                ws.cell(ri, ci).font = font_zerado
                        elif status_val == 'Ocioso':
                            for ci in range(1, ws.max_column+1):
                                ws.cell(ri, ci).fill = fill_ocioso
                                ws.cell(ri, ci).font = font_ocioso
        wb.save(saida)
        log("✅ Relatório de estoque gerado!\n")
        return saida

    except Exception as e:
        log(f"❌ Erro: {e}\n{traceback.format_exc()}")
        return False

def run_estoque(arquivo_estoque, pasta_volumes, arquivo_movimentacao, dias_ocioso, log,
                mes_ref=None):
    saida = _caminho_saida(PASTA_ESTOQUE, 'relatorio_estoque', mes_ref=mes_ref)
    resultado = processar_estoque(arquivo_estoque, pasta_volumes,
                                  arquivo_movimentacao, dias_ocioso, log,
                                  _saida_override=saida)
    if resultado:
        _registrar_extracao('Estoque')
        mes_hist = mes_ref or f'{_mes_ano_referencia()[0]:02d}-{_mes_ano_referencia()[1]}'
        _atualizar_historico('estoque', saida, mes_hist, log)
        return saida
    return False

# ============================================================
# MÓDULO 5 — PRODUTIVIDADE DE EQUIPE
# ============================================================

TURNO_H = 9.0   # horas do turno 08:00–18:00 descontado 1h de almoço (12h–13h)

ETAPAS_PROD = [
    ('Aconselhamento',
     'DATA ÍNICIO ACONSELHAMENTO',     'HORA INÍCIO EMISSÃO ACONSELHAMENTO',
     'DATA FINAL ACONSELHAMENTO',      'HORA IMPRESSÃO ACONSELHAMENTO',      'QUEM?'),
    ('Picking',
     'DATA INÍCIO PICKING',             'HORA INÍCIO PICKING',
     'DATA FINAL PICKING',              'HORA FINAL PICKING',                 'QUEM?.1'),
    ('Manuseio',
     'DATA INÍCIO MANUSEIO',            'HORA INÍCIO MANUSEIO',
     'DATA FINAL MANUSEIO',             'HORA FINALIZAÇÃO CONF/MANUSEIO',     'QUEM?.2'),
    ('Emissão Etiq.',
     'DATA INÍCIO EMISSÃO DE ETIQUETA', 'HORA INÍCIO EMISSÃO ETIQUETA ',
     'DATA FINAL IMPRESSÃO DE ETIQUETA','HORA FINAL IMPRESSÃO ETIQUETA ',     'QUEM?.3'),
    ('Etiquetagem',
     'DATA INÍCIO ETIQUETAGEM',         'HORA INÍCIO ETIQUETAGEM',
     'DATA FINAL ETIQUETAGEM',          'HORA FINAL ETIQUETAGEM',             'QUEM?.4'),
]

_ALIASES_PROD = {
    'flavio':'Flávio', 'flávio':'Flávio', 'julia':'Julia', 'lemule':'Lemuel',
    'isa':'Isabela',   'isabela':'Isabela', 'thaine':'Thaine', 'thaís':'Thaís',
    'jean':'Jean',     'yasmin':'Yasmin',  'glauber':'Glauber', 'marcela':'Marcela',
    'gustavo':'Gustavo','anré':'Anré',     'lemuel':'Lemuel',
}

def _prod_norm_nome(raw):
    k = str(raw).strip().lower()
    return _ALIASES_PROD.get(k, k.title())

def _prod_parse_ini(d_col, h_col):
    try:
        d  = str(d_col).split()[0]
        dt = pd.to_datetime(d + ' ' + str(h_col), errors='coerce')
        if pd.isna(dt) or not (2020 <= dt.year <= 2025): return pd.NaT
        if dt.hour < 5 or dt.hour > 22:                  return pd.NaT
        return dt
    except (ValueError, TypeError):
        return pd.NaT

def _prod_parse_fim(d_col, h_col, ini):
    try:
        d  = str(d_col).split()[0]
        dt = pd.to_datetime(d + ' ' + str(h_col), errors='coerce')
        if pd.isna(dt) or not (2020 <= dt.year <= 2025): return pd.NaT
        if pd.notna(ini) and (dt - ini) > timedelta(hours=9): return pd.NaT
        return dt
    except (ValueError, TypeError):
        return pd.NaT

def _prod_split_colabs(raw):
    """'Lemuel/Thaís' → ['Lemuel','Thaís']. Tempo completo para cada um."""
    raw     = str(raw).strip()
    ignorar = ('nan', 'cancelada', '', 'sem assinatura',
               'processo s/ etiqueta', 's/ colagem de etiqueta')
    if raw.lower() in ignorar: return []
    for sep in ['/', ' e ']:
        if sep in raw:
            return [_prod_norm_nome(x) for x in raw.split(sep)
                    if x.strip() and len(x.strip()) > 1]
    n = _prod_norm_nome(raw)
    return [n] if len(n) > 1 else []

def _prod_merge_min(intervalos):
    """União de intervalos sobrepostos → minutos reais trabalhados."""
    if not intervalos: return 0.0
    ivs    = sorted(intervalos)
    merged = [list(ivs[0])]
    for s, e in ivs[1:]:
        if s <= merged[-1][1]: merged[-1][1] = max(merged[-1][1], e)
        else:                  merged.append([s, e])
    return sum((e - s).total_seconds() / 60 for s, e in merged)

def _prod_extrair_eventos(df_raw):
    eventos = []; ignorados = 0
    for _, row in df_raw.iterrows():
        qtd = row.get('QTDE SAÍDAS', 1)
        try:   qtd = int(qtd)
        except: qtd = 1
        for etapa, d_ini, h_ini, d_fim, h_fim, quem_col in ETAPAS_PROD:
            ini = _prod_parse_ini(row[d_ini], row[h_ini])
            fim = _prod_parse_fim(row[d_fim], row[h_fim], ini)
            if pd.isna(ini) or pd.isna(fim) or fim <= ini:
                ignorados += 1; continue
            for colab in _prod_split_colabs(row[quem_col]):
                eventos.append({'colaborador': colab, 'etapa': etapa,
                                'inicio': ini, 'fim': fim, 'data': ini.date(),
                                'mes': ini.strftime('%Y-%m'), 'qtd_saidas': qtd,
                                'dur_min': (fim - ini).total_seconds() / 60})
    return pd.DataFrame(eventos), ignorados

def _prod_fmt_pct(cell, value):
    s   = Side(style='thin', color='B8CCE4')
    brd = Border(left=s, right=s, top=s, bottom=s)
    if value >= 70:
        fill, fc = PatternFill(fill_type='solid', fgColor='C6EFCE'), '375623'
    elif value >= 40:
        fill, fc = PatternFill(fill_type='solid', fgColor='FFEB9C'), '7F6000'
    else:
        fill, fc = PatternFill(fill_type='solid', fgColor='FFC7CE'), '9C0006'
    cell.fill      = fill
    cell.font      = Font(bold=True, color=fc, name='Calibri', size=10)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border    = brd
    cell.number_format = '0.0'

def _prod_formatar_resumo(ws):
    cores_h = ['1F4E79','2F4F74','375623','9C0006','9C0006','1F4E79','1F4E79','375623','375623']
    for ci in range(1, ws.max_column + 1):
        cor = cores_h[ci - 1] if ci <= len(cores_h) else '1F4E79'
        c   = ws.cell(1, ci)
        c.fill      = PatternFill(fill_type='solid', fgColor=cor)
        c.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        s = Side(style='thin', color='B8CCE4')
        c.border    = Border(left=s, right=s, top=s, bottom=s)
    ws.row_dimensions[1].height = 30
    pct_col = next((ci for ci in range(1, ws.max_column + 1)
                    if '% Utilização' in str(ws.cell(1, ci).value)), None)
    pares = ['EEF2FA', 'F7F9FC']
    for ri in range(2, ws.max_row + 1):
        bg       = pares[(ri - 2) % 2]
        is_total = str(ws.cell(ri, 1).value) in ('TOTAL GERAL', 'TOTAL MÊS')
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(ri, ci)
            s    = Side(style='thin', color='B8CCE4')
            cell.border = Border(left=s, right=s, top=s, bottom=s)
            if is_total:
                cell.fill      = PatternFill(fill_type='solid', fgColor='1F4E79')
                cell.font      = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
            elif pct_col and ci == pct_col:
                try:   _prod_fmt_pct(cell, float(cell.value))
                except: pass
            else:
                cell.fill      = PatternFill(fill_type='solid', fgColor=bg)
                cell.font      = Font(name='Calibri', size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if isinstance(cell.value, float): cell.number_format = '#,##0.00'
        ws.row_dimensions[ri].height = 18
    _autofit(ws)

def _prod_formatar_ociosidade(ws):
    _formatar_aba_generica(ws, cor_header='9C0006')
    col_oc = next((ci for ci in range(1, ws.max_column + 1)
                   if 'Horas Ociosas' in str(ws.cell(1, ci).value)), None)
    if not col_oc: return
    for ri in range(2, ws.max_row + 1):
        cell = ws.cell(ri, col_oc)
        try:
            v = float(cell.value)
            if v > 6:
                cell.fill = PatternFill(fill_type='solid', fgColor='FFC7CE')
                cell.font = Font(bold=True, color='9C0006', name='Calibri', size=10)
            elif v > 3:
                cell.fill = PatternFill(fill_type='solid', fgColor='FFEB9C')
                cell.font = Font(bold=True, color='7F6000', name='Calibri', size=10)
        except (TypeError, ValueError):
            pass

def processar_produtividade(caminho, mes_filtro, log, _saida_override=None):
    try:
        log(f"📂 Carregando: {os.path.basename(caminho)}\n")
        # Detecta aba pela primeira coluna de etapa conhecida
        col_ref = ETAPAS_PROD[0][1]  # 'DATA ÍNICIO ACONSELHAMENTO'
        nome_aba, df_raw = _detectar_aba(caminho, [col_ref], log)
        if df_raw is None:
            # Fallback: tenta primeira aba
            xl = pd.ExcelFile(caminho)
            nome_aba = xl.sheet_names[0]
            df_raw = xl.parse(nome_aba, header=0)
            log(f"⚠️ Coluna '{col_ref}' não encontrada — usando primeira aba: '{nome_aba}'\n")
        else:
            df_raw = pd.read_excel(caminho, sheet_name=nome_aba, header=0)
        log(f"✅ {len(df_raw)} linhas encontradas.\n")
    except Exception as e:
        log(f"❌ Erro ao carregar: {e}\n"); return False

    log("🔄 Extraindo eventos por colaborador e etapa...\n")
    ev, ignorados = _prod_extrair_eventos(df_raw)
    if ev.empty:
        log("❌ Nenhum evento válido encontrado.\n"); return False
    log(f"✅ {len(ev)} eventos válidos | {ignorados} linhas ignoradas (datas inválidas).\n")
    log(f"👥 Colaboradores: {', '.join(sorted(ev['colaborador'].unique()))}\n")

    if mes_filtro:
        ev_filt = ev[ev['mes'] == mes_filtro]
        if ev_filt.empty:
            log(f"⚠️ Sem dados para '{mes_filtro}'. Disponíveis: {sorted(ev['mes'].unique())}\n")
            return False
        ev = ev_filt
        log(f"🔍 Filtrando mês: {mes_filtro}\n")

    log("📊 Calculando resumo mensal...\n")
    resumo_rows = []
    for (colab, mes), grp in ev.groupby(['colaborador', 'mes']):
        dias     = sorted(grp['data'].unique())
        h_ativas = sum(
            _prod_merge_min(list(zip(
                grp[grp['data'] == d]['inicio'],
                grp[grp['data'] == d]['fim']
            ))) / 60
            for d in dias)
        n_dias     = len(dias)
        h_turno    = n_dias * TURNO_H
        h_ociosas  = max(0, h_turno - h_ativas)
        utilizacao = round(min(100, h_ativas / h_turno * 100), 1) if h_turno > 0 else 0
        qtd_saidas = int(grp.drop_duplicates(['data', 'etapa', 'inicio'])['qtd_saidas'].sum())
        resumo_rows.append({
            'Colaborador':      colab,
            'Mês':              mes,
            'Dias Trabalhados': n_dias,
            'H Turno Total':    round(h_turno, 2),
            'H Ativas':         round(h_ativas, 2),
            'H Ociosas':        round(h_ociosas, 2),
            '% Utilização':     utilizacao,
            'Qtd. Saídas':      qtd_saidas,
            'Saídas / Hora':    round(qtd_saidas / h_ativas, 2) if h_ativas > 0 else 0,
        })
    df_resumo = pd.DataFrame(resumo_rows).sort_values(['Mês', 'Colaborador'])

    log("📊 Calculando ranking por etapa...\n")
    etapa_rows = []
    for (colab, etapa, mes), grp in ev.groupby(['colaborador', 'etapa', 'mes']):
        h = _prod_merge_min(list(zip(grp['inicio'], grp['fim']))) / 60
        etapa_rows.append({
            'Etapa':       etapa,
            'Colaborador': colab,
            'Mês':         mes,
            'H na Etapa':  round(h, 2),
            'Eventos':     len(grp),
            'Qtd. Saídas': int(grp['qtd_saidas'].sum()),
            'Saídas/Hora': round(grp['qtd_saidas'].sum() / h, 2) if h > 0 else 0,
        })
    df_etapa = pd.DataFrame(etapa_rows).sort_values(
        ['Etapa', 'Mês', 'H na Etapa'], ascending=[True, True, False])

    log("📊 Calculando ociosidade diária...\n")
    ocio_rows = []
    for (colab, dt), grp in ev.groupby(['colaborador', 'data']):
        h_ativas  = _prod_merge_min(list(zip(grp['inicio'], grp['fim']))) / 60
        h_ociosas = max(0, TURNO_H - h_ativas)
        ocio_rows.append({
            'Colaborador':   colab,
            'Data':          dt.strftime('%d/%m/%Y'),
            'Mês':           dt.strftime('%Y-%m'),
            'H Ativas':      round(h_ativas, 2),
            'Horas Ociosas': round(h_ociosas, 2),
            '% Utilização':  round(min(100, h_ativas / TURNO_H * 100), 1),
            'Qtd. Saídas':   int(grp['qtd_saidas'].sum()),
        })
    df_ocio = pd.DataFrame(ocio_rows).sort_values(['Colaborador', 'Data'])

    saida = _saida_override if _saida_override else \
        _caminho_saida(PASTA_PRODUTIVIDADE, 'relatorio_produtividade')
    if saida is None: return False
    log(f"💾 Gerando Excel: {os.path.basename(saida)}...\n")

    with pd.ExcelWriter(saida, engine='openpyxl') as writer:
        df_resumo.to_excel(writer, sheet_name='Resumo Mensal',     index=False)
        df_etapa.to_excel(writer,  sheet_name='Ranking por Etapa', index=False)
        df_ocio.to_excel(writer,   sheet_name='Ociosidade Diária', index=False)

    wb = load_workbook(saida)
    _prod_formatar_resumo(wb['Resumo Mensal'])
    _formatar_aba_generica(wb['Ranking por Etapa'], cor_header='375623')
    _prod_formatar_ociosidade(wb['Ociosidade Diária'])
    wb.save(saida)

    log(f"\n{'─'*52}\n")
    log(f"📋 {len(df_resumo)} registros no Resumo Mensal\n")
    log(f"📋 {len(df_etapa)} registros no Ranking por Etapa\n")
    log(f"📋 {len(df_ocio)} dias na Ociosidade Diária\n")
    top3 = df_resumo.groupby('Colaborador')['Qtd. Saídas'].sum().nlargest(3)
    log(f"\n🏆 Top 3 por volume total de saídas:\n")
    for nome, qtd in top3.items():
        log(f"   {nome}: {qtd:,} saídas\n")
    log(f"\n✅ Salvo em: {saida}\n")
    log("🎉 Concluído!\n")
    return saida

def run_produtividade(caminho, mes_filtro, log, mes_ref=None):
    pasta = PASTA_PRODUTIVIDADE
    # mes_ref (MM-AAAA) tem prioridade para nomear o arquivo
    if mes_ref:
        try:
            mm, aaaa = mes_ref.split('-')
            sufixo  = f'{int(mm):02d}-{aaaa}'
            ano_str = aaaa
        except Exception:
            m, a    = _mes_ano_referencia()
            sufixo  = f'{m:02d}-{a}'
            ano_str = str(a)
    elif mes_filtro:
        try:
            ano_f, mes_f = mes_filtro.split('-')
            sufixo  = f'{int(mes_f):02d}-{ano_f}'
            ano_str = ano_f
        except Exception:
            sufixo  = mes_filtro
            ano_str = str(date.today().year)
    else:
        m, a    = _mes_ano_referencia()
        sufixo  = f'{m:02d}-{a}'
        ano_str = str(a)

    subpasta = os.path.join(pasta, ano_str)
    try:
        os.makedirs(subpasta, exist_ok=True)
    except Exception:
        log(f"❌ Não foi possível criar a pasta de saída em:\n   {subpasta}\n")
        return False
    nome_arq = f'relatorio_produtividade_{sufixo}.xlsx'
    saida    = os.path.join(subpasta, nome_arq)
    contador = 1
    while os.path.exists(saida):
        saida = os.path.join(subpasta, f'relatorio_produtividade_{sufixo}_{contador}.xlsx')
        contador += 1

    resultado = processar_produtividade(caminho, mes_filtro, log, _saida_override=saida)
    if resultado:
        _registrar_extracao('Produtividade')
        _atualizar_historico('produtividade', saida, mes_ref or sufixo, log)
    return resultado

# ============================================================
# APP PRINCIPAL COM SIDEBAR
# ============================================================
NOME_ABA_PEDIDOS = ''  # preenchido pelo usuário na interface ou nas Configurações

# ============================================================
# MÓDULO 6 — DASHBOARD CONSOLIDADO
# ============================================================
COR_ACCENT6 = '#8b5cf6'  # violeta — dashboard

RELATORIOS_CONFIG_DASH = {
    'pedidos':       {'arquivo_base': 'planilha_de_resultados_detalhada', 'pasta_key': 'pasta_pedidos',      'pasta_default': None},
    'fretes':        {'arquivo_base': 'relatorio_fretes',                 'pasta_key': 'pasta_fretes',        'pasta_default': None},
    'armazenagem':   {'arquivo_base': 'relatorio_armazenagem',            'pasta_key': 'pasta_armazenagem',   'pasta_default': None},
    'estoque':       {'arquivo_base': 'relatorio_estoque',                'pasta_key': 'pasta_estoque',       'pasta_default': None},
    'produtividade': {'arquivo_base': 'relatorio_produtividade',          'pasta_key': 'pasta_produtividade', 'pasta_default': None},
}

def _dash_resolver_pasta(modulo_key):
    """Retorna a pasta de saída do módulo diretamente das constantes."""
    _MAP = {
        'pedidos':       PASTA_PEDIDOS,
        'fretes':        PASTA_FRETES,
        'armazenagem':   PASTA_ARMAZENAGEM,
        'estoque':       PASTA_ESTOQUE,
        'produtividade': PASTA_PRODUTIVIDADE,
    }
    return _MAP.get(modulo_key, '')

def _dash_encontrar_arquivo(modulo_key, mes_ano_str):
    """
    Busca o arquivo do módulo cujo nome contém o padrão MM-AAAA.
    mes_ano_str: '03-2025' (MM-AAAA)
    Retorna caminho completo ou None.
    """
    pasta  = _dash_resolver_pasta(modulo_key)
    base   = RELATORIOS_CONFIG_DASH[modulo_key]['arquivo_base']
    if not pasta or not os.path.isdir(pasta):
        return None

    candidatos = []
    for dirpath, _, files in os.walk(pasta):
        for f in files:
            if f.endswith('.xlsx') and base in f and mes_ano_str in f:
                candidatos.append(os.path.join(dirpath, f))

    if not candidatos:
        return None
    # Se houver mais de um (raro), pega o mais recente
    return sorted(candidatos, key=os.path.getmtime, reverse=True)[0]

# ── Leitores por módulo ────────────────────────────────────────────────────────

def _dash_ler_pedidos(caminho):
    """
    Retorna dict com:
      sla_pct: float (0–100)
      total_ordens: int
      excedidas: int
      por_depositante: list[{nome, sla, total}]
    """
    try:
        df = pd.read_excel(caminho, sheet_name='Resumo Por Depositante', index_col=0)
        df = df[df.index != 'TOTAL GERAL']
        # SLA % pode estar como fração (0–1) ou percentual (0–100)
        sla_col = df['SLA %'] if 'SLA %' in df.columns else df.iloc[:, -1]
        sla_vals = pd.to_numeric(sla_col, errors='coerce').fillna(0)
        if sla_vals.max() <= 1.0:
            sla_vals = sla_vals * 100
        total_geral = pd.to_numeric(df.get('Total Geral', df.iloc[:, -2]), errors='coerce').fillna(0)
        excedidas   = pd.to_numeric(df.get('Excedido D+1', 0), errors='coerce').fillna(0)
        total       = int(total_geral.sum())
        exc_total   = int(excedidas.sum())
        sla_global  = round((total - exc_total) / total * 100, 1) if total > 0 else 0.0
        por_dep = []
        for dep in df.index:
            t  = int(pd.to_numeric(total_geral.get(dep, 0) if hasattr(total_geral, 'get') else total_geral.loc[dep], errors='coerce') or 0)
            s  = round(float(sla_vals.loc[dep]), 1) if dep in sla_vals.index else 0.0
            por_dep.append({'nome': str(dep), 'sla': s, 'total': t})
        por_dep.sort(key=lambda x: x['sla'])
        return {'sla_pct': sla_global, 'total_ordens': total,
                'excedidas': exc_total, 'por_depositante': por_dep}
    except Exception as e:
        return {'erro': str(e)}

def _dash_ler_fretes(caminho):
    """
    Retorna dict com:
      total_frete: float
      por_remetente: list[{nome, valor}]
    """
    try:
        df = pd.read_excel(caminho, sheet_name='Consolidado')
        df = df[df['Remetente'].astype(str) != 'TOTAL GERAL']
        val_col = 'Valor Frete Total' if 'Valor Frete Total' in df.columns else df.columns[1]
        df[val_col] = pd.to_numeric(df[val_col], errors='coerce').fillna(0)
        total = round(df[val_col].sum(), 2)
        por_rem = [
            {'nome': str(r['Remetente']), 'valor': round(float(r[val_col]), 2)}
            for _, r in df.iterrows()
        ]
        por_rem.sort(key=lambda x: x['valor'], reverse=True)
        return {'total_frete': total, 'por_remetente': por_rem}
    except Exception as e:
        return {'erro': str(e)}

def _dash_ler_armazenagem(caminho):
    """
    Retorna dict com:
      total_armazenagem: float
      por_cliente: list[{nome, valor}]
    """
    try:
        df = pd.read_excel(caminho, sheet_name='Armazenagem')
        df = df[df['Cliente'].astype(str).str.upper() != 'TOTAL MÊS']
        df['Soma Armazenagem'] = pd.to_numeric(df['Soma Armazenagem'], errors='coerce').fillna(0)
        total = round(df['Soma Armazenagem'].sum(), 2)
        por_cli = [
            {'nome': str(r['Cliente']), 'valor': round(float(r['Soma Armazenagem']), 2)}
            for _, r in df.iterrows()
        ]
        por_cli.sort(key=lambda x: x['valor'], reverse=True)
        return {'total_armazenagem': total, 'por_cliente': por_cli}
    except Exception as e:
        return {'erro': str(e)}

def _dash_ler_produtividade(caminho):
    """
    Retorna dict com:
      media_utilizacao: float (0–100)
      top_colaboradores: list[{nome, pct}] top 5
    """
    try:
        df = pd.read_excel(caminho, sheet_name='Resumo Mensal')
        pct_col = '% Utilização' if '% Utilização' in df.columns else df.columns[-1]
        df[pct_col] = pd.to_numeric(
            df[pct_col].astype(str).str.replace('%','').str.strip(), errors='coerce'
        ).fillna(0)
        df = df[df['Colaborador'].astype(str).str.upper() != 'TOTAL']
        media = round(df[pct_col].mean(), 1)
        top = df.nlargest(5, pct_col)[['Colaborador', pct_col]].copy()
        top_list = [{'nome': str(r['Colaborador']), 'pct': round(float(r[pct_col]), 1)}
                    for _, r in top.iterrows()]
        return {'media_utilizacao': media, 'top_colaboradores': top_list}
    except Exception as e:
        return {'erro': str(e)}

def _dash_ler_estoque(caminho):
    """
    Retorna dict com:
      pico_total_m3: float
      por_cliente: list[{nome, pico_m3, pico_m2}] top 5
    """
    try:
        df = pd.read_excel(caminho, sheet_name='Resumo por Cliente')
        df = df[df['Cliente'].astype(str).str.upper() != 'TOTAL GERAL']
        df['Pico Volume m³'] = pd.to_numeric(df['Pico Volume m³'], errors='coerce').fillna(0)
        df['Pico Área m²']   = pd.to_numeric(df['Pico Área m²'],   errors='coerce').fillna(0)
        pico_total = round(df['Pico Volume m³'].sum(), 2)
        top = df.nlargest(5, 'Pico Volume m³')[['Cliente','Pico Volume m³','Pico Área m²']].copy()
        top_list = [
            {'nome': str(r['Cliente']),
             'pico_m3':   round(float(r['Pico Volume m³']), 2),
             'pico_m2':   round(float(r['Pico Área m²']),   2)}
            for _, r in top.iterrows()
        ]
        return {'pico_total_m3': pico_total, 'por_cliente': top_list}
    except Exception as e:
        return {'erro': str(e)}

def _dash_coletar_dados(mes_ano_str, overrides, log):
    """
    mes_ano_str: 'MM-AAAA'
    overrides: dict {modulo_key: caminho_manual}
    Retorna dict {modulo_key: dados_dict}
    """
    leitores = {
        'pedidos':       _dash_ler_pedidos,
        'fretes':        _dash_ler_fretes,
        'armazenagem':   _dash_ler_armazenagem,
        'produtividade': _dash_ler_produtividade,
        'estoque':       _dash_ler_estoque,
    }
    resultado = {}
    for key, leitor in leitores.items():
        caminho = overrides.get(key) or _dash_encontrar_arquivo(key, mes_ano_str)
        if not caminho:
            log(f"⚠️  {key.capitalize()}: nenhum arquivo encontrado para {mes_ano_str} — seção será omitida.\n")
            resultado[key] = None
            continue
        log(f"📂 {key.capitalize()}: {os.path.basename(caminho)}\n")
        dados = leitor(caminho)
        if 'erro' in dados:
            log(f"⚠️  {key.capitalize()}: erro ao ler → {dados['erro']}\n")
            resultado[key] = None
        else:
            resultado[key] = dados
    return resultado

# ── Exportação Excel do Dashboard ─────────────────────────────────────────────

def _dash_exportar_excel(dados, mes_ano_str, pasta_saida, log):
    """Gera o Excel do Dashboard consolidado."""
    mm, aaaa = mes_ano_str.split('-')
    nome_arq  = f'dashboard_{mm}-{aaaa}.xlsx'
    subpasta  = os.path.join(pasta_saida, aaaa)
    try:
        os.makedirs(subpasta, exist_ok=True)
    except Exception:
        log(f"❌ Não foi possível criar a pasta de saída em:\n   {subpasta}\n")
        return None
    caminho   = os.path.join(subpasta, nome_arq)

    # Evita sobrescrever
    cnt = 1
    while os.path.exists(caminho):
        caminho = os.path.join(subpasta, f'dashboard_{mm}-{aaaa}_{cnt}.xlsx')
        cnt += 1

    wb = Workbook()
    wb.remove(wb.active)

    AZUL      = '1F3864'
    VIOL      = '4C1D95'
    VERD_ESC  = '14532D'
    VERM_ESC  = '7F1D1D'
    AMBER_ESC = '78350F'
    CINZA     = 'D6DCE4'
    LINHA_PAR = 'F2F7FB'

    def _ws(nome, cor_header):
        ws = wb.create_sheet(nome)
        return ws, cor_header

    def _cabecalho_ws(ws, colunas, cor_bg):
        for i, col in enumerate(colunas, 1):
            c = ws.cell(1, i, col)
            _fmt_cabecalho(c, bg=cor_bg)
        ws.row_dimensions[1].height = 28

    def _linha_ws(ws, row_idx, valores, bg='FFFFFF'):
        for i, v in enumerate(valores, 1):
            c = ws.cell(row_idx, i, v)
            c.fill = _fill_excel(bg)
            c.font = Font(name='Calibri', size=10)
            c.alignment = Alignment(vertical='center')
            c.border = _borda_excel()

    def _total_ws(ws, row_idx, valores):
        for i, v in enumerate(valores, 1):
            _fmt_total(ws.cell(row_idx, i, v))

    # ── Aba 1: Resumo Geral ──────────────────────────────────────────────────
    ws_res = wb.create_sheet('Resumo Geral')
    _cabecalho_ws(ws_res, ['Módulo', 'KPI Principal', 'Valor', 'Observação'], AZUL)

    linhas_resumo = []
    if dados.get('pedidos'):
        d = dados['pedidos']
        linhas_resumo.append(('📦 Pedidos', 'SLA %',
                               f"{d['sla_pct']:.1f}%",
                               f"{d['total_ordens']} ordens · {d['excedidas']} excedidas"))
    if dados.get('fretes'):
        d = dados['fretes']
        linhas_resumo.append(('🚚 Fretes', 'Custo Total Frete',
                               f"R$ {d['total_frete']:,.2f}",
                               f"{len(d['por_remetente'])} remetentes"))
    if dados.get('armazenagem'):
        d = dados['armazenagem']
        linhas_resumo.append(('🏭 Armazenagem', 'Faturamento Total',
                               f"R$ {d['total_armazenagem']:,.2f}",
                               f"{len(d['por_cliente'])} clientes"))
    if dados.get('produtividade'):
        d = dados['produtividade']
        linhas_resumo.append(('👥 Produtividade', 'Utilização Média',
                               f"{d['media_utilizacao']:.1f}%",
                               f"Top: {d['top_colaboradores'][0]['nome']} ({d['top_colaboradores'][0]['pct']:.1f}%)" if d['top_colaboradores'] else ''))
    if dados.get('estoque'):
        d = dados['estoque']
        linhas_resumo.append(('📋 Estoque', 'Pico Total m³',
                               f"{d['pico_total_m3']:,.2f} m³",
                               f"Top: {d['por_cliente'][0]['nome']}" if d['por_cliente'] else ''))

    for i, (mod, kpi, val, obs) in enumerate(linhas_resumo, 2):
        bg = LINHA_PAR if i % 2 == 0 else 'FFFFFF'
        _linha_ws(ws_res, i, [mod, kpi, val, obs], bg)
        # Colorir coluna Valor por módulo
        ws_res.cell(i, 3).font = Font(bold=True, name='Calibri', size=11, color=AZUL)
    _autofit(ws_res)

    # ── Aba 2: Pedidos — SLA por Depositante ────────────────────────────────
    if dados.get('pedidos'):
        d  = dados['pedidos']
        ws_p, _ = _ws('Pedidos — SLA', AZUL)
        _cabecalho_ws(ws_p, ['Depositante', 'Total Ordens', 'SLA %', 'Status'], AZUL)
        for i, dep in enumerate(d['por_depositante'], 2):
            bg = LINHA_PAR if i % 2 == 0 else 'FFFFFF'
            status = '✅ OK' if dep['sla'] >= 90 else ('⚠️ Atenção' if dep['sla'] >= 70 else '❌ Crítico')
            _linha_ws(ws_p, i, [dep['nome'], dep['total'], f"{dep['sla']:.1f}%", status], bg)
            # Colorir SLA
            sla_c = ws_p.cell(i, 3)
            cor_t = '375623' if dep['sla'] >= 90 else ('7F6000' if dep['sla'] >= 70 else '9C0006')
            cor_b = 'C6EFCE' if dep['sla'] >= 90 else ('FFEB9C' if dep['sla'] >= 70 else 'FFC7CE')
            sla_c.fill = _fill_excel(cor_b)
            sla_c.font = Font(bold=True, name='Calibri', size=10, color=cor_t)
        _total_ws(ws_p, len(d['por_depositante']) + 2,
                  ['TOTAL', d['total_ordens'], f"{d['sla_pct']:.1f}%", ''])
        _autofit(ws_p)

    # ── Aba 3: Fretes — Custo por Remetente ─────────────────────────────────
    if dados.get('fretes'):
        d = dados['fretes']
        ws_f, _ = _ws('Fretes — Custos', VIOL)
        _cabecalho_ws(ws_f, ['Remetente', 'Valor Frete Total', '% do Total'], VIOL)
        for i, rem in enumerate(d['por_remetente'], 2):
            bg   = LINHA_PAR if i % 2 == 0 else 'FFFFFF'
            pct  = round(rem['valor'] / d['total_frete'] * 100, 1) if d['total_frete'] > 0 else 0
            _linha_ws(ws_f, i, [rem['nome'], rem['valor'], f"{pct:.1f}%"], bg)
            ws_f.cell(i, 2).number_format = 'R$ #,##0.00'
        _total_ws(ws_f, len(d['por_remetente']) + 2, ['TOTAL GERAL', d['total_frete'], '100%'])
        ws_f.cell(len(d['por_remetente']) + 2, 2).number_format = 'R$ #,##0.00'
        _autofit(ws_f)

    # ── Aba 4: Armazenagem — Faturamento por Cliente ─────────────────────────
    if dados.get('armazenagem'):
        d = dados['armazenagem']
        ws_a, _ = _ws('Armazenagem — Faturamento', VERD_ESC)
        _cabecalho_ws(ws_a, ['Cliente', 'Soma Armazenagem', '% do Total'], VERD_ESC)
        for i, cli in enumerate(d['por_cliente'], 2):
            bg  = LINHA_PAR if i % 2 == 0 else 'FFFFFF'
            pct = round(cli['valor'] / d['total_armazenagem'] * 100, 1) if d['total_armazenagem'] > 0 else 0
            _linha_ws(ws_a, i, [cli['nome'], cli['valor'], f"{pct:.1f}%"], bg)
            ws_a.cell(i, 2).number_format = 'R$ #,##0.00'
        _total_ws(ws_a, len(d['por_cliente']) + 2, ['TOTAL GERAL', d['total_armazenagem'], '100%'])
        ws_a.cell(len(d['por_cliente']) + 2, 2).number_format = 'R$ #,##0.00'
        _autofit(ws_a)

    # ── Aba 5: Produtividade — Top Colaboradores ─────────────────────────────
    if dados.get('produtividade'):
        d = dados['produtividade']
        ws_pr, _ = _ws('Produtividade — Top', AMBER_ESC)
        _cabecalho_ws(ws_pr, ['Posição', 'Colaborador', '% Utilização', 'Status'], AMBER_ESC)
        for i, col in enumerate(d['top_colaboradores'], 2):
            bg     = LINHA_PAR if i % 2 == 0 else 'FFFFFF'
            status = '🏆 Top' if i == 2 else ('⭐' if i <= 4 else '')
            _linha_ws(ws_pr, i, [f'{i-1}º', col['nome'], f"{col['pct']:.1f}%", status], bg)
            pct_c = ws_pr.cell(i, 3)
            cor_t = '14532D' if col['pct'] >= 80 else ('78350F' if col['pct'] >= 60 else '7F1D1D')
            cor_b = 'D1FAE5' if col['pct'] >= 80 else ('FEF3C7' if col['pct'] >= 60 else 'FEE2E2')
            pct_c.fill = _fill_excel(cor_b)
            pct_c.font = Font(bold=True, name='Calibri', size=10, color=cor_t)
        _total_ws(ws_pr, len(d['top_colaboradores']) + 2,
                  ['', 'Média Geral', f"{d['media_utilizacao']:.1f}%", ''])
        _autofit(ws_pr)

    # ── Aba 6: Estoque — Pico por Cliente ────────────────────────────────────
    if dados.get('estoque'):
        d = dados['estoque']
        ws_e, _ = _ws('Estoque — Pico', VERM_ESC)
        _cabecalho_ws(ws_e, ['Posição', 'Cliente', 'Pico Volume m³', 'Pico Área m²'], VERM_ESC)
        for i, cli in enumerate(d['por_cliente'], 2):
            bg = LINHA_PAR if i % 2 == 0 else 'FFFFFF'
            _linha_ws(ws_e, i, [f'{i-1}º', cli['nome'], cli['pico_m3'], cli['pico_m2']], bg)
            ws_e.cell(i, 3).number_format = '#,##0.00'
            ws_e.cell(i, 4).number_format = '#,##0.00'
        _total_ws(ws_e, len(d['por_cliente']) + 2,
                  ['', 'TOTAL', d['pico_total_m3'], ''])
        ws_e.cell(len(d['por_cliente']) + 2, 3).number_format = '#,##0.00'
        _autofit(ws_e)

    wb.save(caminho)
    log(f"✅ Dashboard salvo: {caminho}\n")
    return caminho

def _consolidar_historico(dados, mes_ano_str, log):
    """
    Grava os KPIs do mês em App\\Consolidados\\consolidado_AAAA.xlsx.
    Cada mês ocupa uma aba (ex: Jan-2026). Se já existir, substitui.
    Estrutura da aba: Resumo Geral + tabelas detalhadas de cada módulo.
    """
    mm, aaaa = mes_ano_str.split('-')
    _MESES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                 'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
    nome_aba  = f'{_MESES_PT[int(mm)-1][:3]}-{aaaa}'   # ex: Mar-2026

    pasta = PASTA_CONSOLIDADOS
    try:
        os.makedirs(pasta, exist_ok=True)
    except Exception:
        log(f"⚠️  Não foi possível criar a pasta de consolidados: {pasta}\n")
        return

    arq_consolidado = os.path.join(pasta, f'consolidado_{aaaa}.xlsx')

    # Carrega workbook existente ou cria novo
    if os.path.isfile(arq_consolidado):
        try:
            wb = load_workbook(arq_consolidado)
        except Exception as e:
            log(f"⚠️  Erro ao abrir consolidado existente: {e}\n")
            wb = Workbook(); wb.remove(wb.active)
    else:
        wb = Workbook(); wb.remove(wb.active)

    # Remove aba do mês se já existir (substituição)
    if nome_aba in wb.sheetnames:
        del wb[nome_aba]
        log(f"♻️  Aba '{nome_aba}' substituída no consolidado.\n")

    ws = wb.create_sheet(nome_aba)

    # ── Cores ────────────────────────────────────────────────────────────
    AZUL      = '1F3864'
    VIOL      = '4C1D95'
    VERD_ESC  = '14532D'
    VERM_ESC  = '7F1D1D'
    AMBER_ESC = '78350F'
    LINHA_PAR = 'F2F7FB'

    linha = 1  # cursor de linha atual na aba

    def _titulo_secao(ws, texto, cor, linha):
        """Escreve um título de seção em negrito com fundo colorido."""
        c = ws.cell(linha, 1, texto)
        c.fill = _fill_excel(cor)
        c.font = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[linha].height = 22
        return linha + 1

    def _header(ws, colunas, cor, linha):
        for j, col in enumerate(colunas, 1):
            c = ws.cell(linha, j, col)
            _fmt_cabecalho(c, bg=cor)
        ws.row_dimensions[linha].height = 20
        return linha + 1

    def _row(ws, valores, linha, par=False):
        bg = LINHA_PAR if par else 'FFFFFF'
        for j, v in enumerate(valores, 1):
            c = ws.cell(linha, j, v)
            c.fill = _fill_excel(bg)
            c.font = Font(name='Calibri', size=10)
            c.alignment = Alignment(vertical='center')
            c.border = _borda_excel()
        return linha + 1

    def _total(ws, valores, linha):
        for j, v in enumerate(valores, 1):
            _fmt_total(ws.cell(linha, j, v))
        return linha + 1

    # ── Resumo Geral ─────────────────────────────────────────────────────
    linha = _titulo_secao(ws, f'📊  Resumo Geral — {nome_aba}', AZUL, linha)
    linha = _header(ws, ['Módulo', 'KPI Principal', 'Valor', 'Observação'], AZUL, linha)
    resumo_linhas = []
    if dados.get('pedidos'):
        d = dados['pedidos']
        resumo_linhas.append(('📦 Pedidos', 'SLA %', f"{d['sla_pct']:.1f}%",
                               f"{d['total_ordens']} ordens · {d['excedidas']} excedidas"))
    if dados.get('fretes'):
        d = dados['fretes']
        resumo_linhas.append(('🚚 Fretes', 'Custo Total', f"R$ {d['total_frete']:,.2f}",
                               f"{len(d['por_remetente'])} remetentes"))
    if dados.get('armazenagem'):
        d = dados['armazenagem']
        resumo_linhas.append(('🏭 Armazenagem', 'Faturamento', f"R$ {d['total_armazenagem']:,.2f}",
                               f"{len(d['por_cliente'])} clientes"))
    if dados.get('produtividade'):
        d = dados['produtividade']
        top = d['top_colaboradores'][0]['nome'] if d['top_colaboradores'] else '—'
        resumo_linhas.append(('👥 Produtividade', 'Utilização Média',
                               f"{d['media_utilizacao']:.1f}%", f"Top: {top}"))
    if dados.get('estoque'):
        d = dados['estoque']
        top = d['por_cliente'][0]['nome'] if d['por_cliente'] else '—'
        resumo_linhas.append(('📋 Estoque', 'Pico Total m³',
                               f"{d['pico_total_m3']:,.2f} m³", f"Top: {top}"))
    for i, vals in enumerate(resumo_linhas):
        linha = _row(ws, vals, linha, par=(i % 2 == 0))
    linha += 1  # espaço

    # ── Pedidos ──────────────────────────────────────────────────────────
    if dados.get('pedidos'):
        d = dados['pedidos']
        linha = _titulo_secao(ws, '📦  Pedidos — SLA por Depositante', AZUL, linha)
        linha = _header(ws, ['Depositante', 'Total Ordens', 'SLA %', 'Status'], AZUL, linha)
        for i, dep in enumerate(d['por_depositante']):
            status = '✅ OK' if dep['sla'] >= 90 else ('⚠️ Atenção' if dep['sla'] >= 70 else '❌ Crítico')
            linha = _row(ws, [dep['nome'], dep['total'], f"{dep['sla']:.1f}%", status],
                         linha, par=(i % 2 == 0))
        linha = _total(ws, ['TOTAL', d['total_ordens'], f"{d['sla_pct']:.1f}%", ''], linha)
        linha += 1

    # ── Fretes ───────────────────────────────────────────────────────────
    if dados.get('fretes'):
        d = dados['fretes']
        linha = _titulo_secao(ws, '🚚  Fretes — Custo por Remetente', VIOL, linha)
        linha = _header(ws, ['Remetente', 'Valor Frete Total', '% do Total'], VIOL, linha)
        for i, rem in enumerate(d['por_remetente']):
            pct = round(rem['valor'] / d['total_frete'] * 100, 1) if d['total_frete'] else 0
            linha = _row(ws, [rem['nome'], f"R$ {rem['valor']:,.2f}", f"{pct:.1f}%"],
                         linha, par=(i % 2 == 0))
        linha = _total(ws, ['TOTAL GERAL', f"R$ {d['total_frete']:,.2f}", '100%'], linha)
        linha += 1

    # ── Armazenagem ──────────────────────────────────────────────────────
    if dados.get('armazenagem'):
        d = dados['armazenagem']
        linha = _titulo_secao(ws, '🏭  Armazenagem — Faturamento por Cliente', VERD_ESC, linha)
        linha = _header(ws, ['Cliente', 'Soma Armazenagem', '% do Total'], VERD_ESC, linha)
        for i, cli in enumerate(d['por_cliente']):
            pct = round(cli['valor'] / d['total_armazenagem'] * 100, 1) if d['total_armazenagem'] else 0
            linha = _row(ws, [cli['nome'], f"R$ {cli['valor']:,.2f}", f"{pct:.1f}%"],
                         linha, par=(i % 2 == 0))
        linha = _total(ws, ['TOTAL GERAL', f"R$ {d['total_armazenagem']:,.2f}", '100%'], linha)
        linha += 1

    # ── Produtividade ─────────────────────────────────────────────────────
    if dados.get('produtividade'):
        d = dados['produtividade']
        linha = _titulo_secao(ws, '👥  Produtividade — Top Colaboradores', AMBER_ESC, linha)
        linha = _header(ws, ['Posição', 'Colaborador', '% Utilização', 'Status'], AMBER_ESC, linha)
        for i, col in enumerate(d['top_colaboradores']):
            status = '🏆 Top' if i == 0 else ('⭐' if i < 3 else '')
            linha = _row(ws, [f'{i+1}º', col['nome'], f"{col['pct']:.1f}%", status],
                         linha, par=(i % 2 == 0))
        linha = _total(ws, ['', 'Média Geral', f"{d['media_utilizacao']:.1f}%", ''], linha)
        linha += 1

    # ── Estoque ──────────────────────────────────────────────────────────
    if dados.get('estoque'):
        d = dados['estoque']
        linha = _titulo_secao(ws, '📋  Estoque — Pico por Cliente', VERM_ESC, linha)
        linha = _header(ws, ['Posição', 'Cliente', 'Pico Volume m³', 'Pico Área m²'], VERM_ESC, linha)
        for i, cli in enumerate(d['por_cliente']):
            linha = _row(ws, [f'{i+1}º', cli['nome'],
                               f"{cli['pico_m3']:,.2f}", f"{cli['pico_m2']:,.2f}"],
                         linha, par=(i % 2 == 0))
        linha = _total(ws, ['', 'TOTAL', f"{d['pico_total_m3']:,.2f}", ''], linha)

    # Autofit colunas
    _autofit(ws)

    # Ordena abas por mês cronologicamente
    _ORDEM_MESES = ['Jan','Fev','Mar','Abr','Mai','Jun',
                    'Jul','Ago','Set','Out','Nov','Dez']
    def _sort_key(nome_aba):
        partes = nome_aba.split('-')
        if len(partes) == 2:
            mes_abr, ano = partes
            try:
                return (int(ano), _ORDEM_MESES.index(mes_abr))
            except (ValueError, IndexError):
                return (9999, 99)
        return (9999, 99)

    abas_ordenadas = sorted(wb.sheetnames, key=_sort_key)
    for i, nome in enumerate(abas_ordenadas):
        wb[nome].sheet_properties.tabColor = None
        wb.move_sheet(nome, offset=i - wb.sheetnames.index(nome))

    try:
        wb.save(arq_consolidado)
        log(f"📁 Consolidado atualizado: {os.path.basename(arq_consolidado)} (aba '{nome_aba}')\n")
        _gerar_pbids(log)
    except Exception as e:
        log(f"⚠️  Erro ao salvar consolidado: {e}\n")


def run_dashboard(mes_ano_str, overrides, pasta_saida, log):
    """Ponto de entrada: coleta dados, exporta Excel, retorna caminho do arquivo."""
    log(f"🔍 Buscando arquivos para {mes_ano_str}...\n{'─'*48}\n")
    dados = _dash_coletar_dados(mes_ano_str, overrides, log)

    modulos_ok = [k for k, v in dados.items() if v is not None]
    if not modulos_ok:
        log("❌ Nenhum arquivo encontrado. Verifique as pastas nas Configurações.\n")
        return None

    log(f"\n📊 Gerando Dashboard ({len(modulos_ok)}/5 módulos)...\n")
    caminho = _dash_exportar_excel(dados, mes_ano_str, pasta_saida, log)
    if caminho:
        _consolidar_historico(dados, mes_ano_str, log)
    _registrar_extracao('Dashboard')
    return caminho

# ============================================================
# MÓDULO 7 — CAPACIDADE OPERACIONAL (Scraping ESL)
# ============================================================

def processar_cap_operacional(caminho_pdf, log):
    """
    Lê o PDF de Movimentação de Estoque (Kardex) exportado do ESL e retorna
    dict com dados para o relatório de Capacidade Operacional.

    Como extrair o PDF no ESL:
      Relatório → Movimentação de Estoque → Kardex (opção selecionada)
      Filtrar por empresa BAIA 4 e período desejado → Exportar como PDF
    """
    if _pdfplumber is None:
        log("❌ pdfplumber não instalado.\n   Execute: pip install pdfplumber\n")
        return None

    import re
    from collections import defaultdict

    log(f"📄 Lendo PDF: {os.path.basename(caminho_pdf)}...\n")

    depositante_atual = None
    produto_atual     = None
    registros         = []   # (os_num, depositante, produto, data)

    try:
        with _pdfplumber.open(caminho_pdf) as pdf:
            total_pags = len(pdf.pages)
            log(f"   {total_pags} página(s) encontrada(s).\n")

            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if not text:
                    continue

                for linha in text.splitlines():
                    # Detecta depositante no cabeçalho
                    dep = re.search(r'Depositante:\s+\d+\s+(.+?)\s+Usuário:', linha)
                    if dep:
                        depositante_atual = dep.group(1).strip()
                        continue

                    # Detecta produto atual
                    prod = re.match(r'Saldo Inicial do Produto\s+(.+?)\s+[\d,.]+$', linha)
                    if prod:
                        produto_atual = prod.group(1).strip()
                        continue

                    # Linha de saída (quantidade negativa) com nº OS
                    mov = re.match(
                        r'^(\d{2}/\d{2}/\d{4})\S+\s+\S+\s+(\d{3}\.\d{3})\s+.*?-\d+,\d+\s*$',
                        linha
                    )
                    if mov and depositante_atual and produto_atual:
                        data, os_num = mov.groups()
                        registros.append((os_num, depositante_atual, produto_atual, data))

    except Exception as e:
        log(f"❌ Erro ao ler PDF: {e}\n")
        return None

    if not registros:
        log("⚠️  Nenhuma movimentação de saída encontrada no PDF.\n"
            "   Verifique se o relatório é de Kardex com movimentações do período.\n")
        return None

    log(f"✅ {len(registros)} linhas de saída encontradas.\n")

    # Agrupa: (OS, depositante) → set de produtos únicos (SKUs)
    grupos  = defaultdict(set)
    datas   = defaultdict(set)
    for os_num, dep, prod, data in registros:
        grupos[(os_num, dep)].add(prod)
        datas[os_num].add(data)

    # Monta lista de OS com métricas
    os_lista = []
    for (os_num, dep), prods in grupos.items():
        data_os = sorted(datas[os_num])[0]
        os_lista.append({
            'os':          os_num,
            'depositante': dep,
            'data':        data_os,
            'qtd_skus':    len(prods),
        })
    os_lista.sort(key=lambda x: (x['depositante'], x['os']))

    # Resumo por depositante
    dep_map = defaultdict(lambda: {'total_os': 0, 'total_skus': 0})
    for item in os_lista:
        dep_map[item['depositante']]['total_os']   += 1
        dep_map[item['depositante']]['total_skus'] += item['qtd_skus']

    resumo = []
    for dep, stats in sorted(dep_map.items()):
        total_os   = stats['total_os']
        total_skus = stats['total_skus']
        media      = round(total_skus / total_os, 1) if total_os else 0
        resumo.append({
            'depositante': dep,
            'total_os':    total_os,
            'total_skus':  total_skus,
            'media_skus':  media,
        })

    total_os_geral   = sum(r['total_os']   for r in resumo)
    total_skus_geral = sum(r['total_skus'] for r in resumo)
    media_geral      = round(total_skus_geral / total_os_geral, 1) if total_os_geral else 0

    log(f"📊 {total_os_geral} OS | {total_skus_geral} SKUs | Média: {media_geral} SKUs/OS\n")

    return {
        'os_lista':         os_lista,
        'resumo':           resumo,
        'total_os':         total_os_geral,
        'total_skus':       total_skus_geral,
        'media_geral':      media_geral,
    }


def run_cap_operacional_pdf(caminho_pdf, mes_ref, log,
                            limiar_media=3.0, limiar_alta=5.0):
    """
    Wrapper que processa o PDF e gera o relatório Excel.
    limiar_media: SKUs/OS mínimo para classificar como Média (padrão 3)
    limiar_alta:  SKUs/OS mínimo para classificar como Alta  (padrão 5)
    """
    dados = processar_cap_operacional(caminho_pdf, log)
    if not dados:
        return False

    saida = _caminho_saida(PASTA_CAP_OPERACIONAL, 'relatorio_cap_operacional', mes_ref=mes_ref)
    if not saida:
        log(f"❌ Não foi possível criar pasta de saída em:\n   {PASTA_CAP_OPERACIONAL}\n")
        return False

    log(f"💾 Gerando Excel: {os.path.basename(saida)}...\n")
    log(f"   Complexidade: Baixa < {limiar_media} SKUs/OS | Média {limiar_media}–{limiar_alta} | Alta ≥ {limiar_alta}\n")

    COR_CAP  = '9F1239'
    _DIAS_PT = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado', 'Domingo']

    def _complexidade(media):
        if media >= limiar_alta:
            return 'Alta'
        elif media >= limiar_media:
            return 'Média'
        return 'Baixa'

    def _dia_semana(data_str):
        try:
            d, m, a = data_str.split('/')
            from datetime import date as _date
            return _DIAS_PT[_date(int(a), int(m), int(d)).weekday()]
        except Exception:
            return ''

    wb = Workbook()
    wb.remove(wb.active)

    # ── Aba 1: Resumo por Depositante ────────────────────────────────────
    ws1 = wb.create_sheet('Resumo por Depositante')
    cols1 = ['Depositante', 'Total de OS', 'Total de SKUs', 'Média SKUs/OS', 'Complexidade']
    _fmt_cabecalho_ws(ws1, cols1, COR_CAP)

    for i, r in enumerate(dados['resumo']):
        media = r['media_skus']
        valores = [r['depositante'], r['total_os'], r['total_skus'], media, _complexidade(media)]
        bg = 'F2F7FB' if i % 2 == 0 else 'FFFFFF'
        for j, v in enumerate(valores, 1):
            c = ws1.cell(i + 2, j, v)
            c.fill      = _fill_excel(bg)
            c.font      = Font(name='Calibri', size=10)
            c.border    = _borda_excel()
            c.alignment = Alignment(vertical='center')

    total_row = len(dados['resumo']) + 2
    _fmt_total(ws1.cell(total_row, 1, 'TOTAL GERAL'))
    _fmt_total(ws1.cell(total_row, 2, dados['total_os']))
    _fmt_total(ws1.cell(total_row, 3, dados['total_skus']))
    _fmt_total(ws1.cell(total_row, 4, dados['media_geral']))
    _fmt_total(ws1.cell(total_row, 5, ''))
    _autofit(ws1)

    # ── Aba 2: Detalhe por OS ─────────────────────────────────────────────
    ws2 = wb.create_sheet('Detalhe por OS')
    cols2 = ['Nº OS', 'Depositante', 'Data', 'Dia da Semana', 'Qtd. SKUs', 'Complexidade']
    _fmt_cabecalho_ws(ws2, cols2, COR_CAP)

    for i, item in enumerate(dados['os_lista']):
        dia = _dia_semana(item['data'])
        valores = [item['os'], item['depositante'], item['data'],
                   dia, item['qtd_skus'], _complexidade(item['qtd_skus'])]
        bg = 'F2F7FB' if i % 2 == 0 else 'FFFFFF'
        for j, v in enumerate(valores, 1):
            c = ws2.cell(i + 2, j, v)
            c.fill      = _fill_excel(bg)
            c.font      = Font(name='Calibri', size=10)
            c.border    = _borda_excel()
            c.alignment = Alignment(vertical='center')
    _autofit(ws2)

    # ── Aba 3: Distribuição por Dia ───────────────────────────────────────
    ws3 = wb.create_sheet('Distribuição por Dia')
    cols3 = ['Data', 'Dia da Semana', 'Total de OS', 'Total de SKUs', 'Média SKUs/OS', 'Complexidade do Dia']
    _fmt_cabecalho_ws(ws3, cols3, COR_CAP)

    from collections import defaultdict
    dist_dia = defaultdict(lambda: {'os': set(), 'skus': 0})
    for item in dados['os_lista']:
        dist_dia[item['data']]['os'].add(item['os'])
        dist_dia[item['data']]['skus'] += item['qtd_skus']

    # Ordena por data
    def _sort_data(d):
        try:
            dd, mm, aa = d.split('/')
            return (int(aa), int(mm), int(dd))
        except Exception:
            return (0, 0, 0)

    for i, data in enumerate(sorted(dist_dia.keys(), key=_sort_data)):
        info      = dist_dia[data]
        total_os  = len(info['os'])
        total_sk  = info['skus']
        media_dia = round(total_sk / total_os, 1) if total_os else 0
        dia       = _dia_semana(data)
        valores   = [data, dia, total_os, total_sk, media_dia, _complexidade(media_dia)]
        bg = 'F2F7FB' if i % 2 == 0 else 'FFFFFF'
        for j, v in enumerate(valores, 1):
            c = ws3.cell(i + 2, j, v)
            c.fill      = _fill_excel(bg)
            c.font      = Font(name='Calibri', size=10)
            c.border    = _borda_excel()
            c.alignment = Alignment(vertical='center')

    # Total geral
    total_row3 = len(dist_dia) + 2
    _fmt_total(ws3.cell(total_row3, 1, 'TOTAL'))
    _fmt_total(ws3.cell(total_row3, 2, ''))
    _fmt_total(ws3.cell(total_row3, 3, dados['total_os']))
    _fmt_total(ws3.cell(total_row3, 4, dados['total_skus']))
    _fmt_total(ws3.cell(total_row3, 5, dados['media_geral']))
    _fmt_total(ws3.cell(total_row3, 6, ''))
    _autofit(ws3)

    wb.save(saida)

    _registrar_extracao('Capacidade Operacional')
    _atualizar_historico('cap_operacional', saida, mes_ref, log)
    log(f"✅ Relatório de Capacidade Operacional gerado!\n")
    return saida


def _fmt_cabecalho_ws(ws, colunas, cor_bg, start_row=1):
    """Formata cabeçalho de aba — wrapper de _fmt_cabecalho para múltiplas colunas."""
    for j, col in enumerate(colunas, 1):
        _fmt_cabecalho(ws.cell(start_row, j, col), bg=cor_bg)
    ws.row_dimensions[start_row].height = 28


# ============================================================
# MÓDULO 8 — RECEBIMENTOS E DEVOLUÇÕES
# ============================================================

def processar_recebimentos(caminho_arq, log):
    """
    Lê o relatório de NF de Entrada do ESL (XLS/HTML ou Excel .xlsx/.xls)
    e extrai os recebimentos válidos, excluindo ajustes e manuseios.

    Colunas esperadas (índice):
      0=Sequência, 2=Nota Fiscal, 6=Depositante, 8=Data Entrada,
      9=Hora Entrada, 14=Fornecedor, 19=Valor Nota, 23=Peso,
      25=Comentário Adicional, 26=Usuário Cadastro

    Tipos válidos: Entrada c/ NF, Entrada s/ NF, Devolução,
                   Retirada, Entrada c/ CT-e, Entrada c/ OC
    Ignorados:    Ajuste (qualquer), Manuseio
    """
    from collections import defaultdict
    import re as _re

    _DIAS_PT = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo']

    def _normalizar_data(val):
        """Converte qualquer formato de data para DD/MM/AAAA."""
        if val is None or str(val) == 'nan':
            return ''
        # datetime / date do pandas/Excel
        if hasattr(val, 'strftime'):
            return val.strftime('%d/%m/%Y')
        s = str(val).strip()
        # '2026-01-02 00:00:00' ou '2026-01-02'
        m = _re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
        if m:
            return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
        # já está em DD/MM/AAAA
        if _re.match(r'^\d{2}/\d{2}/\d{4}', s):
            return s[:10]
        return s

    def _normalizar_hora(val):
        """Retorna HH:MM ou vazio."""
        if val is None or str(val) == 'nan':
            return ''
        if hasattr(val, 'strftime'):
            return val.strftime('%H:%M')
        s = str(val).strip()
        m = _re.match(r'^(\d{1,2}):(\d{2})', s)
        if m:
            return f"{int(m.group(1)):02d}:{m.group(2)}"
        return ''

    def _dia_sem(data_dd_mm_aaaa):
        try:
            d, m, a = data_dd_mm_aaaa.split('/')
            from datetime import date as _dt
            return _DIAS_PT[_dt(int(a), int(m), int(d)).weekday()]
        except Exception:
            return ''

    ext = os.path.splitext(caminho_arq)[1].lower()
    log(f"📄 Lendo arquivo: {os.path.basename(caminho_arq)}...\n")

    # ── Leitura do arquivo ────────────────────────────────────────────────
    df = None
    try:
        df = pd.read_excel(caminho_arq)
        log(f"   {len(df)} registros encontrados.\n")
    except Exception as e:
        log(f"❌ Não foi possível ler o arquivo: {e}\n")
        return None

    if df is None or len(df) == 0:
        log("⚠️  Arquivo vazio ou sem dados.\n")
        return None

    # ── Classifica e filtra ───────────────────────────────────────────────
    def _str(val):
        return str(val).strip() if val is not None and str(val) != 'nan' else ''

    def _eh_ajuste(row):
        nf  = _str(row.iloc[2]).upper()
        cmt = _str(row.iloc[25]).upper()
        return nf == 'AJUSTE' or 'AJUSTE' in cmt

    def _eh_manuseio(row):
        return _str(row.iloc[2]).upper() == 'MANUSEIO'

    def _classificar(row):
        nf  = _str(row.iloc[2]).upper()
        cmt = _str(row.iloc[25]).upper()
        if nf.startswith('DEV-') or 'DEVOLUC' in cmt or 'DEVELU' in cmt \
                or 'ENTRADA DE DEVOL' in cmt or 'ENTRADA DEVOL' in cmt \
                or 'ENTRADA REFERENTE A DEVOL' in cmt:
            return 'Devolução'
        if nf == 'RETIRADA':        return 'Retirada'
        if nf.startswith('CTE') or nf.startswith('CT-E'):
            return 'Entrada c/ CT-e'
        if nf.startswith('OC '):    return 'Entrada c/ OC'
        if nf in ('SEM NF', 'SEN NF', 'SEMNF', 'AGUAR DECL'):
            return 'Entrada s/ NF'
        return 'Entrada c/ NF'

    def _parse_valor(val):
        """Converte valor monetário BR — trata string '5.697,98', int, float."""
        if val is None or str(val) == 'nan':
            return 0.0
        if isinstance(val, float):
            return val
        if isinstance(val, int):
            return float(val)
        # String BR: '5.697,98' → 5697.98 | '10160' → 10160.0
        s = str(val).strip().replace('.', '').replace(',', '.')
        try:    return float(s)
        except (ValueError, TypeError): return 0.0

    def _parse_peso(val):
        """Peso já vem em kg diretamente do Excel — apenas converte o tipo."""
        if val is None or str(val) == 'nan':
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace('.', '').replace(',', '.')
        try:    return float(s)
        except (ValueError, TypeError): return 0.0

    ignorados_ajuste   = 0
    ignorados_manuseio = 0
    registros          = []

    for _, row in df.iterrows():
        if _eh_ajuste(row):
            ignorados_ajuste += 1
            continue
        if _eh_manuseio(row):
            ignorados_manuseio += 1
            continue

        data        = _normalizar_data(row.iloc[8])
        hora        = _normalizar_hora(row.iloc[9])
        depositante = _str(row.iloc[6])
        nf          = _str(row.iloc[2])
        fornecedor  = _str(row.iloc[14])
        valor       = _parse_valor(row.iloc[19])
        peso        = _parse_peso(row.iloc[23])
        comentario  = _str(row.iloc[25])
        usuario     = _str(row.iloc[26])
        sequencia   = _str(row.iloc[0])

        if not data or not depositante:
            continue

        registros.append({
            'sequencia':   sequencia,
            'depositante': depositante,
            'data':        data,
            'hora':        hora,
            'dia_semana':  _dia_sem(data),
            'nota_fiscal': nf,
            'fornecedor':  fornecedor,
            'tipo':        _classificar(row),
            'valor':       valor,
            'peso':        peso,
            'comentario':  comentario,
            'usuario':     usuario,
        })

    log(f"   Ajustes ignorados:   {ignorados_ajuste}\n")
    log(f"   Manuseios ignorados: {ignorados_manuseio}\n")
    log(f"✅ {len(registros)} recebimentos válidos.\n")

    if not registros:
        log("⚠️  Nenhum recebimento válido encontrado.\n")
        return None

    # ── Resumo por depositante ────────────────────────────────────────────
    dep_map = defaultdict(lambda: {
        'total_valor': 0.0, 'total_peso': 0.0, 'total_nf': 0,
        'entradas_nf': 0, 'devolucoes': 0, 'retiradas': 0,
        'sem_nf': 0, 'cte': 0, 'oc': 0,
    })
    for r in registros:
        d = dep_map[r['depositante']]
        d['total_valor'] += r['valor']
        d['total_peso']  += r['peso']
        d['total_nf']    += 1
        if r['tipo'] == 'Entrada c/ NF':    d['entradas_nf'] += 1
        elif r['tipo'] == 'Devolução':       d['devolucoes']  += 1
        elif r['tipo'] == 'Retirada':        d['retiradas']   += 1
        elif r['tipo'] == 'Entrada s/ NF':   d['sem_nf']      += 1
        elif r['tipo'] == 'Entrada c/ CT-e': d['cte']         += 1
        elif r['tipo'] == 'Entrada c/ OC':   d['oc']          += 1

    resumo = [
        {'depositante': dep, **s}
        for dep, s in sorted(dep_map.items())
    ]

    # ── Distribuição por dia ──────────────────────────────────────────────
    dia_map = defaultdict(lambda: {'valor': 0.0, 'eventos': 0, 'dia_semana': ''})
    for r in registros:
        dia_map[r['data']]['valor']      += r['valor']
        dia_map[r['data']]['eventos']    += 1
        dia_map[r['data']]['dia_semana']  = r['dia_semana']

    total_valor  = sum(r['total_valor'] for r in resumo)
    total_peso   = sum(r['total_peso']  for r in resumo)
    log(f"📊 {len(registros)} NFs | R$ {total_valor:,.2f} | {total_peso:,.2f} kg | {len(resumo)} depositantes\n")

    return {
        'registros':     registros,
        'resumo':        resumo,
        'dia_map':       dict(dia_map),
        'total_valor':   total_valor,
        'total_peso':    total_peso,
        'total_eventos': len(registros),
    }


def run_recebimentos(caminho_arq, mes_ref, log):
    """Wrapper — processa arquivo e gera Excel de Recebimentos e Devoluções."""
    dados = processar_recebimentos(caminho_arq, log)
    if not dados:
        return False

    saida = _caminho_saida(PASTA_RECEBIMENTOS, 'relatorio_recebimentos', mes_ref=mes_ref)
    if not saida:
        log(f"❌ Não foi possível criar pasta de saída em:\n   {PASTA_RECEBIMENTOS}\n")
        return False

    log(f"💾 Gerando Excel: {os.path.basename(saida)}...\n")
    COR_REC   = '0C4A6E'
    FMT_MON   = 'R$ #,##0.00'
    FMT_PESO  = '#,##0.00'
    FMT_DATA  = 'DD/MM/YYYY'

    def _cel(ws, row, col, val, bg, fmt=None, align_left=False):
        c = ws.cell(row, col, val)
        c.fill      = _fill_excel(bg)
        c.font      = Font(name='Calibri', size=10)
        c.border    = _borda_excel()
        c.alignment = Alignment(vertical='center',
                                horizontal='left' if align_left else 'center')
        if fmt:
            c.number_format = fmt
        return c

    wb = Workbook(); wb.remove(wb.active)

    # ── Aba 1: Resumo por Depositante ─────────────────────────────────────
    ws1 = wb.create_sheet('Resumo por Depositante')
    _fmt_cabecalho_ws(ws1,
        ['Depositante', 'Total Recebimentos', 'Valor Total (R$)', 'Peso Total (kg)',
         'Entradas c/ NF', 'Entradas s/ NF', 'CT-e', 'OC', 'Devoluções', 'Retiradas'],
        COR_REC)
    for i, r in enumerate(dados['resumo']):
        bg = 'F2F7FB' if i % 2 == 0 else 'FFFFFF'
        _cel(ws1, i+2, 1,  r['depositante'],        bg, align_left=True)
        _cel(ws1, i+2, 2,  r['total_nf'],            bg)
        _cel(ws1, i+2, 3,  r['total_valor'],         bg, FMT_MON)
        _cel(ws1, i+2, 4,  r['total_peso'],          bg, FMT_PESO)
        _cel(ws1, i+2, 5,  r['entradas_nf'],         bg)
        _cel(ws1, i+2, 6,  r['sem_nf'],              bg)
        _cel(ws1, i+2, 7,  r['cte'],                 bg)
        _cel(ws1, i+2, 8,  r['oc'],                  bg)
        _cel(ws1, i+2, 9,  r['devolucoes'],          bg)
        _cel(ws1, i+2, 10, r['retiradas'],           bg)

    tr = len(dados['resumo']) + 2
    tots = ['TOTAL GERAL', dados['total_eventos'],
            dados['total_valor'], dados['total_peso'],
            sum(r['entradas_nf'] for r in dados['resumo']),
            sum(r['sem_nf']      for r in dados['resumo']),
            sum(r['cte']         for r in dados['resumo']),
            sum(r['oc']          for r in dados['resumo']),
            sum(r['devolucoes']  for r in dados['resumo']),
            sum(r['retiradas']   for r in dados['resumo'])]
    fmts_t = [None, None, FMT_MON, FMT_PESO] + [None]*6
    for j, (v, f) in enumerate(zip(tots, fmts_t), 1):
        c = _fmt_total(ws1.cell(tr, j, v))
        if f:
            ws1.cell(tr, j).number_format = f
    _autofit(ws1)

    # ── Aba 2: Detalhe por NF ──────────────────────────────────────────────
    ws2 = wb.create_sheet('Detalhe por NF')
    _fmt_cabecalho_ws(ws2,
        ['Seq.', 'Data', 'Hora', 'Dia da Semana', 'Depositante',
         'Fornecedor', 'Nota Fiscal', 'Tipo', 'Valor (R$)', 'Peso (kg)', 'Comentário'],
        COR_REC)
    _COR_TIPO = {
        'Devolução':       'FFF3CD',
        'Retirada':        'D1ECF1',
        'Entrada s/ NF':   'F8D7DA',
        'Entrada c/ CT-e': 'E8F4FD',
        'Entrada c/ OC':   'EDE7F6',
    }

    def _data_sort(d):
        try:
            dd, mm, aa = d.split('/'); return (int(aa), int(mm), int(dd))
        except Exception:
            return (0, 0, 0)

    regs_ord = sorted(dados['registros'],
                      key=lambda x: (_data_sort(x['data']), x['depositante']))
    for i, r in enumerate(regs_ord):
        bg = _COR_TIPO.get(r['tipo'], 'F2F7FB' if i % 2 == 0 else 'FFFFFF')
        _cel(ws2, i+2, 1,  r['sequencia'],   bg)
        _cel(ws2, i+2, 2,  r['data'],        bg, FMT_DATA)
        _cel(ws2, i+2, 3,  r['hora'],        bg)
        _cel(ws2, i+2, 4,  r['dia_semana'],  bg)
        _cel(ws2, i+2, 5,  r['depositante'], bg, align_left=True)
        _cel(ws2, i+2, 6,  r['fornecedor'],  bg, align_left=True)
        _cel(ws2, i+2, 7,  r['nota_fiscal'], bg)
        _cel(ws2, i+2, 8,  r['tipo'],        bg)
        _cel(ws2, i+2, 9,  r['valor'],       bg, FMT_MON)
        _cel(ws2, i+2, 10, r['peso'],        bg, FMT_PESO)
        _cel(ws2, i+2, 11, r['comentario'],  bg, align_left=True)
    _autofit(ws2)

    # ── Aba 3: Distribuição por Dia ────────────────────────────────────────
    ws3 = wb.create_sheet('Distribuição por Dia')
    _fmt_cabecalho_ws(ws3,
        ['Data', 'Dia da Semana', 'Nº de Entradas', 'Valor Total (R$)'],
        COR_REC)

    for i, data in enumerate(sorted(dados['dia_map'].keys(), key=_data_sort)):
        info = dados['dia_map'][data]; bg = 'F2F7FB' if i % 2 == 0 else 'FFFFFF'
        _cel(ws3, i+2, 1, data,               bg, FMT_DATA)
        _cel(ws3, i+2, 2, info['dia_semana'], bg)
        _cel(ws3, i+2, 3, info['eventos'],    bg)
        _cel(ws3, i+2, 4, info['valor'],      bg, FMT_MON)

    tr3 = len(dados['dia_map']) + 2
    _fmt_total(ws3.cell(tr3, 1, 'TOTAL'))
    _fmt_total(ws3.cell(tr3, 2, ''))
    c3 = _fmt_total(ws3.cell(tr3, 3, dados['total_eventos']))
    c4 = _fmt_total(ws3.cell(tr3, 4, dados['total_valor']))
    ws3.cell(tr3, 4).number_format = FMT_MON
    _autofit(ws3)

    wb.save(saida)
    _registrar_extracao('Recebimentos e Devoluções')
    _atualizar_historico('recebimentos', saida, mes_ref, log)
    log("✅ Relatório de Recebimentos e Devoluções gerado!\n")
    return saida


# ============================================================
# MÓDULO 9 — FINANCEIRO (Contas a Pagar / Contas a Receber)
# ============================================================

def processar_financeiro(caminho, tipo, log):
    """
    Lê Excel de Contas a Pagar ou Contas a Receber.
    tipo: 'pagar' | 'receber'
    Colunas esperadas: Fornecedor/Cliente, Documento, Emissão, Vencto,
                       Valor Principal, Valor Título, Saldo, Natureza,
                       Comentário, Processo
    Status calculado: Saldo=0 → Pago/Recebido | Vencto<hoje → Atrasado | resto → Pendente
    """
    from collections import defaultdict

    log(f"📄 Lendo arquivo: {os.path.basename(caminho)}...\n")
    try:
        df = pd.read_excel(caminho)
        log(f"   {len(df)} registros encontrados.\n")
    except Exception as e:
        log(f"❌ Não foi possível ler o arquivo: {e}\n")
        return None

    if df is None or len(df) == 0:
        log("⚠️  Arquivo vazio.\n")
        return None

    hoje = date.today()
    _DIAS_PT = ['Segunda','Terça','Quarta','Quinta','Sexta','Sábado','Domingo']

    def _str(v):
        return str(v).strip() if v is not None and str(v) != 'nan' else ''

    def _data(v):
        if v is None or str(v) == 'nan': return ''
        if hasattr(v, 'strftime'): return v.strftime('%d/%m/%Y')
        import re as _re
        s = str(v).strip()
        m = _re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
        if m: return f"{m.group(3)}/{m.group(2)}/{m.group(1)}"
        return s[:10] if len(s) >= 10 else s

    def _data_obj(v):
        try:
            if hasattr(v, 'date'): return v.date()
            if hasattr(v, 'year'): return v
            import re as _re
            s = str(v).strip()
            m = _re.match(r'^(\d{2})/(\d{2})/(\d{4})', s)
            if m: return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
            m2 = _re.match(r'^(\d{4})-(\d{2})-(\d{2})', s)
            if m2: return date(int(m2.group(1)), int(m2.group(2)), int(m2.group(3)))
        except Exception:
            pass
        return None

    def _float(v):
        if v is None or str(v) == 'nan': return 0.0
        if isinstance(v, (int, float)): return float(v)
        try: return float(str(v).replace('.','').replace(',','.'))
        except (ValueError, TypeError): return 0.0

    def _status(saldo, vencto_val):
        if saldo <= 0.001:
            return 'Pago' if tipo == 'pagar' else 'Recebido'
        dt = _data_obj(vencto_val)
        if dt and dt < hoje:
            return 'Atrasado'
        return 'Pendente'

    # Detecta coluna de contraparte (Fornecedor ou Cliente)
    col_parte = 'Fornecedor' if 'Fornecedor' in df.columns else 'Cliente'

    registros = []
    for _, row in df.iterrows():
        vencto_raw  = row.get('Vencto', '')
        saldo       = _float(row.get('Saldo', 0))
        valor_tit   = _float(row.get('Valor Título', 0))
        valor_princ = _float(row.get('Valor Principal', 0))
        status      = _status(saldo, vencto_raw)

        registros.append({
            'parte':        _str(row.get(col_parte, '')),
            'documento':    _str(row.get('Documento', '')),
            'emissao':      _data(row.get('Emissão', '')),
            'vencto':       _data(vencto_raw),
            'vencto_raw':   vencto_raw,
            'valor_princ':  valor_princ,
            'valor_titulo': valor_tit,
            'saldo':        saldo,
            'natureza':     _str(row.get('Natureza', '')),
            'comentario':   _str(row.get('Comentário', '')),
            'processo':     _str(row.get('Processo', '')),
            'status':       status,
        })

    # KPIs
    total_geral   = sum(r['valor_titulo'] for r in registros)
    total_pago    = sum(r['valor_titulo'] for r in registros if r['status'] in ('Pago','Recebido'))
    total_pendente= sum(r['saldo']        for r in registros if r['status'] == 'Pendente')
    total_atrasado= sum(r['saldo']        for r in registros if r['status'] == 'Atrasado')
    qtd_atrasado  = sum(1 for r in registros if r['status'] == 'Atrasado')
    qtd_pendente  = sum(1 for r in registros if r['status'] == 'Pendente')
    qtd_quitado   = sum(1 for r in registros if r['status'] in ('Pago','Recebido'))

    # Resumo por natureza
    nat_map = defaultdict(lambda: {'total': 0.0, 'saldo': 0.0, 'qtd': 0})
    for r in registros:
        nat = r['natureza'] or 'Sem categoria'
        nat_map[nat]['total'] += r['valor_titulo']
        nat_map[nat]['saldo'] += r['saldo']
        nat_map[nat]['qtd']   += 1

    log(f"📊 {len(registros)} títulos | Total: R$ {total_geral:,.2f}\n")
    log(f"   Quitados: {qtd_quitado} | Pendentes: {qtd_pendente} | Atrasados: {qtd_atrasado}\n")

    return {
        'registros':       registros,
        'col_parte':       col_parte,
        'tipo':            tipo,
        'total_geral':     total_geral,
        'total_pago':      total_pago,
        'total_pendente':  total_pendente,
        'total_atrasado':  total_atrasado,
        'qtd_atrasado':    qtd_atrasado,
        'qtd_pendente':    qtd_pendente,
        'qtd_quitado':     qtd_quitado,
        'nat_map':         dict(nat_map),
    }


def run_financeiro(caminho_pagar, caminho_receber, mes_ref, log):
    """
    Gera/atualiza o relatório financeiro do mês.
    - Se já existe relatorio_financeiro_MM-AAAA.xlsx, abre e substitui apenas
      as abas do tipo gerado (Pagar ou Receber), preservando as demais.
    - Consolidado anual recebe uma aba por mês com Pagar e Receber combinados.
    """
    # ── Resolve caminho de saída ──────────────────────────────────────────
    # Determina mês/ano
    try:
        mm, aaaa = mes_ref.split('-')
        mes, ano = int(mm), int(aaaa)
    except Exception:
        mes, ano = _mes_ano_referencia()

    # Tenta pasta de rede; fallback para pasta local
    pasta = PASTA_FINANCEIRO
    subpasta = os.path.join(pasta, str(ano))
    try:
        os.makedirs(subpasta, exist_ok=True)
    except Exception:
        pasta_fallback = os.path.join(os.path.expanduser('~'), 'Documents',
                                      'CentralRelatorios', 'Financeiro')
        subpasta = os.path.join(pasta_fallback, str(ano))
        log(f"⚠️  Pasta de rede inacessível: {pasta}\n")
        log(f"   Usando pasta local: {pasta_fallback}\n")
        try:
            os.makedirs(subpasta, exist_ok=True)
        except Exception as e:
            log(f"❌ Não foi possível criar pasta de saída: {e}\n")
            return False

    # Arquivo do mês — único, acumula Pagar e Receber
    nome_arq = f'relatorio_financeiro_{mes:02d}-{ano}.xlsx'
    saida    = os.path.join(subpasta, nome_arq)

    FMT_MON  = 'R$ #,##0.00'
    FMT_DATA = 'DD/MM/YYYY'
    COR_PAG_HEX = 'B91C1C'
    COR_REC_HEX = '15803D'
    _COR_STATUS = {
        'Atrasado': 'FECACA', 'Pendente': 'FEF9C3',
        'Pago':     'DCFCE7', 'Recebido': 'DCFCE7',
    }

    def _cel(ws, row, col, val, bg, fmt=None, left=False):
        c = ws.cell(row, col, val)
        c.fill      = _fill_excel(bg)
        c.font      = Font(name='Calibri', size=10)
        c.border    = _borda_excel()
        c.alignment = Alignment(vertical='center',
                                horizontal='left' if left else 'center')
        if fmt: c.number_format = fmt
        return c

    def _escrever_aba_detalhe(wb_dest, dados, cor_hex, nome_aba):
        """Cria/substitui aba de detalhe no workbook."""
        if nome_aba in wb_dest.sheetnames:
            del wb_dest[nome_aba]
        ws = wb_dest.create_sheet(nome_aba)
        parte_label = 'Fornecedor' if dados['tipo'] == 'pagar' else 'Cliente'
        cols = [parte_label, 'Documento', 'Emissão', 'Vencimento',
                'Valor (R$)', 'Saldo (R$)', 'Natureza', 'Processo',
                'Comentário', 'Status']
        _fmt_cabecalho_ws(ws, cols, cor_hex)
        regs_ord = sorted(dados['registros'],
                          key=lambda x: (
                              0 if x['status'] == 'Atrasado' else
                              1 if x['status'] == 'Pendente' else 2,
                              x['vencto']))
        for i, r in enumerate(regs_ord):
            bg = _COR_STATUS.get(r['status'], 'FFFFFF')
            _cel(ws, i+2, 1,  r['parte'],        bg, left=True)
            _cel(ws, i+2, 2,  r['documento'],    bg, left=True)
            _cel(ws, i+2, 3,  r['emissao'],      bg, FMT_DATA)
            _cel(ws, i+2, 4,  r['vencto'],       bg, FMT_DATA)
            _cel(ws, i+2, 5,  r['valor_titulo'], bg, FMT_MON)
            _cel(ws, i+2, 6,  r['saldo'],        bg, FMT_MON)
            _cel(ws, i+2, 7,  r['natureza'],     bg, left=True)
            _cel(ws, i+2, 8,  r['processo'],     bg)
            _cel(ws, i+2, 9,  r['comentario'],   bg, left=True)
            _cel(ws, i+2, 10, r['status'],       bg)
        tr = len(regs_ord) + 2
        _fmt_total(ws.cell(tr, 1, 'TOTAL GERAL'))
        for j in range(2, 5): _fmt_total(ws.cell(tr, j, ''))
        _fmt_total(ws.cell(tr, 5, dados['total_geral']))
        ws.cell(tr, 5).number_format = FMT_MON
        _fmt_total(ws.cell(tr, 6, dados['total_atrasado'] + dados['total_pendente']))
        ws.cell(tr, 6).number_format = FMT_MON
        for j in range(7, 11): _fmt_total(ws.cell(tr, j, ''))
        _autofit(ws)

    def _escrever_aba_categorias_combinadas(wb_dest, resultados, cor_pag, cor_rec):
        """
        Aba 1 — 'Resumo por Categoria'.
        Duas tabelas empilhadas: Pagar (vermelho) e Receber (verde).
        Se apenas um lado foi gerado agora, copia os dados do outro lado
        diretamente da aba de detalhe já salva no workbook.
        """
        def _recuperar_dados_existentes(wb_dest, tipo):
            """Lê totais e categorias da aba de detalhe já salva."""
            nome_aba = 'Contas a Pagar' if tipo == 'pagar' else 'Contas a Receber'
            if nome_aba not in wb_dest.sheetnames:
                return None
            ws_ref = wb_dest[nome_aba]
            # Lê todos os valores para reconstruir nat_map e totais
            from collections import defaultdict
            nat_map = defaultdict(lambda: {'total': 0.0, 'saldo': 0.0, 'qtd': 0})
            total_geral = total_pago = total_pendente = total_atrasado = 0.0
            qtd_quitado = qtd_pendente = qtd_atrasado = 0
            col_valor = col_saldo = col_status = col_nat = None
            headers = [ws_ref.cell(1, c).value for c in range(1, ws_ref.max_column + 1)]
            for j, h in enumerate(headers, 1):
                if h in ('Valor (R$)',):    col_valor  = j
                if h == 'Saldo (R$)':      col_saldo  = j
                if h == 'Status':          col_status = j
                if h == 'Natureza':        col_nat    = j
            if not all([col_valor, col_saldo, col_status, col_nat]):
                return None
            for row in ws_ref.iter_rows(min_row=2, values_only=True):
                if row[0] in ('TOTAL GERAL', None): continue
                try:
                    val    = float(row[col_valor-1] or 0)
                    saldo  = float(row[col_saldo-1] or 0)
                    status = str(row[col_status-1] or '')
                    nat    = str(row[col_nat-1] or 'Sem categoria')
                    total_geral += val
                    nat_map[nat]['total'] += val
                    nat_map[nat]['saldo'] += saldo
                    nat_map[nat]['qtd']   += 1
                    if status in ('Pago', 'Recebido'):
                        total_pago    += val; qtd_quitado  += 1
                    elif status == 'Pendente':
                        total_pendente += saldo; qtd_pendente += 1
                    elif status == 'Atrasado':
                        total_atrasado += saldo; qtd_atrasado += 1
                except Exception:
                    continue
            return {
                'tipo': tipo, 'total_geral': total_geral, 'total_pago': total_pago,
                'total_pendente': total_pendente, 'total_atrasado': total_atrasado,
                'qtd_quitado': qtd_quitado, 'qtd_pendente': qtd_pendente,
                'qtd_atrasado': qtd_atrasado, 'nat_map': dict(nat_map),
            }

        # Completa resultados com dados já salvos do lado oposto
        if '_preservar_pagar' in resultados and 'pagar' not in resultados:
            dados_existentes = _recuperar_dados_existentes(wb_dest, 'pagar')
            if dados_existentes:
                resultados['pagar'] = dados_existentes
        if '_preservar_receber' in resultados and 'receber' not in resultados:
            dados_existentes = _recuperar_dados_existentes(wb_dest, 'receber')
            if dados_existentes:
                resultados['receber'] = dados_existentes

        # Remove flags de controle
        resultados.pop('_preservar_pagar', None)
        resultados.pop('_preservar_receber', None)

        if 'Resumo por Categoria' in wb_dest.sheetnames:
            del wb_dest['Resumo por Categoria']
        ws = wb_dest.create_sheet('Resumo por Categoria', 0)

        linha = 1
        for tipo_key, cor_hex, titulo in [
            ('pagar',   cor_pag, '📕  CONTAS A PAGAR'),
            ('receber', cor_rec, '📗  CONTAS A RECEBER'),
        ]:
            if tipo_key not in resultados:
                continue
            dados = resultados[tipo_key]

            # Título da seção
            ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
            c = ws.cell(linha, 1, titulo)
            c.fill      = _fill_excel(cor_hex)
            c.font      = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            c.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[linha].height = 22
            linha += 1

            # KPIs
            for j, label in enumerate(['Total Geral', 'Quitados', 'Pendentes', 'Atrasados'], 1):
                c = ws.cell(linha, j, label)
                c.fill = _fill_excel(cor_hex); c.border = _borda_excel()
                c.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
                c.alignment = Alignment(horizontal='center', vertical='center')
            linha += 1
            for j, val in enumerate([dados['total_geral'], dados['total_pago'],
                                      dados['total_pendente'], dados['total_atrasado']], 1):
                c = ws.cell(linha, j, val)
                c.fill = _fill_excel('F9FAFB'); c.border = _borda_excel()
                c.font = Font(name='Calibri', size=10, bold=True)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.number_format = FMT_MON
            linha += 1

            # Cabeçalho categoria
            for j, hdr in enumerate(['Categoria', 'Qtd.', 'Valor Total (R$)', 'Saldo (R$)'], 1):
                _fmt_cabecalho(ws.cell(linha, j, hdr), bg=cor_hex)
            linha += 1

            # Linhas de categoria
            nat_sorted = sorted(dados['nat_map'].items(), key=lambda x: -x[1]['total'])
            for i, (nat, s) in enumerate(nat_sorted):
                bg = 'F2F7FB' if i % 2 == 0 else 'FFFFFF'
                for j, (v, fmt, left) in enumerate([
                    (nat,        None,    True),
                    (s['qtd'],   None,    False),
                    (s['total'], FMT_MON, False),
                    (s['saldo'], FMT_MON, False),
                ], 1):
                    c = ws.cell(linha, j, v)
                    c.fill = _fill_excel(bg); c.font = Font(name='Calibri', size=10)
                    c.border = _borda_excel()
                    c.alignment = Alignment(vertical='center',
                                            horizontal='left' if left else 'center')
                    if fmt: c.number_format = fmt
                linha += 1

            # Total
            _fmt_total(ws.cell(linha, 1, 'TOTAL'))
            _fmt_total(ws.cell(linha, 2, sum(s['qtd']   for s in dados['nat_map'].values())))
            _fmt_total(ws.cell(linha, 3, dados['total_geral']))
            ws.cell(linha, 3).number_format = FMT_MON
            _fmt_total(ws.cell(linha, 4, dados['total_atrasado'] + dados['total_pendente']))
            ws.cell(linha, 4).number_format = FMT_MON
            linha += 2

        # Autofit
        for col in ws.columns:
            max_len = 0; col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if cell.value: max_len = max(max_len, len(str(cell.value)))
                except Exception: pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 45)

    # ── Abre ou cria o workbook do mês ───────────────────────────────────
    arquivo_existe = os.path.isfile(saida)
    if arquivo_existe:
        wb = load_workbook(saida)
        log(f"📂 Atualizando relatório existente: {nome_arq}\n")
    else:
        wb = Workbook(); wb.remove(wb.active)
        log(f"💾 Criando relatório: {nome_arq}\n")

    resultados = {}

    # Se o arquivo já existe e estamos gerando só um lado,
    # recupera os dados do lado oposto que já estava no arquivo
    if arquivo_existe and caminho_pagar and not caminho_receber:
        # Tenta recuperar Receber de uma geração anterior usando o arquivo de detalhe
        # Não temos o caminho original, então marcamos para não sobrescrever a aba
        resultados['_preservar_receber'] = True
    if arquivo_existe and caminho_receber and not caminho_pagar:
        resultados['_preservar_pagar'] = True

    if caminho_pagar:
        log(f"\n{'─'*52}\n📕 Processando Contas a Pagar...\n")
        dados_pagar = processar_financeiro(caminho_pagar, 'pagar', log)
        if dados_pagar:
            resultados['pagar'] = dados_pagar

    if caminho_receber:
        log(f"\n{'─'*52}\n📗 Processando Contas a Receber...\n")
        dados_receber = processar_financeiro(caminho_receber, 'receber', log)
        if dados_receber:
            resultados['receber'] = dados_receber

    if not any(k for k in resultados if not k.startswith('_')):
        log("❌ Nenhum dado processado.\n")
        return False

    # ── Aba 1: Categorias combinadas (Pagar + Receber) ────────────────────
    _escrever_aba_categorias_combinadas(wb, resultados, COR_PAG_HEX, COR_REC_HEX)

    # ── Abas 2 e 3: Detalhe individual ────────────────────────────────────
    # Só reescreve o lado que foi processado agora; o lado preservado já está salvo
    if 'pagar' in resultados and 'registros' in resultados['pagar']:
        _escrever_aba_detalhe(wb, resultados['pagar'], COR_PAG_HEX, 'Contas a Pagar')
    if 'receber' in resultados and 'registros' in resultados['receber']:
        _escrever_aba_detalhe(wb, resultados['receber'], COR_REC_HEX, 'Contas a Receber')

    # Garante ordem correta das abas
    nomes_ordem = ['Resumo por Categoria', 'Contas a Pagar', 'Contas a Receber']
    for i, nome in enumerate(nomes_ordem):
        if nome in wb.sheetnames:
            wb.move_sheet(nome, offset=wb.sheetnames.index(nome) - i)

    wb.save(saida)
    log(f"✅ Relatório salvo: {nome_arq}\n")

    # ── Consolidado anual ─────────────────────────────────────────────────
    _atualizar_historico_financeiro(resultados, mes, ano, log)

    _registrar_extracao('Financeiro')
    log(f"\n✅ Relatório Financeiro gerado!\n")
    return saida


def _atualizar_historico_financeiro(resultados, mes, ano, log):
    """
    Mantém Financeiro_AAAA.xlsx com uma aba por mês.
    Cada aba do mês contém: resumo Pagar + resumo Receber lado a lado por categoria,
    diferenciados por cor (vermelho = Pagar, verde = Receber).
    """
    _MESES_PT = ['Jan','Fev','Mar','Abr','Mai','Jun',
                 'Jul','Ago','Set','Out','Nov','Dez']
    prefixo   = f'{_MESES_PT[mes-1]}-{ano}'

    try:
        os.makedirs(PASTA_CONSOLIDADOS, exist_ok=True)
    except Exception as e:
        log(f"⚠️  Consolidado: {e}\n"); return

    cons_path = os.path.join(PASTA_CONSOLIDADOS, f'Financeiro_{ano}.xlsx')

    if os.path.isfile(cons_path):
        wb_c = load_workbook(cons_path)
    else:
        wb_c = Workbook(); wb_c.remove(wb_c.active)

    # Remove aba do mês se existir
    if prefixo in wb_c.sheetnames:
        del wb_c[prefixo]

    ws = wb_c.create_sheet(prefixo)

    FMT_MON = 'R$ #,##0.00'
    COR_PAG_HEX = 'B91C1C'
    COR_REC_HEX = '15803D'

    def _cel(ws, row, col, val, bg, fmt=None, bold=False, left=False):
        c = ws.cell(row, col, val)
        c.fill      = _fill_excel(bg)
        c.font      = Font(name='Calibri', size=10, bold=bold)
        c.border    = _borda_excel()
        c.alignment = Alignment(vertical='center',
                                horizontal='left' if left else 'center')
        if fmt: c.number_format = fmt
        return c

    linha = 1

    for tipo_key, cor_hex, titulo in [
        ('pagar',   COR_PAG_HEX, '📕  CONTAS A PAGAR'),
        ('receber', COR_REC_HEX, '📗  CONTAS A RECEBER'),
    ]:
        if tipo_key not in resultados:
            continue
        dados = resultados[tipo_key]

        # Título da seção
        ws.merge_cells(start_row=linha, start_column=1, end_row=linha, end_column=4)
        c = ws.cell(linha, 1, titulo)
        c.fill      = _fill_excel(cor_hex)
        c.font      = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[linha].height = 22
        linha += 1

        # KPIs resumo
        kpis = [
            ('Total Geral',   dados['total_geral'],     cor_hex),
            ('Quitados',      dados['total_pago'],      '15803D' if tipo_key == 'receber' else '166534'),
            ('Pendentes',     dados['total_pendente'],  'B45309'),
            ('Atrasados',     dados['total_atrasado'],  'DC2626'),
        ]
        for j, (label, val, bg) in enumerate(kpis, 1):
            _cel(ws, linha, j, label, bg, bold=True)
            ws.cell(linha, j).font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
        linha += 1
        for j, (_, val, bg) in enumerate(kpis, 1):
            _cel(ws, linha, j, val, 'F9FAFB', FMT_MON)
        linha += 1

        # Cabeçalho categoria
        _fmt_cabecalho(ws.cell(linha, 1, 'Categoria'), bg=cor_hex)
        _fmt_cabecalho(ws.cell(linha, 2, 'Qtd.'),      bg=cor_hex)
        _fmt_cabecalho(ws.cell(linha, 3, 'Valor (R$)'), bg=cor_hex)
        _fmt_cabecalho(ws.cell(linha, 4, 'Saldo (R$)'), bg=cor_hex)
        linha += 1

        # Linhas de categoria
        nat_sorted = sorted(dados['nat_map'].items(), key=lambda x: -x[1]['total'])
        for i, (nat, s) in enumerate(nat_sorted):
            bg = 'F2F7FB' if i % 2 == 0 else 'FFFFFF'
            _cel(ws, linha, 1, nat,        bg, left=True)
            _cel(ws, linha, 2, s['qtd'],   bg)
            _cel(ws, linha, 3, s['total'], bg, FMT_MON)
            _cel(ws, linha, 4, s['saldo'], bg, FMT_MON)
            linha += 1

        # Total
        _fmt_total(ws.cell(linha, 1, 'TOTAL'))
        _fmt_total(ws.cell(linha, 2, sum(s['qtd']   for s in dados['nat_map'].values())))
        _fmt_total(ws.cell(linha, 3, dados['total_geral']))
        ws.cell(linha, 3).number_format = FMT_MON
        _fmt_total(ws.cell(linha, 4, dados['total_atrasado'] + dados['total_pendente']))
        ws.cell(linha, 4).number_format = FMT_MON
        linha += 2  # espaço entre seções

    # Autofit
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value: max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 14), 45)

    wb_c.save(cons_path)
    status = '🔄 Consolidado atualizado' if os.path.isfile(cons_path) else '📚 Consolidado criado'
    log(f"{status}: Financeiro_{ano}.xlsx\n")


# ============================================================
# MÓDULO 10 — FATURAMENTO ARMAZENAGEM
# ============================================================

def processar_faturamento_armazenagem(arquivo_mov, arquivo_volumes, log,
                                     arquivo_estoque=None):
    """
    Gera dados do relatório de pico de armazenagem por cliente/família/grupo.

    Lógica:
      1. Arquivo de volumes (por família) → data e m³ do pico por subdivisão
         — nome real extraído da linha 3 do cabeçalho:
           'Familia do Produto = X' → família
           'Grupo do Produto = X'   → grupo (IPSEN, YELUM)
           nenhum dos dois          → depositante único
      2. DB de configuração → tipo do cliente e preços
      3. DB de famílias     → mapeamento SKU → família/grupo
      4. Arquivo de movimentação → saldo qtd e valor monetário de cada SKU no pico
         Fallback: arquivo de estoque estático para SKUs sem movimentação no período
      5. Agrupa por cliente → subdivisão → SKUs com faturamento calculado
    """
    import re as _re
    from collections import defaultdict

    # ── Carrega DBs ───────────────────────────────────────────────────────
    cfg          = _carregar_db_precos_arm()
    db_fam       = _carregar_db_familias()
    clientes_cfg = cfg.get('clientes', {}) if cfg else {}
    heranca      = cfg.get('heranca', {}) if cfg else {}
    ipsen_rh_grp = cfg.get('ipsen_rh_grupos', []) if cfg else []
    ipsen_rh_pre = cfg.get('ipsen_rh_preco') if cfg else None

    if not clientes_cfg:
        log("⚠️  DB de configuração vazio — use 'Carregar Configuração' primeiro.\n"
            "   Relatório gerado sem agrupamento e sem preços.\n\n")

    # ── Carrega estoque estático (fallback) ───────────────────────────────
    # est_db[cliente_norm] = {codigo: (saldo_qtd, saldo_valor)}
    est_db = {}
    if arquivo_estoque:
        try:
            xl_est = pd.ExcelFile(arquivo_estoque)
            for aba in xl_est.sheet_names:
                df_e = xl_est.parse(aba, header=None)
                cli_norm = aba.strip().upper()
                est_db[cli_norm] = {}
                for _, row in df_e.iloc[4:].iterrows():
                    cod = str(row[0]).strip() if pd.notna(row[0]) else ''
                    if not cod or cod in ('nan','None',''): continue
                    qtd = float(row[6]) if pd.notna(row[6]) else 0.0
                    val = float(row[8]) if pd.notna(row[8]) else 0.0
                    if qtd > 0:
                        est_db[cli_norm][cod] = (qtd, val)
            log(f"📦 Estoque estático carregado: {len(est_db)} clientes como fallback.\n\n")
        except Exception as e:
            log(f"⚠️  Erro ao carregar estoque estático: {e}\n\n")

    # ── PASSO 1: Pico m³ por subdivisão (arquivo de volumes por família) ──
    log("📊 Lendo picos de volume por família/grupo...\n")
    # pico_sub[nome_real] = {data, vol, depositante, tipo_sub ('familia'|'grupo'|'unico')}
    pico_sub = {}

    try:
        xl_vol = pd.ExcelFile(arquivo_volumes)
        for aba in xl_vol.sheet_names:
            try:
                df_v = xl_vol.parse(aba, header=None)
                if df_v.empty or len(df_v) < 4: continue

                # Extrai nome real e depositante do cabeçalho (linha 3)
                filtro = str(df_v.iloc[3].iloc[0]) \
                         if pd.notna(df_v.iloc[3].iloc[0]) else ''

                m_fam = _re.search(
                    r'Familia do Produto\s*=\s*(.+?);\s*Per', filtro)
                m_grp = _re.search(
                    r'Grupo do Produto\s*=\s*(.+?);\s*Per', filtro)
                m_dep = _re.search(
                    r'Depositante\s*=\s*(.+?)(?:;|$)', filtro)

                if m_fam:
                    nome_real = m_fam.group(1).strip()
                    tipo_sub  = 'familia'
                elif m_grp:
                    nome_real = m_grp.group(1).strip()
                    tipo_sub  = 'grupo'
                else:
                    # Depositante único — usa nome da aba
                    nome_real = aba.strip()
                    tipo_sub  = 'unico'

                depositante = m_dep.group(1).strip() if m_dep else aba.strip()

                # Dados começam na linha 12 (após cabeçalhos do relatório)
                dados = df_v.iloc[11:].copy()
                col_data = pd.to_datetime(dados.iloc[:, 0],
                                          dayfirst=True, errors='coerce')
                col_vol  = pd.to_numeric(dados.iloc[:, 14], errors='coerce')
                df_tmp   = pd.DataFrame(
                    {'data': col_data, 'vol': col_vol}).dropna()
                if df_tmp.empty: continue

                # Filtra apenas o mês de referência — ignora o dia 01 do
                # mês seguinte que o sistema insere como saldo de fechamento
                meses = df_tmp['data'].dt.month
                mes_ref = int(meses.mode()[0])
                df_mes = df_tmp[df_tmp['data'].dt.month == mes_ref]
                if df_mes.empty:
                    df_mes = df_tmp  # fallback: usa tudo

                idx = df_mes['vol'].idxmax()
                pico_sub[nome_real] = {
                    'data':        df_mes.loc[idx, 'data'].date(),
                    'vol':         round(float(df_mes.loc[idx, 'vol']), 4),
                    'depositante': depositante,
                    'tipo_sub':    tipo_sub,
                    'aba':         aba.strip(),
                }
                log(f"   📋 {nome_real} [{tipo_sub}] — dep: {depositante} — "
                    f"{pico_sub[nome_real]['vol']:.2f} m³ "
                    f"em {pico_sub[nome_real]['data'].strftime('%d/%m/%Y')}\n")
            except Exception as e:
                log(f"   ⚠️  Aba '{aba}': {e}\n")
    except Exception as e:
        log(f"❌ Erro ao ler arquivo de volumes: {e}\n"); return None

    if not pico_sub:
        log("❌ Nenhum dado de volume encontrado.\n"); return None
    log(f"✅ {len(pico_sub)} subdivisões com pico.\n\n")

    # ── PASSO 2: Parse do arquivo de movimentação ─────────────────────────
    log("📄 Lendo arquivo de movimentação...\n")

    def _sv(v): return str(v).strip() \
        if v is not None and str(v) not in ('nan','None') else ''
    def _fv(v):
        try: return float(v)
        except (ValueError, TypeError): return 0.0

    mov_db = {}
    try:
        xl_mov = pd.ExcelFile(arquivo_mov)
        for aba in xl_mov.sheet_names:
            try:
                df_m = xl_mov.parse(aba, header=None)
                regs = defaultdict(list)
                sku_atual = None
                for _, row in df_m.iterrows():
                    data_raw = _sv(row.iloc[0])
                    prod_raw = _sv(row.iloc[2])
                    saldo_q  = _fv(row.iloc[11])
                    saldo_v  = _fv(row.iloc[14])
                    if not prod_raw: continue
                    if data_raw in ('Saldo Inicial', 'Saldo Final', ''):
                        sku_atual = prod_raw
                        regs[sku_atual].append((data_raw, saldo_q, saldo_v))
                    else:
                        if prod_raw: sku_atual = prod_raw
                        if sku_atual:
                            try:
                                d, m, a = data_raw.split('/')
                                dt = date(int(a), int(m), int(d))
                                regs[sku_atual].append((dt, saldo_q, saldo_v))
                            except (ValueError, AttributeError):
                                pass
                mov_db[aba.strip()] = dict(regs)
                log(f"   ✅ {aba}: {len(regs)} SKUs\n")
            except Exception as e:
                log(f"   ⚠️  Aba '{aba}': {e}\n")
    except Exception as e:
        log(f"❌ Erro ao ler movimentação: {e}\n"); return None

    log(f"✅ {len(mov_db)} clientes na movimentação.\n\n")

    # ── PASSO 3: Saldo na data do pico ────────────────────────────────────
    def _saldo_em(registros, data_alvo):
        ini_q = ini_v = 0.0
        ult_q = ult_v = None; ult_dt = None
        for marc, sq, sv in registros:
            if marc == 'Saldo Inicial':
                ini_q, ini_v = sq, sv
            elif isinstance(marc, date) and marc <= data_alvo:
                if ult_dt is None or marc >= ult_dt:
                    ult_dt, ult_q, ult_v = marc, sq, sv
        return (ult_q, ult_v) if ult_q is not None else (ini_q, ini_v)

    def _norm(s):
        import re as _re3
        s = s.upper().strip()
        # Normaliza sufixos jurídicos e artigos que variam entre arquivos
        s = _re3.sub(r'\b(S/A|S\.A\.?|LTDA\.?|DO|DE|DA|DOS|DAS)\b', '', s)
        s = _re3.sub(r'[/\.\-]', ' ', s)  # / . - viram espaço
        s = _re3.sub(r'\s+', ' ', s).strip()
        return s

    def _match_mov(nome):
        n = _norm(nome)
        # 1. Match exato normalizado
        for k in mov_db:
            if _norm(k) == n: return k
        # 2. Um contém o outro (normalizado)
        for k in mov_db:
            nk = _norm(k)
            if n in nk or nk in n: return k
        # 3. Primeira palavra igual
        for k in mov_db:
            if _norm(k).split()[0] == n.split()[0]: return k
        return None

    def _match_cfg_cli(nome):
        n = _norm(nome)
        for k in clientes_cfg:
            nk = _norm(k)
            if nk == n or n in nk or nk in n \
               or nk.split()[0] == n.split()[0]: return k
        return None

    def _extrair_candidatos(cod):
        """
        Gera variações do código para match tolerante.
        O Kardex usa formatos como:
          '(P12315-01)DYS-BR-002'   → candidato: 'P12315-01'
          '(GCR TNE(1576))SOM-BR'   → candidato: 'GCR TNE'
          'BIOGEN-252139(P611-01'    → candidato: 'P611-01'
          'BIOGEN-265009*'           → candidato: 'BIOGEN-265009'
          'DYS-BR-002205(P12309-01)' → candidato: 'DYS-BR-002205'
        """
        import re as _r
        cands = set()
        c = cod.strip()
        cands.add(c)
        # Remove asterisco e sufixos como '.', '-MAR/2025', ' (17)' no final
        cands.add(_r.sub(r'[*.]$', '', c).strip())
        cands.add(_r.sub(r'[-\s]+(?:JAN|FEV|MAR|ABR|MAI|JUN|JUL|AGO|SET|OUT|NOV|DEZ)[\/\-]\d+.*$', '', c, flags=_r.IGNORECASE).strip())
        cands.add(_r.sub(r'\s*\(\d+\)\s*$', '', c).strip())  # remove ' (17)' no fim
        # Extrai todos os conteúdos entre parênteses
        for m in _r.findall(r'\(([^)]+)\)', c):
            cands.add(m.strip())
            # Remove sufixo numérico: 'GCR TNE(1576)' → 'GCR TNE'
            cands.add(_r.sub(r'\(\d+\)$', '', m).strip())
        # Prefixo antes do primeiro '('
        if '(' in c:
            cands.add(c.split('(')[0].strip())
        # Sufixo após o último '(' — pega código incompleto como 'P611-01'
        if '(' in c:
            cands.add(c.split('(')[-1].strip())
        # Sufixo após o último ')'
        if ')' in c:
            cands.add(c.split(')')[-1].strip())
        # Remove prefixo '(xxx)' do início → resto
        m = _r.match(r'\([^)]+\)(.+)', c)
        if m: cands.add(m.group(1).strip())
        # Código antes de '-BR-' ou '-BRA-'
        m = _r.match(r'(.+?)(?:-BR|-BRA)[-_]', c)
        if m: cands.add(m.group(1).strip())
        # Parênteses aninhados: '(GCR TNE(1576))' → 'GCR TNE'
        m = _r.match(r'\(([^(]+?)\s*\(', c)
        if m: cands.add(m.group(1).strip())
        # Código base sem sufixo após '/'
        m = _r.match(r'([^/]+)/', c)
        if m: cands.add(m.group(1).strip())
        return {x for x in cands if x}

    def _familia_sku(aba_mov, sku_raw, usar_grupo=False):
        import re as _re2
        # sku_raw vem como "[CÓDIGO] Descrição" ou "[(P12315-01)DYS-BR-002] Descrição"
        m = _re2.search(r'\[(.+?)\]', sku_raw)
        cod = m.group(1).strip() if m else sku_raw.strip()

        for cli_db, skus in db_fam.items():
            na = _norm(aba_mov); nk = _norm(cli_db)
            if na == nk or na in nk or nk in na:
                def _get_val(inf):
                    if usar_grupo:
                        grp = inf.get('grupo', '')
                        return grp if grp and grp not in ('nan','None','') \
                               else inf.get('familia', 'SEM GRUPO')
                    return inf.get('familia', 'SEM FAMÍLIA')

                # 1. Match direto
                if cod in skus:
                    return _get_val(skus[cod])
                # 2. Case-insensitive
                cod_u = cod.upper()
                for cod_db, inf in skus.items():
                    if cod_db.upper() == cod_u:
                        return _get_val(inf)
                # 3. Match tolerante com candidatos
                for cand in _extrair_candidatos(cod):
                    if not cand: continue
                    if cand in skus:
                        return _get_val(skus[cand])
                    cu = cand.upper()
                    for cod_db, inf in skus.items():
                        if cod_db.upper() == cu:
                            return _get_val(inf)
                # 4. Match parcial
                for cod_db, inf in skus.items():
                    cd = cod_db.upper()
                    if (len(cd) >= 6 and len(cod_u) >= 6 and
                            (cod_u.startswith(cd) or cd.startswith(cod_u) or
                             cod_u in cd or cd in cod_u)):
                        return _get_val(inf)
                break
        return 'SEM GRUPO' if usar_grupo else 'SEM FAMÍLIA'

    def _preco_subdivisao(depositante_norm, nome_sub, cfg_key):
        """Retorna config de preço para a subdivisão — trata IPSEN RH."""
        cfg_base = clientes_cfg.get(cfg_key, {}) if cfg_key else {}
        # Verifica se é grupo RH do IPSEN
        if ipsen_rh_pre and ipsen_rh_grp:
            for rh in ipsen_rh_grp:
                if _norm(nome_sub) == _norm(rh):
                    return ipsen_rh_pre
        return cfg_base

    # ── PASSO 4: Agrupa resultado por cliente → subdivisão → SKUs ────────
    log("🔗 Cruzando dados...\n")

    # Agrupa subdivisões por depositante
    por_cliente = defaultdict(list)
    for nome_sub, info_sub in pico_sub.items():
        dep = info_sub['depositante']
        por_cliente[dep].append((nome_sub, info_sub))

    resultado = {}

    for depositante, subdivisoes in por_cliente.items():
        cfg_key  = _match_cfg_cli(depositante)
        cfg_dep  = clientes_cfg.get(cfg_key, {}) if cfg_key else {}
        mov_cli  = _match_mov(depositante)
        tipo_dep = cfg_dep.get('tipo', 'unico')

        skus_total    = []
        subdiv_dados  = {}
        fat_total     = {'preco_m3': 0, 'seguro': 0,
                         'subtotal': 0, 'iss': 0, 'total': 0}

        for nome_sub, info_sub in subdivisoes:
            data_pico = info_sub['data']
            vol_pico  = info_sub['vol']

            # SKUs na data do pico — via Kardex
            skus_sub = []
            codigos_kardex = set()  # rastreia o que já foi incluído
            if mov_cli:
                for sku_raw, regs in mov_db[mov_cli].items():
                    qtd, val = _saldo_em(regs, data_pico)
                    if qtd <= 0: continue
                    # Família/grupo do SKU
                    if tipo_dep == 'unico':
                        sub_sku = depositante
                    else:
                        usar_grp = (tipo_dep == 'grupo')
                        sub_sku = _familia_sku(mov_cli, sku_raw, usar_grupo=usar_grp)
                    # Só inclui se pertence a esta subdivisão
                    if tipo_dep != 'unico' and \
                       _norm(sub_sku) != _norm(nome_sub):
                        continue
                    # Extrai código limpo para deduplicação
                    import re as _re3
                    m = _re3.search(r'\[(.+?)\]', sku_raw)
                    cod_limpo = m.group(1).strip() if m else sku_raw.strip()
                    codigos_kardex.add(cod_limpo.upper())
                    skus_sub.append({
                        'sku':    sku_raw,
                        'qtd':    qtd,
                        'valor':  val,
                        'subdiv': nome_sub,
                    })

            # Fallback: estoque estático para SKUs sem movimentação no período
            if est_db:
                dep_norm = _norm(depositante)
                est_cli = None
                for k in est_db:
                    if k == dep_norm or dep_norm in k or k in dep_norm:
                        est_cli = k; break
                if est_cli:
                    for cod_est, (qtd_est, val_est) in est_db[est_cli].items():
                        if cod_est.upper() in codigos_kardex:
                            continue  # já incluído via Kardex
                        # Verifica família
                        sku_raw_fake = f'[{cod_est}] (estoque)'
                        if tipo_dep == 'unico':
                            sub_sku = depositante
                        else:
                            usar_grp = (tipo_dep == 'grupo')
                            sub_sku = _familia_sku(
                                mov_cli or depositante,
                                sku_raw_fake, usar_grupo=usar_grp)
                        if tipo_dep != 'unico' and \
                           _norm(sub_sku) != _norm(nome_sub):
                            continue
                        skus_sub.append({
                            'sku':    f'[{cod_est}]',
                            'qtd':    qtd_est,
                            'valor':  val_est,
                            'subdiv': nome_sub,
                            'fonte':  'estoque',
                        })

            skus_sub.sort(key=lambda x: -x['qtd'])

            # Faturamento da subdivisão
            cfg_sub = _preco_subdivisao(depositante, nome_sub, cfg_key)
            preco   = cfg_sub.get('preco_m3', 0) * vol_pico
            seguro  = cfg_sub.get('seguro_pct', 0) / 100 * \
                      cfg_sub.get('seguro_base', 1000) * vol_pico
            sub_val = preco + seguro
            iss     = cfg_sub.get('iss_pct', 0) / 100 * sub_val
            total   = round(sub_val + iss, 4)

            subdiv_dados[nome_sub] = {
                'data_pico': data_pico,
                'vol_pico':  vol_pico,
                'skus':      skus_sub,
                'fat': {
                    'preco_m3': cfg_sub.get('preco_m3', 0),
                    'seguro':   round(seguro, 4),
                    'subtotal': round(sub_val, 4),
                    'iss':      round(iss, 4),
                    'total':    total,
                },
            }
            skus_total.extend(skus_sub)

            # Acumula totais do cliente
            fat_total['preco_m3'] = cfg_dep.get('preco_m3', 0)
            fat_total['seguro']   = round(
                fat_total['seguro'] + seguro, 4)
            fat_total['subtotal'] = round(
                fat_total['subtotal'] + sub_val, 4)
            fat_total['iss']      = round(
                fat_total['iss'] + iss, 4)
            fat_total['total']    = round(
                fat_total['total'] + total, 4)

        log(f"   ✅ {depositante}: {len(subdivisoes)} subdivisão(ões) · "
            f"{len(skus_total)} SKUs · "
            f"Total R$ {fat_total['total']:,.2f}\n")

        resultado[depositante] = {
            'subdivisoes':     subdiv_dados,
            'skus':            skus_total,
            'tipo':            tipo_dep,
            'cfg':             cfg_dep,
            'cfg_key':         cfg_key,
            'mov_cli':         mov_cli,
            'fat':             fat_total,
            'heranca':         heranca.get(depositante),
        }

    return resultado

    # ── Carrega DBs de configuração ───────────────────────────────────────
    cfg = _carregar_db_precos_arm()
    db_fam = _carregar_db_familias()
    clientes_cfg = cfg.get('clientes', {}) if cfg else {}

    if not clientes_cfg:
        log("⚠️  DB de configuração vazio — use 'Carregar Configuração' primeiro.\n"
            "   O relatório será gerado sem agrupamento por família/grupo e sem preços.\n\n")

    # ── PASSO 1: Pico m³ por cliente (arquivo de volumes) ─────────────────
    log("📊 Lendo picos de volume por cliente...\n")
    pico_vol = {}
    try:
        xl_vol = pd.ExcelFile(arquivo_volumes)
        for aba in xl_vol.sheet_names:
            try:
                df_v  = xl_vol.parse(aba, header=None)
                dados = df_v.iloc[3:-1].copy()
                if dados.empty: continue
                col_data = pd.to_datetime(dados[0], dayfirst=True, errors='coerce')
                col_vol  = pd.to_numeric(dados[14], errors='coerce')
                df_tmp   = pd.DataFrame({'data': col_data, 'vol': col_vol}).dropna()
                if df_tmp.empty: continue
                # Filtra apenas o mês de referência
                mes_ref = int(df_tmp['data'].dt.month.mode()[0])
                df_mes = df_tmp[df_tmp['data'].dt.month == mes_ref]
                if df_mes.empty: df_mes = df_tmp
                idx = df_mes['vol'].idxmax()
                pico_vol[aba.strip()] = {
                    'data': df_mes.loc[idx, 'data'].date(),
                    'vol':  round(float(df_mes.loc[idx, 'vol']), 4),
                }
                log(f"   📋 {aba}: {pico_vol[aba.strip()]['vol']:.2f} m³ "
                    f"em {pico_vol[aba.strip()]['data'].strftime('%d/%m/%Y')}\n")
            except Exception as e:
                log(f"   ⚠️  Aba '{aba}': {e}\n")
    except Exception as e:
        log(f"❌ Erro ao ler arquivo de volumes: {e}\n"); return None

    if not pico_vol:
        log("❌ Nenhum dado de volume encontrado.\n"); return None
    log(f"✅ {len(pico_vol)} clientes com pico.\n\n")

    # ── PASSO 2: Parse do arquivo de movimentação ─────────────────────────
    log("📄 Lendo arquivo de movimentação...\n")

    def _sv(v): return str(v).strip() if v is not None and str(v) not in ('nan','None') else ''
    def _fv(v):
        try: return float(v)
        except (ValueError, TypeError): return 0.0

    # mov_db[aba] = {sku: [(marcador_ou_data, saldo_qtd, saldo_val), ...]}
    mov_db = {}
    try:
        xl_mov = pd.ExcelFile(arquivo_mov)
        for aba in xl_mov.sheet_names:
            try:
                df_m = xl_mov.parse(aba, header=None)
                regs = defaultdict(list)
                sku_atual = None
                for _, row in df_m.iterrows():
                    data_raw = _sv(row.iloc[0])
                    prod_raw = _sv(row.iloc[2])
                    saldo_q  = _fv(row.iloc[11])
                    saldo_v  = _fv(row.iloc[14])  # valor monetário — última coluna
                    if not prod_raw: continue
                    if data_raw in ('Saldo Inicial', 'Saldo Final', ''):
                        sku_atual = prod_raw
                        regs[sku_atual].append((data_raw, saldo_q, saldo_v))
                    else:
                        if prod_raw: sku_atual = prod_raw
                        if sku_atual:
                            try:
                                d, m, a = data_raw.split('/')
                                dt = date(int(a), int(m), int(d))
                                regs[sku_atual].append((dt, saldo_q, saldo_v))
                            except (ValueError, AttributeError):
                                pass
                mov_db[aba.strip()] = dict(regs)
                log(f"   ✅ {aba}: {len(regs)} SKUs\n")
            except Exception as e:
                log(f"   ⚠️  Aba '{aba}': {e}\n")
    except Exception as e:
        log(f"❌ Erro ao ler movimentação: {e}\n"); return None

    log(f"✅ {len(mov_db)} clientes na movimentação.\n\n")

    # ── PASSO 3: Saldo na data do pico ────────────────────────────────────
    def _saldo_em(registros, data_alvo):
        """Retorna (qtd, valor) do último registro <= data_alvo."""
        ini_q = ini_v = 0.0
        ult_q = ult_v = None
        ult_dt = None
        for item in registros:
            marc, sq, sv = item
            if marc == 'Saldo Inicial':
                ini_q, ini_v = sq, sv
            elif isinstance(marc, date) and marc <= data_alvo:
                if ult_dt is None or marc >= ult_dt:
                    ult_dt, ult_q, ult_v = marc, sq, sv
        if ult_q is not None:
            return ult_q, ult_v
        return ini_q, ini_v

    def _norm(s): return s.upper().strip().replace('  ', ' ')

    def _match_mov(nome, mov_db):
        n = _norm(nome)
        for k in mov_db:
            nk = _norm(k)
            if nk == n or n in nk or nk in n: return k
        return None

    def _match_cfg(nome, clientes_cfg):
        n = _norm(nome)
        for k in clientes_cfg:
            nk = _norm(k)
            if nk == n or n in nk or nk in n or \
               nk.split()[0] == n.split()[0]: return k
        return None

    def _get_familia_sku_local(aba_mov, sku_raw):
        """Extrai família do SKU a partir do DB de famílias."""
        sku_limpo = sku_raw.strip()
        # Tenta match no DB de famílias pelo nome da aba
        for cli_db, skus in db_fam.items():
            nk = _norm(cli_db)
            na = _norm(aba_mov)
            if nk == na or na in nk or nk in na:
                # Tenta match do código
                if sku_limpo in skus:
                    return skus[sku_limpo].get('familia', 'SEM FAMÍLIA')
                # Tenta sem colchetes
                cod_c = sku_limpo.strip('[]').strip()
                for cod_db, info_db in skus.items():
                    if cod_c == cod_db.strip('[]').strip():
                        return info_db.get('familia', 'SEM FAMÍLIA')
                break
        return 'SEM FAMÍLIA'

    log("🔗 Cruzando dados...\n")
    resultado = {}

    for nome_vol, info_pico in pico_vol.items():
        data_pico = info_pico['data']
        vol_pico  = info_pico['vol']

        # Casa volumes → movimentação → configuração
        mov_cli = _match_mov(nome_vol)
        cfg_key = _match_cfg_cli(nome_vol)
        cfg_cli = clientes_cfg.get(cfg_key, {}) if cfg_key else {}

        tipo        = cfg_cli.get('tipo', 'unico')
        subdivisoes = cfg_cli.get('subdivisoes', [])

        # Coleta SKUs com saldo > 0 na data do pico
        skus_pico = []
        if mov_cli:
            for sku_raw, regs in mov_db[mov_cli].items():
                qtd, val = _saldo_em(regs, data_pico)
                if qtd > 0:
                    familia = _get_familia_sku_local(mov_cli, sku_raw) \
                              if tipo != 'unico' else nome_vol
                    skus_pico.append({
                        'sku':    sku_raw,
                        'qtd':    qtd,
                        'valor':  val,
                        'subdiv': familia,
                    })
            skus_pico.sort(key=lambda x: (-x['qtd']))
            log(f"   ✅ {nome_vol}: {len(skus_pico)} SKUs | "
                f"tipo={tipo} | cfg={'sim' if cfg_key else 'não'}\n")
        else:
            log(f"   ⚠️  {nome_vol}: não encontrado na movimentação\n")

        # Calcula valor a faturar
        preco_m3 = cfg_cli.get('preco_m3', 0)
        seguro   = cfg_cli.get('seguro_pct', 0) / 100 * \
                   cfg_cli.get('seguro_base', 1000) * vol_pico
        subtotal = preco_m3 * vol_pico + seguro
        iss      = cfg_cli.get('iss_pct', 0) / 100 * subtotal
        total_fat = round(subtotal + iss, 4)

        resultado[nome_vol] = {
            'data_pico':   data_pico,
            'vol_pico':    vol_pico,
            'skus':        skus_pico,
            'tipo':        tipo,
            'subdivisoes': subdivisoes,
            'cfg':         cfg_cli,
            'fat': {
                'preco_m3':  preco_m3,
                'seguro':    round(seguro, 4),
                'subtotal':  round(subtotal, 4),
                'iss':       round(iss, 4),
                'total':     total_fat,
            },
            'cfg_key': cfg_key,
            'mov_cli': mov_cli,
        }

    return resultado


def run_faturamento_armazenagem(arquivo_mov, arquivo_volumes, mes_ref, log,
                               arquivo_estoque=None):
    """Gera relatório Excel: Índice + Resumo + aba por cliente."""
    from collections import defaultdict
    dados = processar_faturamento_armazenagem(
        arquivo_mov, arquivo_volumes, log,
        arquivo_estoque=arquivo_estoque)
    if not dados: return False

    saida = _caminho_saida(PASTA_FINANCEIRO, 'faturamento_armazenagem', mes_ref=mes_ref)
    if not saida:
        pasta_fallback = os.path.join(os.path.expanduser('~'), 'Documents',
                                      'CentralRelatorios', 'Financeiro')
        saida = _caminho_saida(pasta_fallback, 'faturamento_armazenagem', mes_ref=mes_ref)
        if not saida:
            log("❌ Não foi possível criar pasta de saída.\n"); return False

    log(f"\n💾 Gerando Excel: {os.path.basename(saida)}...\n")

    COR_VIO  = '6D28D9'
    COR_SUB  = '312E81'
    FMT_MON  = 'R$ #,##0.00'
    FMT_VOL  = '#,##0.00'

    CORES_CLI = ['1E3A5F','14532D','7C2D12','064E3B','1E1B4B',
                 '3B1F6B','0F3460','1A3C34','5C1A1A','1C3553']

    def _cor_cli(idx): return CORES_CLI[idx % len(CORES_CLI)]

    def _cel(ws, r, c, v, bg, fmt=None, left=False, bold=False, fg='111827'):
        x = ws.cell(r, c, v); x.fill = _fill_excel(bg)
        x.font = Font(name='Calibri', size=10, bold=bold, color=fg)
        x.border = _borda_excel()
        x.alignment = Alignment(
            horizontal='left' if left else 'center', vertical='center')
        if fmt: x.number_format = fmt
        return x

    def _nav(ws, r, c, label, cor, link):
        x = ws.cell(r, c, label); x.fill = _fill_excel(cor)
        x.font = Font(name='Calibri', size=9, bold=True, color='FFFFFF')
        x.border = _borda_excel()
        x.alignment = Alignment(horizontal='center', vertical='center')
        x.hyperlink = link

    wb = Workbook(); wb.remove(wb.active)
    clientes_ord = sorted(dados.items())

    # ── ABA ÍNDICE ────────────────────────────────────────────────────────
    ws_idx = wb.create_sheet('Índice')
    ws_idx.merge_cells('A1:F1')
    c = ws_idx.cell(1, 1, '🏭  FATURAMENTO ARMAZENAGEM — ÍNDICE')
    c.fill = _fill_excel(COR_VIO)
    c.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_idx.row_dimensions[1].height = 28

    for j, h in enumerate(
            ['Cliente','Tipo','Nº Subdivisões','Volume m³ Total',
             'Total a Faturar','Detalhes'], 1):
        x = ws_idx.cell(2, j, h); x.fill = _fill_excel('3730A3')
        x.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        x.border = _borda_excel()
        x.alignment = Alignment(horizontal='center', vertical='center')
    ws_idx.row_dimensions[2].height = 20

    # ── Gera siglas únicas para nomes de aba ─────────────────────────
    siglas_map = {}   # {cli: sigla_unica}
    usadas = set()
    for cli, _ in clientes_ord:
        base = cli.split()[0][:12]
        # Remove caracteres inválidos para nomes de aba
        base = ''.join(c for c in base if c not in r'\/?*[]:/\'')
        sigla = base
        n = 1
        while sigla in usadas:
            sufixo = str(n)
            sigla = base[:12 - len(sufixo)] + sufixo
            n += 1
        usadas.add(sigla)
        siglas_map[cli] = sigla

    def _link(nome_aba):
        """Formato de hiperlink interno compatível com todas as versões do Excel."""
        # Sem aspas simples quando o nome não tem espaços/especiais
        if ' ' in nome_aba or any(c in nome_aba for c in '!@#$%&()-+'):
            return f"#'{nome_aba}'!A1"
        return f"#{nome_aba}!A1"

    for i, (cli, info) in enumerate(clientes_ord):
        r = i + 3; cor = _cor_cli(i)
        bg = 'F5F3FF' if i % 2 == 0 else 'FFFFFF'
        sigla = siglas_map[cli]
        vol_total = sum(sd['vol_pico']
                        for sd in info['subdivisoes'].values())
        her = f" [HUMANIA]" if info.get('heranca') else ''
        _cel(ws_idx, r, 1, cli + her,                      bg, left=True, bold=True)
        _cel(ws_idx, r, 2, info['tipo'].capitalize(),      bg)
        _cel(ws_idx, r, 3, len(info['subdivisoes']),       bg)
        _cel(ws_idx, r, 4, round(vol_total, 4),            bg, FMT_VOL)
        _cel(ws_idx, r, 5, info['fat']['total'],           bg, FMT_MON,
             bold=info['fat']['total'] > 0)
        _nav(ws_idx, r, 6, f'→ {sigla}', cor, _link(sigla))
        ws_idx.row_dimensions[r].height = 20

    tr = len(clientes_ord) + 3
    _cel(ws_idx, tr, 1, 'TOTAL', COR_SUB, bold=True, left=True, fg='FFFFFF')
    for c_ in range(2, 5): _cel(ws_idx, tr, c_, '', COR_SUB, fg='FFFFFF')
    vol_geral = sum(
        sum(sd['vol_pico'] for sd in info['subdivisoes'].values())
        for _, info in clientes_ord)
    _cel(ws_idx, tr, 4, round(vol_geral, 4),
         COR_SUB, FMT_VOL, fg='FFFFFF', bold=True)
    _cel(ws_idx, tr, 5,
         round(sum(info['fat']['total'] for _, info in clientes_ord), 4),
         COR_SUB, FMT_MON, fg='FFFFFF', bold=True)
    _cel(ws_idx, tr, 6, '', COR_SUB, fg='FFFFFF')

    for i, w in enumerate([34,10,14,14,16,12], 1):
        ws_idx.column_dimensions[get_column_letter(i)].width = w

    # ── ABA RESUMO ────────────────────────────────────────────────────────
    ws_res = wb.create_sheet('Resumo')
    ws_res.merge_cells('A1:G1')
    c = ws_res.cell(1, 1, '🏭  FATURAMENTO ARMAZENAGEM — RESUMO POR CLIENTE')
    c.fill = _fill_excel(COR_VIO)
    c.font = Font(name='Calibri', size=13, bold=True, color='FFFFFF')
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_res.row_dimensions[1].height = 28
    _nav(ws_res, 1, 8, '← Índice', '4338CA', _link('Índice'))

    row_res = 2
    COLS_RES = ['Subdivisão','Data Pico','Volume m³',
                'Preço/m³','Seguro','ISS','Total a Faturar']

    for i, (cli, info) in enumerate(clientes_ord):
        cor   = _cor_cli(i)
        sigla = siglas_map[cli]
        fat   = info['fat']
        her   = f" [herda HUMANIA]" if info.get('heranca') else ''

        # Bloco cliente
        ws_res.merge_cells(start_row=row_res, start_column=1,
                           end_row=row_res, end_column=7)
        c = ws_res.cell(row_res, 1, f'  {cli}{her}')
        c.fill = _fill_excel(cor)
        c.font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='left', vertical='center')
        _nav(ws_res, row_res, 8, f'→ {sigla}', cor, _link(sigla))
        ws_res.row_dimensions[row_res].height = 22; row_res += 1

        # Cabeçalho
        for j, h in enumerate(COLS_RES, 1):
            x = ws_res.cell(row_res, j, h); x.fill = _fill_excel('3730A3')
            x.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            x.border = _borda_excel()
            x.alignment = Alignment(horizontal='center', vertical='center')
        ws_res.row_dimensions[row_res].height = 18; row_res += 1

        # Linhas por subdivisão
        for k, (nome_sub, sd) in enumerate(
                sorted(info['subdivisoes'].items())):
            bg  = 'F5F3FF' if k % 2 == 0 else 'FFFFFF'
            sf  = sd['fat']
            _cel(ws_res, row_res, 1, nome_sub,
                 bg, left=True)
            _cel(ws_res, row_res, 2,
                 sd['data_pico'].strftime('%d/%m/%Y'), bg)
            _cel(ws_res, row_res, 3, sd['vol_pico'],  bg, FMT_VOL)
            _cel(ws_res, row_res, 4, sf['preco_m3'],  bg, FMT_MON)
            _cel(ws_res, row_res, 5, sf['seguro'],    bg, FMT_MON)
            _cel(ws_res, row_res, 6, sf['iss'],       bg, FMT_MON)
            _cel(ws_res, row_res, 7, sf['total'],     bg, FMT_MON, bold=True)
            row_res += 1

        # Total cliente
        _cel(ws_res, row_res, 1, f'TOTAL — {cli}',
             COR_SUB, bold=True, left=True, fg='FFFFFF')
        _cel(ws_res, row_res, 2, '', COR_SUB, fg='FFFFFF')
        vol_cli = sum(sd['vol_pico']
                      for sd in info['subdivisoes'].values())
        _cel(ws_res, row_res, 3, round(vol_cli, 4),
             COR_SUB, FMT_VOL, fg='FFFFFF', bold=True)
        _cel(ws_res, row_res, 4, fat['preco_m3'],
             COR_SUB, FMT_MON, fg='FFFFFF')
        _cel(ws_res, row_res, 5, fat['seguro'],
             COR_SUB, FMT_MON, fg='FFFFFF')
        _cel(ws_res, row_res, 6, fat['iss'],
             COR_SUB, FMT_MON, fg='FFFFFF')
        _cel(ws_res, row_res, 7, fat['total'],
             COR_SUB, FMT_MON, fg='FFFFFF', bold=True)
        row_res += 2

    for i, w in enumerate([34,12,13,13,13,12,16,10], 1):
        ws_res.column_dimensions[get_column_letter(i)].width = w

    # ── ABAS POR CLIENTE ──────────────────────────────────────────────────
    COLS_DET = ['Subdivisão','SKU / Código','Qtd no Pico','Saldo (R$)']

    for i, (cli, info) in enumerate(clientes_ord):
        cor   = _cor_cli(i)
        sigla = siglas_map[cli]
        ws_c  = wb.create_sheet(sigla)
        fat   = info['fat']

        # Título
        vol_cli = sum(sd['vol_pico'] for sd in info['subdivisoes'].values())
        ws_c.merge_cells('A1:D1')
        c = ws_c.cell(1, 1,
            f'  {cli}  |  {len(info["subdivisoes"])} subdivisão(ões)'
            f'  |  {vol_cli:.2f} m³ total'
            f'  |  Faturamento: R$ {fat["total"]:,.2f}')
        c.fill = _fill_excel(cor)
        c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws_c.row_dimensions[1].height = 22
        _nav(ws_c, 1, 5, '← Índice',  '4338CA', _link('Índice'))
        _nav(ws_c, 1, 6, '📋 Resumo', COR_VIO,  _link('Resumo'))

        # Linha de faturamento total
        ws_c.merge_cells('A2:D2')
        c = ws_c.cell(2, 1,
            f'  Seguro total: R$ {fat["seguro"]:.2f}'
            f'  |  ISS total: R$ {fat["iss"]:.2f}'
            f'  |  TOTAL A FATURAR: R$ {fat["total"]:.2f}')
        c.fill = _fill_excel(COR_SUB)
        c.font = Font(name='Calibri', size=9, color='FFFFFF')
        c.alignment = Alignment(horizontal='left', vertical='center')
        ws_c.row_dimensions[2].height = 18

        row_d = 3

        # Itera subdivisões em ordem
        for nome_sub, sd in sorted(info['subdivisoes'].items()):
            sf      = sd['fat']
            skus_sd = sd['skus']

            # Sub-título com dados do pico e faturamento da subdivisão
            ws_c.merge_cells(start_row=row_d, start_column=1,
                             end_row=row_d, end_column=4)
            c = ws_c.cell(row_d, 1,
                f'  {nome_sub}'
                f'  |  Pico: {sd["data_pico"].strftime("%d/%m/%Y")}'
                f'  |  {sd["vol_pico"]:.2f} m³'
                f'  |  Preço/m³: R$ {sf["preco_m3"]:.2f}'
                f'  |  Total: R$ {sf["total"]:.2f}')
            c.fill = _fill_excel(COR_SUB)
            c.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
            c.alignment = Alignment(horizontal='left', vertical='center')
            ws_c.row_dimensions[row_d].height = 20; row_d += 1

            if not skus_sd:
                # Distingue: volume zero (grupo sem estoque no período)
                # vs volume > 0 mas SKUs não encontrados no Kardex
                if sd['vol_pico'] == 0:
                    msg = '⚪  Sem estoque no período — volume m³ = 0,00'
                    cor_msg = '9CA3AF'
                else:
                    msg = (f'⚠️  Volume de {sd["vol_pico"]:.2f} m³ identificado no pico, '
                           f'mas nenhum SKU encontrado no Kardex para esta subdivisão.\n'
                           f'     Verifique se os SKUs desta família estão cadastrados '
                           f'corretamente no Banco de Dados de Famílias.')
                    cor_msg = 'B45309'
                ws_c.merge_cells(start_row=row_d, start_column=1,
                                 end_row=row_d, end_column=4)
                x = ws_c.cell(row_d, 1, msg)
                x.font = Font(name='Calibri', size=9, italic=True, color=cor_msg)
                x.fill = _fill_excel('F9FAFB')
                x.alignment = Alignment(horizontal='left', vertical='center',
                                        wrap_text=True)
                ws_c.row_dimensions[row_d].height = 28 if sd['vol_pico'] > 0 else 18
                row_d += 2
                continue

            # Cabeçalho
            for j, h in enumerate(COLS_DET, 1):
                x = ws_c.cell(row_d, j, h); x.fill = _fill_excel('3730A3')
                x.font = Font(name='Calibri', size=10,
                              bold=True, color='FFFFFF')
                x.border = _borda_excel()
                x.alignment = Alignment(
                    horizontal='center', vertical='center')
            ws_c.row_dimensions[row_d].height = 18; row_d += 1

            for k, s in enumerate(
                    sorted(skus_sd, key=lambda x: -x['qtd'])):
                bg = 'F5F3FF' if k % 2 == 0 else 'FFFFFF'
                _cel(ws_c, row_d, 1, nome_sub,           bg, left=True)
                _cel(ws_c, row_d, 2, s['sku'],           bg, left=True)
                _cel(ws_c, row_d, 3, s['qtd'],           bg)
                _cel(ws_c, row_d, 4, round(s['valor'], 2), bg, FMT_MON)
                row_d += 1

            _fmt_total(ws_c.cell(row_d, 1, f'TOTAL {nome_sub}'))
            _fmt_total(ws_c.cell(row_d, 2, ''))
            _fmt_total(ws_c.cell(row_d, 3,
                sum(s['qtd']   for s in skus_sd)))
            _fmt_total(ws_c.cell(row_d, 4,
                round(sum(s['valor'] for s in skus_sd), 2)))
            ws_c.cell(row_d, 4).number_format = FMT_MON
            row_d += 2

        for j, w in enumerate([28, 22, 14, 16, 10, 10], 1):
            ws_c.column_dimensions[get_column_letter(j)].width = w

    wb.save(saida)
    _registrar_extracao('Financeiro')
    log(f"✅ Faturamento Armazenagem gerado: {os.path.basename(saida)}\n")
    return saida


class App:
    def __init__(self, root):
        self.root        = root
        self.estado      = _carregar_estado()
        self._pagina_ativa = None

        # Garante que config_pastas sempre reflita as constantes atuais do código
        self.estado.setdefault('config_pastas', {})
        self.estado['config_pastas'].update({
            'pasta_pedidos':          PASTA_PEDIDOS,
            'pasta_fretes':           PASTA_FRETES,
            'pasta_armazenagem':      PASTA_ARMAZENAGEM,
            'pasta_estoque':          PASTA_ESTOQUE,
            'pasta_produtividade':    PASTA_PRODUTIVIDADE,
            'pasta_cap_operacional':  PASTA_CAP_OPERACIONAL,
            'pasta_recebimentos':      PASTA_RECEBIMENTOS,
            'pasta_financeiro':        PASTA_FINANCEIRO,
            'pasta_consolidados':     PASTA_CONSOLIDADOS,
        })
        _salvar_estado(self.estado)

        root.title('Central de Relatórios')
        root.geometry('960x620')
        root.minsize(800, 520)
        root.resizable(True, True)
        root.configure(bg=COR_BG)
        root.state('zoomed')

        self._build()
        self._ir_para('home')
        self._verificar_lembrete()

    # ----------------------------------------------------------
    def _build(self):
        # Layout principal: sidebar | conteúdo
        self.frm_sidebar  = tk.Frame(self.root, bg=COR_SIDEBAR, width=220)
        self.frm_sidebar.pack(side='left', fill='y')
        self.frm_sidebar.pack_propagate(False)

        self.frm_conteudo = tk.Frame(self.root, bg=COR_BG)
        self.frm_conteudo.pack(side='left', fill='both', expand=True)

        self._build_sidebar()

    # ----------------------------------------------------------
    def _build_sidebar(self):
        sb = self.frm_sidebar

        # Canvas scrollável para a sidebar — sem scrollbar visível
        self._sb_canvas = tk.Canvas(sb, bg=COR_SIDEBAR, highlightthickness=0,
                                    width=220)
        self._sb_canvas.pack(side='left', fill='both', expand=True)

        # Frame interno onde todos os widgets são colocados
        sb = tk.Frame(self._sb_canvas, bg=COR_SIDEBAR)
        self._sb_inner = sb
        self._sb_canvas_window = self._sb_canvas.create_window(
            (0, 0), window=sb, anchor='nw')

        def _on_configure(e):
            self._sb_canvas.configure(
                scrollregion=self._sb_canvas.bbox('all'))
            self._sb_canvas.itemconfig(
                self._sb_canvas_window,
                width=self._sb_canvas.winfo_width())

        sb.bind('<Configure>', _on_configure)
        self._sb_canvas.bind('<Configure>', _on_configure)

        # Scroll isolado: só ativo quando o mouse está sobre a sidebar
        def _on_mousewheel(e):
            self._sb_canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')

        def _sb_enter(e):
            self._sb_canvas.bind_all('<MouseWheel>', _on_mousewheel)

        def _sb_leave(e):
            self._sb_canvas.unbind_all('<MouseWheel>')

        self._sb_canvas.bind('<Enter>', _sb_enter)
        self._sb_canvas.bind('<Leave>', _sb_leave)
        sb.bind('<Enter>', _sb_enter)
        sb.bind('<Leave>', _sb_leave)

        def _bind_scroll_hover(widget):
            """Propaga Enter/Leave para ativar/desativar scroll da sidebar."""
            widget.bind('<Enter>', _sb_enter)
            widget.bind('<Leave>', _sb_leave)
            for child in widget.winfo_children():
                _bind_scroll_hover(child)

        sb.after(100, lambda: _bind_scroll_hover(sb))

        tk.Frame(sb, bg=COR_ACCENT, height=4).pack(fill='x')

        _base     = os.path.dirname(os.path.abspath(__file__))
        _ico_path = os.path.join(_base, 'central_relatorios.ico')
        _logo_carregado = False
        try:
            import base64 as _b64, io as _io
            from PIL import Image as _Img, ImageTk as _ITk
            _raw  = _b64.b64decode(_LOGO_B64)
            _img  = _Img.open(_io.BytesIO(_raw)).convert('RGBA')
            self._logo_img = _ITk.PhotoImage(_img)
            tk.Label(sb, image=self._logo_img, bg=COR_SIDEBAR).pack(pady=(14, 0))
            _logo_carregado = True
        except Exception:
            pass
        if not _logo_carregado:
            tk.Label(sb, text='📊', font=('Segoe UI', 28),
                     bg=COR_SIDEBAR, fg=COR_TEXTO).pack(pady=(20, 0))
        try:
            if os.path.isfile(_ico_path):
                self.root.iconbitmap(_ico_path)
        except Exception:
            pass

        tk.Label(sb, text='Central de\nRelatórios',
                 font=('Segoe UI', 11, 'bold'),
                 bg=COR_SIDEBAR, fg=COR_TEXTO,
                 justify='center').pack(pady=(4, 20))
        tk.Frame(sb, bg=COR_BORDA, height=1).pack(fill='x', padx=16)

        self._nav_btns = {}

        # Estado de expansão dos grupos
        grupos_abertos = {'operacional': tk.BooleanVar(value=True),
                          'financeiro':  tk.BooleanVar(value=True)}

        # Sub-itens de cada grupo — frames que serão mostrados/ocultados
        frm_sub = {}

        def _btn_nav(key, label, is_sub=False, parent=None):
            """Cria botão de navegação na sidebar."""
            container = parent if parent else sb
            fonte = ('Segoe UI', 9)          if is_sub else ('Segoe UI', 10, 'bold')
            pady  = 6                         if is_sub else 10
            btn = tk.Button(container, text=label, font=fonte,
                            bg=COR_SIDEBAR, fg=COR_TEXTO_SUB,
                            relief='flat', anchor='w',
                            padx=20, pady=pady, cursor='hand2',
                            activebackground=COR_SEL,
                            command=lambda k=key: self._ir_para(k))
            btn.pack(fill='x', padx=8, pady=1)
            btn.bind('<Enter>', lambda e, b=btn, k=key:
                     (b.config(bg=COR_SEL) if k != self._pagina_ativa else None,
                      _sb_enter(e)))
            btn.bind('<Leave>', lambda e, b=btn, k=key:
                     (b.config(bg=COR_SEL if k == self._pagina_ativa else COR_SIDEBAR),
                      _sb_leave(e)))
            self._nav_btns[key] = btn
            return btn

        def _btn_grupo(key, label, icone_ab, icone_fe, var_aberto):
            """Cria botão de grupo colapsável com seta."""
            frm_sub[key] = tk.Frame(sb, bg=COR_SIDEBAR)

            def _toggle():
                if var_aberto.get():
                    frm_sub[key].pack_forget()
                    var_aberto.set(False)
                    btn_g.config(text=f'{icone_fe}  {label}  ▸')
                else:
                    frm_sub[key].pack(fill='x', after=btn_g)
                    var_aberto.set(True)
                    btn_g.config(text=f'{icone_ab}  {label}  ▾')

            btn_g = tk.Button(sb, text=f'{icone_ab}  {label}  ▾',
                              font=('Segoe UI', 10, 'bold'),
                              bg=COR_SIDEBAR, fg=COR_TEXTO,
                              relief='flat', anchor='w',
                              padx=20, pady=10, cursor='hand2',
                              activebackground=COR_SEL,
                              command=_toggle)
            btn_g.pack(fill='x', padx=8, pady=1)
            btn_g.bind('<Enter>', _sb_enter)
            btn_g.bind('<Leave>', _sb_leave)
            frm_sub[key].pack(fill='x')
            return btn_g, frm_sub[key]

        # ── Navegação principal ───────────────────────────────────────────
        _btn_nav('home', '🏠  Home')
        tk.Frame(sb, bg=COR_BORDA, height=1).pack(fill='x', padx=16, pady=4)
        _btn_nav('dashboard', '📈  Dashboard')
        tk.Frame(sb, bg=COR_BORDA, height=1).pack(fill='x', padx=16, pady=4)

        # ── Grupo Operacional (colapsável) ────────────────────────────────
        _, frm_op = _btn_grupo('operacional', 'Operacional', '⚙️', '⚙️',
                               grupos_abertos['operacional'])
        for key, label in [
            ('relatorios',     '   📊  Operacional'),
            ('pedidos',        '   📦  Pedidos'),
            ('fretes',         '   🚚  Fretes'),
            ('armazenagem',    '   🏭  Armazenagem'),
            ('estoque',        '   📋  Estoque'),
            ('produtividade',  '   👥  Produtividade'),
            ('cap_operacional','   ⚙️  Cap. Operacional'),
            ('recebimentos',   '   📥  Recebimentos'),
        ]:
            _btn_nav(key, label, is_sub=True, parent=frm_op)

        tk.Frame(sb, bg=COR_BORDA, height=1).pack(fill='x', padx=16, pady=4)

        # ── Grupo Financeiro (colapsável) ─────────────────────────────────
        _, frm_fin = _btn_grupo('financeiro', 'Financeiro', '💰', '💰',
                                grupos_abertos['financeiro'])
        for key, label in [
            ('financeiro', '   🏠  Visão Geral'),
            ('contas',     '   💰  Contas'),
            ('fat_arm',    '   🏭  Fat. Armazenagem'),
        ]:
            _btn_nav(key, label, is_sub=True, parent=frm_fin)

        tk.Frame(sb, bg=COR_BORDA, height=1).pack(fill='x', padx=16, pady=4)

        # ── Demais ────────────────────────────────────────────────────────
        _btn_nav('arquivos', '📁  Arquivos')
        tk.Frame(sb, bg=COR_BORDA, height=1).pack(fill='x', padx=16, pady=4)
        _btn_nav('config', '⚙️  Configurações')

        # Rodapé — fixo fora do canvas, sempre visível
        frm_rodape = self.frm_sidebar
        tk.Frame(frm_rodape, bg=COR_BORDA, height=1).pack(fill='x', padx=16, side='bottom', pady=(0, 8))
        tk.Label(frm_rodape, text=f'v3.1  ·  {date.today().strftime("%d/%m/%Y")}',
                 font=('Segoe UI', 7), bg=COR_SIDEBAR,
                 fg=COR_BORDA).pack(side='bottom', pady=(0, 4))

    # Páginas que usam canvas com scroll — recebem bind global de MouseWheel
    # Para adicionar uma nova página com scroll, basta incluir sua chave aqui
    _PAGINAS_COM_SCROLL = {'arquivos', 'config', 'de_para', 'estoque', 'home'}

    # ----------------------------------------------------------
    def _ir_para(self, pagina):
        # Desliga scroll global da página anterior
        try:
            self.root.unbind_all('<MouseWheel>')
        except Exception:
            pass

        self._pagina_ativa = pagina
        for key, btn in self._nav_btns.items():
            btn.config(bg=COR_SEL if key == pagina else COR_SIDEBAR,
                       fg=COR_TEXTO if key == pagina else COR_TEXTO_SUB)
        for w in self.frm_conteudo.winfo_children():
            w.destroy()
        try:
            {
                'home':        self._pagina_home,
                'relatorios':  self._pagina_relatorios,
                'operacional': self._pagina_relatorios,
                'dashboard':   self._pagina_dashboard,
                'pedidos':     self._pagina_pedidos,
                'fretes':      self._pagina_fretes,
                'armazenagem': self._pagina_armazenagem,
                'arquivos':    self._pagina_arquivos,
                'estoque':     self._pagina_estoque,
                'produtividade': self._pagina_produtividade,
                'cap_operacional': self._pagina_cap_operacional,
                'recebimentos':    self._pagina_recebimentos,
                'financeiro':      self._pagina_financeiro,
                'contas':          self._pagina_financeiro_contas,
                'fat_arm':         self._pagina_fat_arm,
                'de_para':     self._pagina_de_para,
                'config':      self._pagina_config,
            }[pagina]()
        except Exception as _err:
            tk.Label(self.frm_conteudo,
                     text=f'Erro ao carregar página "{pagina}":\n\n{traceback.format_exc()}',
                     font=('Consolas', 8), bg=COR_BG, fg='#ef4444',
                     justify='left', wraplength=900).pack(padx=32, pady=20, anchor='w')

    # ----------------------------------------------------------
    # HOME
    # ----------------------------------------------------------
    def _pagina_home(self):
        c = self.frm_conteudo

        # ── Scroll via Canvas ─────────────────────────────────────────────
        canvas = tk.Canvas(c, bg=COR_BG, highlightthickness=0)
        sb     = ttk.Scrollbar(c, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        inner = tk.Frame(canvas, bg=COR_BG)
        win_id = canvas.create_window((0, 0), window=inner, anchor='nw')

        def _on_resize(e):
            canvas.itemconfig(win_id, width=e.width)
        canvas.bind('<Configure>', _on_resize)

        def _on_frame_resize(e):
            canvas.configure(scrollregion=canvas.bbox('all'))
        inner.bind('<Configure>', _on_frame_resize)

        def _on_mousewheel(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), 'units')
        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>', _on_mousewheel))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

        # A partir daqui tudo vai para 'inner' em vez de 'c'
        c = inner

        # Cabeçalho
        frm_h = tk.Frame(c, bg=COR_BG)
        frm_h.pack(fill='x', padx=32, pady=(28, 4))
        tk.Label(frm_h, text='🏠  Painel Principal',
                 font=('Segoe UI', 18, 'bold'),
                 bg=COR_BG, fg=COR_TEXTO).pack(side='left')

        _MESES = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                   'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
        _DIAS  = ['Segunda-feira','Terça-feira','Quarta-feira','Quinta-feira',
                  'Sexta-feira','Sábado','Domingo']
        hoje_d = date.today()
        data_pt = f"{_DIAS[hoje_d.weekday()]}, {hoje_d.day} de {_MESES[hoje_d.month-1]} de {hoje_d.year}"
        tk.Label(c, text=f'Hoje: {data_pt}',
                 font=('Segoe UI', 9), bg=COR_BG,
                 fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 20))

        # Cards dos relatórios — layout original 3 por linha, uniforme
        estado    = _carregar_estado()
        extrações = estado.get('extrações', {})
        dias_mod  = estado.get('dias_extracao_modulos', {})

        modulos_lista = list(RELATORIOS_CONFIG.items())
        COLS = 3

        self._spin_dias_mod = {}
        for linha_idx in range(0, len(modulos_lista), COLS):
            grupo = modulos_lista[linha_idx:linha_idx + COLS]
            frm_linha = tk.Frame(c, bg=COR_BG)
            frm_linha.pack(fill='x', padx=28, pady=(0, 8))
            for col_idx, (nome, cfg) in enumerate(grupo):
                frm_linha.columnconfigure(col_idx, weight=1, uniform='card')
                ultima  = extrações.get(nome, '—')
                dia_mod = dias_mod.get(nome, estado.get('dia_extracao', 6))
                proxima = _proxima_extracao(dia_mod)
                self._card_home(frm_linha, cfg['icone'], nome,
                                ultima, proxima, cfg['cor'], col_idx, dia_mod)

        # Dica rápida
        frm_dica = tk.Frame(c, bg=COR_CARD,
                            highlightbackground=COR_BORDA, highlightthickness=1)
        frm_dica.pack(fill='x', padx=28, pady=(24, 32))
        tk.Label(frm_dica,
                 text='💡  Clique em "Relatórios" na barra lateral para gerar novos relatórios, '
                      'ou em "Arquivos" para visualizar os já gerados.',
                 font=('Segoe UI', 9), bg=COR_CARD, fg=COR_TEXTO_SUB,
                 wraplength=660, justify='left').pack(padx=16, pady=12)

    def _card_home(self, parent, icone, titulo, ultima, proxima, cor, col, dia_mod=6):
        # Verifica se a extração está atrasada
        hoje = date.today()
        atrasado = False
        try:
            dt_proxima = datetime.strptime(proxima, '%d/%m/%Y').date()
            atrasado = dt_proxima < hoje
        except Exception:
            pass

        cor_borda = COR_ERRO if atrasado else COR_BORDA
        frm = tk.Frame(parent, bg=COR_CARD,
                       highlightbackground=cor_borda, highlightthickness=2 if atrasado else 1)
        frm.grid(row=0, column=col, padx=10, pady=4, sticky='nsew')
        parent.columnconfigure(col, weight=1)

        # Barra superior — vermelha se atrasado
        tk.Frame(frm, bg=COR_ERRO if atrasado else cor, height=4).pack(fill='x')

        # Badge de atraso
        if atrasado:
            tk.Label(frm, text='⚠️  EXTRAÇÃO ATRASADA',
                     font=('Segoe UI', 7, 'bold'),
                     bg=COR_ERRO, fg='white',
                     padx=6, pady=2).pack(fill='x')

        tk.Label(frm, text=icone, font=('Segoe UI', 26),
                 bg=COR_CARD, fg=COR_ERRO if atrasado else cor).pack(pady=(14, 2))
        tk.Label(frm, text=titulo, font=('Segoe UI', 10, 'bold'),
                 bg=COR_CARD, fg=COR_TEXTO, wraplength=160,
                 justify='center').pack()

        tk.Frame(frm, bg=COR_BORDA, height=1).pack(fill='x', padx=16, pady=10)

        for label, valor, cor_val in [
            ('Última extração', ultima, COR_TEXTO),
            ('Próxima extração', proxima,
             COR_ERRO if atrasado else (COR_AVISO if ultima == '—' else COR_ACCENT3)),
        ]:
            tk.Label(frm, text=label, font=('Segoe UI', 7),
                     bg=COR_CARD, fg=COR_TEXTO_SUB).pack()
            tk.Label(frm, text=valor, font=('Segoe UI', 11, 'bold'),
                     bg=COR_CARD, fg=cor_val).pack(pady=(0, 4))

        # Botão Agendar Próxima Extração
        tk.Frame(frm, bg=COR_BORDA, height=1).pack(fill='x', padx=16, pady=(4, 8))

        btn_agendar = tk.Label(frm, text='📅  Agendar Próxima Extração',
                               font=('Segoe UI', 8, 'bold'),
                               bg=cor, fg='white',
                               padx=10, pady=5, cursor='hand2')
        btn_agendar.pack(pady=(0, 14))

        def _hover_in(e):  btn_agendar.config(bg=COR_ACCENT)
        def _hover_out(e): btn_agendar.config(bg=cor)
        btn_agendar.bind('<Enter>', _hover_in)
        btn_agendar.bind('<Leave>', _hover_out)

        def _abrir_calendario(nome=titulo, dia_atual=dia_mod):
            # Determina o mês a exibir: mês seguinte à última extração
            try:
                ultimo = datetime.strptime(ultima, '%d/%m/%Y') if ultima != '—' else date.today()
                if isinstance(ultimo, datetime):
                    ultimo = ultimo.date()
                if ultimo.month == 12:
                    cal_ano, cal_mes = ultimo.year + 1, 1
                else:
                    cal_ano, cal_mes = ultimo.year, ultimo.month + 1
            except Exception:
                hoje2 = date.today()
                if hoje2.month == 12:
                    cal_ano, cal_mes = hoje2.year + 1, 1
                else:
                    cal_ano, cal_mes = hoje2.year, hoje2.month + 1

            _, num_dias = monthrange(cal_ano, cal_mes)

            # Popup
            popup = tk.Toplevel(self.root)
            popup.title(f'Agendar — {nome}')
            popup.resizable(False, False)
            popup.configure(bg=COR_BG)
            popup.grab_set()

            # Centraliza na tela
            popup.update_idletasks()
            pw, ph = 260, 290
            sx = self.root.winfo_x() + (self.root.winfo_width()  - pw) // 2
            sy = self.root.winfo_y() + (self.root.winfo_height() - ph) // 2
            popup.geometry(f'{pw}x{ph}+{sx}+{sy}')

            tk.Frame(popup, bg=cor, height=4).pack(fill='x')
            tk.Label(popup, text=f'📅  {nome}',
                     font=('Segoe UI', 10, 'bold'),
                     bg=COR_BG, fg=COR_TEXTO).pack(pady=(10, 2))

            _MESES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                         'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
            tk.Label(popup,
                     text=f'{_MESES_PT[cal_mes-1]} de {cal_ano}',
                     font=('Segoe UI', 9, 'bold'),
                     bg=COR_BG, fg=cor).pack(pady=(0, 6))

            frm_cal = tk.Frame(popup, bg=COR_BG)
            frm_cal.pack(padx=16)

            # Cabeçalho dias da semana
            _DIAS_CAB = ['S', 'T', 'Q', 'Q', 'S', 'S', 'D']
            frm_hdr = tk.Frame(frm_cal, bg=COR_BG)
            frm_hdr.pack()
            for d in _DIAS_CAB:
                tk.Label(frm_hdr, text=d, font=('Segoe UI', 8, 'bold'),
                         bg=COR_BG, fg=COR_TEXTO_SUB, width=3).pack(side='left')

            # Grade de dias
            primeiro_dia = date(cal_ano, cal_mes, 1).weekday()
            dia_celulas  = [None] * primeiro_dia + list(range(1, num_dias + 1))
            while len(dia_celulas) % 7 != 0:
                dia_celulas.append(None)

            _sel = tk.IntVar(value=dia_atual if dia_atual <= num_dias else 1)
            dia_labels = {}

            def _selecionar(d):
                for dd, lb in dia_labels.items():
                    lb.config(bg=COR_CARD if dd != d else cor,
                              fg=COR_TEXTO if dd != d else 'white',
                              font=('Segoe UI', 8, 'bold') if dd == d else ('Segoe UI', 8))
                _sel.set(d)

            for i in range(0, len(dia_celulas), 7):
                frm_row = tk.Frame(frm_cal, bg=COR_BG)
                frm_row.pack()
                for dia in dia_celulas[i:i+7]:
                    if dia is None:
                        tk.Label(frm_row, text='', width=3, bg=COR_BG).pack(side='left')
                    else:
                        eh_sel = (dia == _sel.get())
                        lbl = tk.Label(frm_row, text=str(dia),
                                       font=('Segoe UI', 8, 'bold') if eh_sel else ('Segoe UI', 8),
                                       bg=cor if eh_sel else COR_CARD,
                                       fg='white' if eh_sel else COR_TEXTO,
                                       width=3, cursor='hand2')
                        lbl.bind('<Button-1>', lambda e, d=dia: _selecionar(d))
                        lbl.bind('<Enter>',   lambda e, lb=lbl, d=dia:
                                 lb.config(bg=cor, fg='white') if d != _sel.get() else None)
                        lbl.bind('<Leave>',   lambda e, lb=lbl, d=dia:
                                 lb.config(bg=COR_CARD, fg=COR_TEXTO) if d != _sel.get() else None)
                        lbl.pack(side='left', pady=1)
                        dia_labels[dia] = lbl

            def _confirmar():
                d = _sel.get()
                if not 1 <= d <= 28:
                    messagebox.showwarning('Atenção',
                        'Selecione um dia entre 1 e 28 para garantir que ocorra em todos os meses.',
                        parent=popup)
                    return
                est = _carregar_estado()
                dias = est.get('dias_extracao_modulos', {})
                dias[nome] = d
                est['dias_extracao_modulos'] = dias
                _salvar_estado(est)
                popup.destroy()
                self._ir_para('home')

            frm_btns = tk.Frame(popup, bg=COR_BG)
            frm_btns.pack(pady=(10, 12))

            btn_c = tk.Label(frm_btns, text='✔  Confirmar',
                             font=('Segoe UI', 9, 'bold'),
                             bg=cor, fg='white', padx=12, pady=5, cursor='hand2')
            btn_c.bind('<Button-1>', lambda e: _confirmar())
            btn_c.bind('<Enter>',   lambda e: btn_c.config(bg=COR_ACCENT))
            btn_c.bind('<Leave>',   lambda e: btn_c.config(bg=cor))
            btn_c.pack(side='left', padx=6)

            btn_x = tk.Label(frm_btns, text='✕  Cancelar',
                             font=('Segoe UI', 9),
                             bg=COR_CARD2, fg=COR_TEXTO_SUB,
                             padx=12, pady=5, cursor='hand2')
            btn_x.bind('<Button-1>', lambda e: popup.destroy())
            btn_x.bind('<Enter>',   lambda e: btn_x.config(fg=COR_TEXTO))
            btn_x.bind('<Leave>',   lambda e: btn_x.config(fg=COR_TEXTO_SUB))
            btn_x.pack(side='left', padx=6)

        btn_agendar.bind('<Button-1>', lambda e: _abrir_calendario())

    # ----------------------------------------------------------
    # RELATÓRIOS
    # ----------------------------------------------------------
    def _pagina_relatorios(self):
        c = self.frm_conteudo

        tk.Label(c, text='📊  Relatórios',
                 font=('Segoe UI', 18, 'bold'),
                 bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', padx=32, pady=(28, 4))
        tk.Label(c, text='Selecione um módulo para gerar o relatório',
                 font=('Segoe UI', 10), bg=COR_BG,
                 fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 24))

        # Frame que expande e distribui os cards igualmente
        frm = tk.Frame(c, bg=COR_BG)
        frm.pack(fill='both', expand=True, padx=32, pady=(0, 32))

        modulos = [
            ('📦', 'Pedidos e\nRecebimentos', 'Análise de SLA · D+0 · D+1 · Excedidos', COR_ACCENT,  lambda: self._ir_para('pedidos')),
            ('🚚', 'Fretes',                 'Embarques · RESCOM · Portadores · Insumos', COR_ACCENT2, lambda: self._ir_para('fretes')),
            ('🏭', 'Armazenagem',            'Faturamento mensal por cliente',            COR_ACCENT3, lambda: self._ir_para('armazenagem')),
            ('📋', 'Estoque',                  'Volume ocupado · Ociosidade por cliente',              COR_ACCENT4, lambda: self._ir_para('estoque')),
            ('👥', 'Produtividade\nde Equipe', 'Utilização · Ociosidade · Ranking por etapa',          COR_ACCENT5, lambda: self._ir_para('produtividade')),
            ('⚙️', 'Capacidade\nOperacional',  'OS por depositante · SKUs · Extração do ESL',          COR_ACCENT7, lambda: self._ir_para('cap_operacional')),
            ('📥', 'Recebimentos\ne Devoluções',   'Entradas · Devoluções · Retiradas por depositante',   COR_ACCENT8, lambda: self._ir_para('recebimentos')),
        ]

        for i in range(len(modulos)):
            frm.columnconfigure(i, weight=1, uniform='card')
        frm.rowconfigure(0, weight=1)

        for i, (icone, titulo, desc, cor, cmd) in enumerate(modulos):
            self._card_relatorio(frm, icone, titulo, desc, cor, cmd, i)

    def _card_relatorio(self, parent, icone, titulo, desc, cor, cmd, col):
        frm = tk.Frame(parent, bg=COR_CARD,
                       highlightbackground=COR_BORDA, highlightthickness=1)
        frm.grid(row=0, column=col, padx=12, pady=8, sticky='nsew')

        tk.Frame(frm, bg=cor, height=6).pack(fill='x')

        # Conteúdo centralizado verticalmente
        frm_inner = tk.Frame(frm, bg=COR_CARD)
        frm_inner.pack(expand=True, fill='both', pady=20)

        tk.Label(frm_inner, text=icone, font=('Segoe UI', 42),
                 bg=COR_CARD, fg=cor).pack(pady=(0, 8))
        tk.Label(frm_inner, text=titulo, font=('Segoe UI', 14, 'bold'),
                 bg=COR_CARD, fg=COR_TEXTO, justify='center').pack()
        tk.Label(frm_inner, text=desc, font=('Segoe UI', 9),
                 bg=COR_CARD, fg=COR_TEXTO_SUB,
                 justify='center', wraplength=220).pack(pady=(6, 24))

        btn = tk.Button(frm_inner, text='  Abrir Módulo  →',
                        font=('Segoe UI', 11, 'bold'),
                        bg=cor, fg='white', relief='flat',
                        padx=24, pady=12, cursor='hand2',
                        command=cmd)
        btn.pack()
        btn.bind('<Enter>', lambda e: btn.config(bg=_escurecer(cor)))
        btn.bind('<Leave>', lambda e: btn.config(bg=cor))
        frm.bind('<Enter>', lambda e: frm.config(highlightbackground=cor))
        frm.bind('<Leave>', lambda e: frm.config(highlightbackground=COR_BORDA))

    # ----------------------------------------------------------
    # ARQUIVOS
    # ----------------------------------------------------------
    def _buscar_arquivos_modulo(self, nome, cfg):
        """Busca todos os arquivos xlsx do módulo na pasta e subpastas."""
        pasta = cfg['pasta']
        base  = cfg.get('arquivo_base', '')
        arquivos = []
        if not os.path.isdir(pasta):
            return arquivos
        for dirpath, _, files in os.walk(pasta):
            for f in files:
                if f.endswith('.xlsx') and f.startswith(base):
                    caminho = os.path.join(dirpath, f)
                    mtime   = os.path.getmtime(caminho)
                    arquivos.append({
                        'nome':    f,
                        'caminho': caminho,
                        'mtime':   mtime,
                        'dt':      datetime.fromtimestamp(mtime),
                        'tamanho': os.path.getsize(caminho) / 1024,
                    })
        arquivos.sort(key=lambda x: x['mtime'], reverse=True)
        return arquivos

    def _pagina_arquivos(self):
        c = self.frm_conteudo
        self._filtro_periodo = tk.StringVar(value='')

        # Cabeçalho
        frm_h = tk.Frame(c, bg=COR_BG)
        frm_h.pack(fill='x', padx=32, pady=(28, 4))
        tk.Label(frm_h, text='📁  Arquivos Gerados',
                 font=('Segoe UI', 18, 'bold'),
                 bg=COR_BG, fg=COR_TEXTO).pack(side='left')

        # Filtro por período
        frm_f = tk.Frame(frm_h, bg=COR_CARD2,
                         highlightbackground=COR_BORDA, highlightthickness=1)
        frm_f.pack(side='right')
        tk.Label(frm_f, text='Filtrar por período (MM/AAAA):',
                 font=('Segoe UI', 8), bg=COR_CARD2,
                 fg=COR_TEXTO_SUB).pack(side='left', padx=(10, 4), pady=6)
        entry_filtro = estilo_entry(frm_f, width=10)
        entry_filtro.pack(side='left', ipady=4, pady=6)
        tk.Button(frm_f, text='🔍', font=('Segoe UI', 9),
                  bg=COR_ACCENT, fg='white', relief='flat',
                  padx=8, pady=4, cursor='hand2',
                  command=lambda: self._atualizar_lista_arquivos(
                      frm_lista, entry_filtro.get().strip()
                  )).pack(side='left', padx=(4, 4), pady=6)
        tk.Button(frm_f, text='✕', font=('Segoe UI', 9),
                  bg=COR_CARD2, fg=COR_TEXTO_SUB, relief='flat',
                  padx=6, pady=4, cursor='hand2',
                  command=lambda: [entry_filtro.delete(0, tk.END),
                                   self._atualizar_lista_arquivos(frm_lista, '')]
                  ).pack(side='left', padx=(0, 8), pady=6)

        tk.Label(c, text='Clique em um arquivo para abrir no Excel ou visualizar como PDF',
                 font=('Segoe UI', 10), bg=COR_BG,
                 fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 8))

        # Status feedback
        self._lbl_status = tk.Label(c, text='', font=('Segoe UI', 9),
                                    bg=COR_BG, fg=COR_ACCENT3)
        self._lbl_status.pack(anchor='w', padx=32)

        # Área scrollável
        frm_outer = tk.Frame(c, bg=COR_BG)
        frm_outer.pack(fill='both', expand=True, padx=28, pady=(4, 16))

        canvas = tk.Canvas(frm_outer, bg=COR_BG, highlightthickness=0)
        sb     = ttk.Scrollbar(frm_outer, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        frm_lista = tk.Frame(canvas, bg=COR_BG)
        canvas_win = canvas.create_window((0, 0), window=frm_lista, anchor='nw')

        def _on_resize(e):
            canvas.itemconfig(canvas_win, width=e.width)
        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox('all'))
        canvas.bind('<Configure>', _on_resize)
        frm_lista.bind('<Configure>', _on_frame_configure)
        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>',
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), 'units')))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

        self._frm_lista  = frm_lista
        self._lbl_status_ref = self._lbl_status
        self._atualizar_lista_arquivos(frm_lista, '')

    def _atualizar_lista_arquivos(self, frm_lista, filtro):
        for w in frm_lista.winfo_children():
            w.destroy()

        total_encontrados = 0

        for nome, cfg in RELATORIOS_CONFIG.items():
            arquivos = self._buscar_arquivos_modulo(nome, cfg)

            # Aplica filtro por nome do arquivo (MM-AAAA no nome)
            if filtro:
                try:
                    # Aceita MM/AAAA ou MM-AAAA
                    filtro_norm = filtro.strip().replace('/', '-')
                    arquivos = [a for a in arquivos if filtro_norm in a['nome']]
                except (AttributeError, TypeError):
                    pass

            # Cabeçalho do módulo
            frm_mod = tk.Frame(frm_lista, bg=COR_BG)
            frm_mod.pack(fill='x', pady=(12, 4))
            tk.Frame(frm_mod, bg=cfg['cor'], width=4).pack(side='left', fill='y')
            tk.Label(frm_mod,
                     text=f"{cfg['icone']}  {nome}  —  {len(arquivos)} arquivo(s)",
                     font=('Segoe UI', 11, 'bold'),
                     bg=COR_BG, fg=COR_TEXTO).pack(side='left', padx=10)

            if not arquivos:
                tk.Label(frm_lista,
                         text='   Nenhum arquivo encontrado para este módulo.',
                         font=('Segoe UI', 9), bg=COR_BG,
                         fg=COR_TEXTO_SUB).pack(anchor='w', padx=20, pady=2)
            else:
                for arq in arquivos:
                    self._card_arquivo(frm_lista, arq, cfg['cor'])
                    total_encontrados += 1

        if total_encontrados == 0 and filtro:
            tk.Label(frm_lista,
                     text=f'Nenhum arquivo encontrado para o período "{filtro}".',
                     font=('Segoe UI', 10), bg=COR_BG,
                     fg=COR_AVISO).pack(pady=20)

    def _card_arquivo(self, parent, arq, cor):
        frm = tk.Frame(parent, bg=COR_CARD,
                       highlightbackground=COR_BORDA, highlightthickness=1)
        frm.pack(fill='x', padx=20, pady=3)

        tk.Frame(frm, bg=cor, width=4).pack(side='left', fill='y')

        frm_info = tk.Frame(frm, bg=COR_CARD)
        frm_info.pack(side='left', fill='both', expand=True, padx=14, pady=10)

        tk.Label(frm_info, text=arq['nome'],
                 font=('Segoe UI', 10, 'bold'),
                 bg=COR_CARD, fg=COR_TEXTO, anchor='w').pack(anchor='w')
        tk.Label(frm_info,
                 text=f"Gerado em {arq['dt'].strftime('%d/%m/%Y às %H:%M')}  ·  {arq['tamanho']:.0f} KB",
                 font=('Segoe UI', 8),
                 bg=COR_CARD, fg=COR_TEXTO_SUB, anchor='w').pack(anchor='w')

        frm_btns = tk.Frame(frm, bg=COR_CARD)
        frm_btns.pack(side='right', padx=12, pady=8)

        tk.Button(frm_btns, text='📊 Excel',
                  font=('Segoe UI', 8, 'bold'),
                  bg=COR_CARD2, fg=COR_TEXTO, relief='flat',
                  padx=10, pady=5, cursor='hand2',
                  command=lambda p=arq['caminho']: _abrir_excel(p)
                  ).pack(side='left', padx=(0, 6))

        tk.Button(frm_btns, text='📄 PDF',
                  font=('Segoe UI', 8, 'bold'),
                  bg=cor, fg='white', relief='flat',
                  padx=10, pady=5, cursor='hand2',
                  command=lambda p=arq['caminho']: threading.Thread(
                      target=_exportar_pdf,
                      args=(p, lambda m: self.root.after(
                          0, self._lbl_status.config, {'text': m})),
                      daemon=True).start()
                  ).pack(side='left', padx=(0, 6))

        def _excluir(p=arq['caminho'], card=frm):
            if messagebox.askyesno('Excluir arquivo',
                    f'Tem certeza que deseja excluir:\n{os.path.basename(p)}?',
                    icon='warning'):
                try:
                    os.remove(p)
                    card.destroy()
                    self._lbl_status.config(
                        text=f'🗑️ {os.path.basename(p)} excluído.', fg=COR_AVISO)
                except Exception as e:
                    messagebox.showerror('Erro', f'Não foi possível excluir:\n{e}')

        tk.Button(frm_btns, text='🗑️',
                  font=('Segoe UI', 10),
                  bg=COR_CARD2, fg=COR_ERRO, relief='flat',
                  padx=8, pady=5, cursor='hand2',
                  command=_excluir
                  ).pack(side='left')


    # ----------------------------------------------------------
    # HELPERS DE PÁGINA
    # ----------------------------------------------------------
    def _cabecalho(self, titulo, subtitulo, cor):
        tk.Frame(self.frm_conteudo, bg=cor, height=5).pack(fill='x')
        tk.Label(self.frm_conteudo, text=titulo,
                 font=('Segoe UI', 18, 'bold'),
                 bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', padx=32, pady=(20, 2))
        tk.Label(self.frm_conteudo, text=subtitulo,
                 font=('Segoe UI', 10), bg=COR_BG,
                 fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 10))

    def _separador(self):
        tk.Frame(self.frm_conteudo, bg=COR_BORDA, height=1).pack(fill='x', pady=8)

    def _file_row(self, pasta_padrao, label='Arquivo de entrada:', config_key=None,
                  filetypes=None):
        tk.Label(self.frm_conteudo, text=label, font=('Segoe UI', 9),
                 bg=COR_BG, fg=COR_TEXTO_SUB).pack(anchor='w', padx=32)
        frm = tk.Frame(self.frm_conteudo, bg=COR_BG)
        frm.pack(fill='x', padx=32, pady=(2, 10))
        entry = estilo_entry(frm, width=62)
        entry.pack(side='left', ipady=5)
        _filetypes = filetypes or [('Excel', '*.xlsx *.xls')]
        def _pick():
            estado = _carregar_estado()
            chave_entrada = f'entrada_{config_key}' if config_key else None
            pasta = (estado.get('config_pastas', {}).get(chave_entrada)
                     or pasta_padrao)
            p = filedialog.askopenfilename(
                initialdir=_pasta_inicial(pasta),
                filetypes=_filetypes)
            if p:
                entry.delete(0, tk.END)
                entry.insert(0, p)
                if chave_entrada:
                    est = _carregar_estado()
                    est.setdefault('config_pastas', {})[chave_entrada] = os.path.dirname(p)
                    _salvar_estado(est)
        def _limpar():
            entry.delete(0, tk.END)
        tk.Button(frm, text='📁', font=('Segoe UI', 11), bg=COR_CARD2,
                  fg=COR_TEXTO, relief='flat', padx=8, pady=4,
                  cursor='hand2', command=_pick).pack(side='left', padx=(6, 0))
        btn_limpar = tk.Button(frm, text='✕', font=('Segoe UI', 10), bg=COR_CARD2,
                               fg=COR_TEXTO_SUB, relief='flat', padx=8, pady=4,
                               cursor='hand2', command=_limpar)
        btn_limpar.pack(side='left', padx=(4, 0))
        btn_limpar.bind('<Enter>', lambda e: btn_limpar.config(fg=COR_ERRO))
        btn_limpar.bind('<Leave>', lambda e: btn_limpar.config(fg=COR_TEXTO_SUB))
        return entry

    def _log_area(self):
        frm_log, txt = criar_log(self.frm_conteudo)
        frm_log.pack(fill='both', expand=True, padx=32, pady=(0, 8))
        return txt

    def _run_btn(self, label, cor):
        btn = estilo_btn(self.frm_conteudo, label, cor, lambda: None, width=24)
        btn.pack(pady=(0, 16))
        return btn

    def _log_e_btn(self, label, cor):
        """
        Cria log + botão numa estrutura onde o botão fica fixo na base
        e o log preenche todo o espaço restante sem comprimir o botão.
        Retorna (txt, btn).
        """
        # Frame que ocupa todo o espaço restante
        frm_outer = tk.Frame(self.frm_conteudo, bg=COR_BG)
        frm_outer.pack(fill='both', expand=True, padx=32, pady=(0, 16))

        # Botão fixo na base
        btn = estilo_btn(frm_outer, label, cor, lambda: None, width=24)
        btn.pack(side='bottom', pady=(8, 0))

        # Log preenche o espaço acima do botão
        frm_log, txt = criar_log(frm_outer)
        frm_log.pack(fill='both', expand=True, pady=(0, 0))

        return txt, btn

    # ----------------------------------------------------------
    # DASHBOARD
    # ----------------------------------------------------------
    def _pagina_dashboard(self):
        self._cabecalho('📈  Dashboard Consolidado',
                        'Visão executiva de todos os módulos por mês', COR_ACCENT6)
        self._separador()
        c = self.frm_conteudo

        # ── Linha: seletor de mês + botão Gerar (alinhados à esquerda) ─────────
        frm_topo = tk.Frame(c, bg=COR_BG)
        frm_topo.pack(anchor='w', padx=32, pady=(6, 4))

        tk.Label(frm_topo, text='Mês (MM-AAAA):',
                 font=('Segoe UI', 9), bg=COR_BG, fg=COR_TEXTO_SUB).pack(side='left', padx=(0, 8))

        hoje    = date.today()
        mes_ref = f'{(hoje.month - 1) or 12:02d}-{hoje.year if hoje.month > 1 else hoje.year - 1}'
        entry_mes = estilo_entry(frm_topo, width=12)
        entry_mes.insert(0, mes_ref)
        entry_mes.pack(side='left', ipady=5)

        tk.Label(frm_topo, text='(ex: 03-2025)',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB).pack(side='left', padx=(6, 20))

        btn = estilo_btn(frm_topo, '📈  Gerar Dashboard', COR_ACCENT6, lambda: None, width=22)
        btn.pack(side='left')

        self._separador()

        # ── Overrides manuais por módulo ─────────────────────────────────────
        tk.Label(c, text='Substituição manual de arquivo (opcional):',
                 font=('Segoe UI', 9, 'bold'), bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', padx=32, pady=(4, 2))
        tk.Label(c, text='Deixe em branco para usar o arquivo detectado automaticamente pelo mês.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 6))

        _MODULOS_LABEL = [
            ('pedidos',       '📦 Pedidos',       PASTA_PEDIDOS,      'pasta_pedidos'),
            ('fretes',        '🚚 Fretes',         PASTA_FRETES,       'pasta_fretes'),
            ('armazenagem',   '🏭 Armazenagem',    PASTA_ARMAZENAGEM,  'pasta_armazenagem'),
            ('estoque',       '📋 Estoque',        PASTA_ESTOQUE,      'pasta_estoque'),
            ('produtividade', '👥 Produtividade',  PASTA_PRODUTIVIDADE,'pasta_produtividade'),
        ]
        override_entries = {}
        for key, label, pasta_def, cfg_key in _MODULOS_LABEL:
            frm_row = tk.Frame(c, bg=COR_BG)
            frm_row.pack(fill='x', padx=32, pady=2)
            tk.Label(frm_row, text=label, font=('Segoe UI', 9), width=18, anchor='w',
                     bg=COR_BG, fg=COR_TEXTO_SUB).pack(side='left')
            ent = estilo_entry(frm_row, width=50)
            ent.pack(side='left', ipady=4)
            def _pick(e=ent, p=pasta_def, k=cfg_key):
                estado_cfg = _carregar_estado().get('config_pastas', {})
                pasta_real = estado_cfg.get(k, p)
                path = filedialog.askopenfilename(
                    initialdir=_pasta_inicial(pasta_real),
                    filetypes=[('Excel', '*.xlsx')])
                if path:
                    e.delete(0, tk.END)
                    e.insert(0, path)
            tk.Button(frm_row, text='📁', font=('Segoe UI', 10), bg=COR_CARD2,
                      fg=COR_TEXTO, relief='flat', padx=6, pady=2,
                      cursor='hand2', command=_pick).pack(side='left', padx=(4, 0))
            btn_x = tk.Button(frm_row, text='✕', font=('Segoe UI', 10), bg=COR_CARD2,
                               fg=COR_TEXTO_SUB, relief='flat', padx=6, pady=2,
                               cursor='hand2', command=lambda e=ent: e.delete(0, tk.END))
            btn_x.pack(side='left', padx=(4, 0))
            btn_x.bind('<Enter>', lambda e, b=btn_x: b.config(fg=COR_ERRO))
            btn_x.bind('<Leave>', lambda e, b=btn_x: b.config(fg=COR_TEXTO_SUB))
            override_entries[key] = ent

        self._separador()

        # ── Log (altura fixa, não expande) ───────────────────────────────────
        frm_log, txt_log = criar_log(c)
        frm_log.pack(fill='x', padx=32, pady=(0, 4))
        frm_log.configure(height=90)
        frm_log.pack_propagate(False)

        # ── Área de preview com scroll (cards + tabelas) ─────────────────────
        frm_preview_outer = tk.Frame(c, bg=COR_BG)
        frm_preview_outer.pack(fill='both', expand=True, padx=28, pady=(4, 0))

        canvas_prev = tk.Canvas(frm_preview_outer, bg=COR_BG, highlightthickness=0)
        sb_prev     = ttk.Scrollbar(frm_preview_outer, orient='vertical', command=canvas_prev.yview)
        canvas_prev.configure(yscrollcommand=sb_prev.set)
        sb_prev.pack(side='right', fill='y')
        canvas_prev.pack(side='left', fill='both', expand=True)

        frm_preview = tk.Frame(canvas_prev, bg=COR_BG)
        cw_prev = canvas_prev.create_window((0, 0), window=frm_preview, anchor='nw')

        canvas_prev.bind('<Configure>', lambda e: canvas_prev.itemconfig(cw_prev, width=e.width))
        frm_preview.bind('<Configure>', lambda e: canvas_prev.configure(scrollregion=canvas_prev.bbox('all')))
        canvas_prev.bind('<Enter>', lambda e: canvas_prev.bind_all('<MouseWheel>',
            lambda e: canvas_prev.yview_scroll(int(-1*(e.delta/120)), 'units')))
        canvas_prev.bind('<Leave>', lambda e: canvas_prev.unbind_all('<MouseWheel>'))

        self._dash_preview_frame = frm_preview

        # ── Lógica do botão ──────────────────────────────────────────────────
        def _run():
            mes_str = entry_mes.get().strip()
            if not re.match(r'^\d{2}-\d{4}$', mes_str):
                messagebox.showwarning('Atenção', 'Formato inválido. Use MM-AAAA (ex: 03-2025).')
                return

            overrides  = {k: v.get().strip() for k, v in override_entries.items() if v.get().strip()}
            estado_cfg = _carregar_estado()
            pasta_out  = estado_cfg.get('config_pastas', {}).get('pasta_dashboards', r'Z:\GUSTAVO\App\Dashboards')
            if not os.path.isdir(pasta_out):
                pasta_out = os.path.expanduser('~')

            txt_log.configure(state='normal'); txt_log.delete('1.0', tk.END); txt_log.configure(state='disabled')
            log_write(txt_log, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*50}\n")
            btn.configure(state='disabled', text='⏳  Carregando...')

            for w in frm_preview.winfo_children():
                w.destroy()

            def _t():
                dados = _dash_coletar_dados(mes_str, overrides,
                                            lambda m: self.root.after(0, log_write, txt_log, m))
                caminho    = None
                modulos_ok = [k for k, v in dados.items() if v is not None]
                if modulos_ok:
                    log_fn   = lambda m: self.root.after(0, log_write, txt_log, m)
                    pasta_ok = estado_cfg.get('config_pastas', {}).get('pasta_dashboards', r'Z:\GUSTAVO\App\Dashboards')
                    if not os.path.isdir(pasta_ok):
                        pasta_ok = os.path.expanduser('~')
                    caminho = _dash_exportar_excel(dados, mes_str, pasta_ok, log_fn)

                self.root.after(0, lambda: btn.configure(state='normal', text='📈  Gerar Dashboard'))
                self.root.after(0, lambda: self._dash_renderizar_preview(frm_preview, dados, mes_str))
                if caminho:
                    self.root.after(0, lambda: log_write(txt_log, '\n🎉 Concluído!\n'))
                    self.root.after(0, lambda: self._resumo_sucesso('Dashboard', caminho, txt_log))
                else:
                    self.root.after(0, lambda: messagebox.showerror(
                        'Erro', '❌ Nenhum módulo disponível. Verifique o log.'))

            threading.Thread(target=_t, daemon=True).start()

        btn.configure(command=_run)


    def _dash_renderizar_preview(self, frm, dados, mes_str):
        """Renderiza cards de KPI no topo + tabelas resumidas por módulo abaixo."""
        for w in frm.winfo_children():
            w.destroy()

        mm, aaaa   = mes_str.split('-')
        _MESES_PT  = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                      'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
        mes_label  = f"{_MESES_PT[int(mm)-1]} / {aaaa}"

        # ── Título ───────────────────────────────────────────────────────────
        tk.Label(frm, text=f'Dashboard — {mes_label}',
                 font=('Segoe UI', 13, 'bold'), bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', pady=(8, 6), padx=4)

        # ════════════════════════════════════════════════════════════════════
        # BLOCO 1 — CARDS de KPI
        # ════════════════════════════════════════════════════════════════════
        cards_def = []
        if dados.get('pedidos'):
            d = dados['pedidos']
            cor_v = COR_ACCENT3 if d['sla_pct'] >= 90 else (COR_AVISO if d['sla_pct'] >= 70 else COR_ERRO)
            cards_def.append(('📦', 'SLA do Mês',        f"{d['sla_pct']:.1f}%",
                               f"{d['total_ordens']} ordens · {d['excedidas']} excedidas",
                               COR_ACCENT, cor_v))
        if dados.get('fretes'):
            d = dados['fretes']
            cards_def.append(('🚚', 'Custo Fretes',      f"R$ {d['total_frete']:,.0f}",
                               f"{len(d['por_remetente'])} remetentes",
                               COR_ACCENT2, COR_TEXTO))
        if dados.get('armazenagem'):
            d = dados['armazenagem']
            cards_def.append(('🏭', 'Armazenagem',       f"R$ {d['total_armazenagem']:,.0f}",
                               f"{len(d['por_cliente'])} clientes",
                               COR_ACCENT3, COR_TEXTO))
        if dados.get('produtividade'):
            d = dados['produtividade']
            cor_v = COR_ACCENT3 if d['media_utilizacao'] >= 80 else (COR_AVISO if d['media_utilizacao'] >= 60 else COR_ERRO)
            top_n = d['top_colaboradores'][0]['nome'] if d['top_colaboradores'] else '—'
            cards_def.append(('👥', 'Utilização Média',  f"{d['media_utilizacao']:.1f}%",
                               f"Top: {top_n}", COR_ACCENT5, cor_v))
        if dados.get('estoque'):
            d = dados['estoque']
            top_n = d['por_cliente'][0]['nome'] if d['por_cliente'] else '—'
            cards_def.append(('📋', 'Pico Estoque',      f"{d['pico_total_m3']:,.1f} m³",
                               f"Top: {top_n}", COR_ACCENT4, COR_TEXTO))

        if not cards_def:
            tk.Label(frm, text='Nenhum dado disponível.',
                     font=('Segoe UI', 9), bg=COR_BG, fg=COR_TEXTO_SUB).pack(pady=12)
            return

        frm_cards = tk.Frame(frm, bg=COR_BG)
        frm_cards.pack(fill='x', pady=(0, 16), padx=4)
        for i, (icone, titulo, valor, sub, cor_b, cor_v) in enumerate(cards_def):
            frm_cards.columnconfigure(i, weight=1, uniform='dc')
            card = tk.Frame(frm_cards, bg=COR_CARD,
                            highlightbackground=cor_b, highlightthickness=2)
            card.grid(row=0, column=i, padx=6, pady=4, sticky='nsew')
            tk.Frame(card, bg=cor_b, height=4).pack(fill='x')
            tk.Label(card, text=icone, font=('Segoe UI', 20),
                     bg=COR_CARD, fg=cor_b).pack(pady=(10, 2))
            tk.Label(card, text=titulo, font=('Segoe UI', 8, 'bold'),
                     bg=COR_CARD, fg=COR_TEXTO_SUB).pack()
            tk.Label(card, text=valor, font=('Segoe UI', 14, 'bold'),
                     bg=COR_CARD, fg=cor_v).pack(pady=(3, 1))
            tk.Label(card, text=sub, font=('Segoe UI', 8),
                     bg=COR_CARD, fg=COR_TEXTO_SUB,
                     wraplength=160, justify='center').pack(pady=(0, 10))

    # ----------------------------------------------------------
    # PEDIDOS
    # ----------------------------------------------------------
    def _pagina_pedidos(self):
        self._cabecalho('📦  Pedidos e Recebimentos',
                        'Análise de SLA com feriados brasileiros', COR_ACCENT)
        self._separador()
        entry_arq = self._file_row(PASTA_PEDIDOS, config_key='pasta_pedidos')
        var_mes   = _dropdown_mes(self.frm_conteudo, cor_label=COR_ACCENT)

        tk.Label(self.frm_conteudo,
                 text='ℹ  A aba de dados é detectada automaticamente.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 10))
        self._separador()

        txt, btn = self._log_e_btn('📦  Gerar Relatório', COR_ACCENT)

        def _run():
            c = entry_arq.get().strip()
            if not c:
                messagebox.showwarning('Atenção', 'Selecione um arquivo.')
                return
            txt.configure(state='normal'); txt.delete('1.0', tk.END); txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Gerando...')
            mes_ref = var_mes.get()
            def _t():
                ok = run_pedidos(c, lambda m: self.root.after(0, log_write, txt, m),
                                 mes_ref=mes_ref)
                self.root.after(0, lambda: btn.configure(state='normal', text='📦  Gerar Relatório'))
                if ok: self.root.after(0, lambda: self._resumo_sucesso('Pedidos e Recebimentos', ok, txt))
                else:  self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
            threading.Thread(target=_t, daemon=True).start()
        btn.configure(command=_run)

    # ----------------------------------------------------------
    # FRETES
    # ----------------------------------------------------------
    def _pagina_fretes(self):
        self._cabecalho('🚚  Fretes',
                        'Embarques · RESCOM · Portadores · Custo de Insumos', COR_ACCENT2)
        self._separador()
        entry_arq = self._file_row(PASTA_FRETES, 'Arquivo de fretes:', config_key='pasta_fretes')
        var_mes   = _dropdown_mes(self.frm_conteudo, cor_label=COR_ACCENT2)

        tk.Label(self.frm_conteudo,
                 text='ℹ  Aba de Embarques, RESCOM e PORTADORES são detectadas automaticamente.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 10))
        self._separador()

        txt, btn = self._log_e_btn('🚚  Gerar Relatório', COR_ACCENT2)

        def _run():
            c = entry_arq.get().strip()
            if not c:
                messagebox.showwarning('Atenção', 'Selecione um arquivo.')
                return
            txt.configure(state='normal'); txt.delete('1.0', tk.END); txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Gerando...')
            mes_ref = var_mes.get()
            def _t():
                ok = run_fretes(c, lambda m: self.root.after(0, log_write, txt, m),
                                mes_ref=mes_ref)
                self.root.after(0, lambda: btn.configure(state='normal', text='🚚  Gerar Relatório'))
                if ok: self.root.after(0, lambda: self._resumo_sucesso('Fretes', ok, txt))
                else:  self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
            threading.Thread(target=_t, daemon=True).start()
        btn.configure(command=_run)

    # ----------------------------------------------------------
    # ARMAZENAGEM
    # ----------------------------------------------------------
    def _pagina_armazenagem(self):
        self._cabecalho('🏭  Armazenagem',
                        'Faturamento mensal por cliente', COR_ACCENT3)
        self._separador()
        entry_arq = self._file_row(PASTA_ARMAZENAGEM, config_key='pasta_armazenagem')
        var_mes   = _dropdown_mes(self.frm_conteudo, cor_label=COR_ACCENT3)
        self._separador()

        txt, btn = self._log_e_btn('🏭  Gerar Relatório', COR_ACCENT3)

        def _run():
            c = entry_arq.get().strip()
            if not c:
                messagebox.showwarning('Atenção', 'Selecione um arquivo.')
                return
            txt.configure(state='normal'); txt.delete('1.0', tk.END); txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Gerando...')
            mes_ref = var_mes.get()
            # Armazenagem usa formato AAAA-MM internamente para filtro de período
            try:
                mm, aaaa = mes_ref.split('-')
                mes_filtro = f'{aaaa}-{mm}'
            except Exception:
                mes_filtro = None
            def _t():
                ok = run_armazenagem(c, mes_filtro,
                                     lambda m: self.root.after(0, log_write, txt, m),
                                     mes_ref=mes_ref)
                self.root.after(0, lambda: btn.configure(state='normal', text='🏭  Gerar Relatório'))
                if ok: self.root.after(0, lambda: self._resumo_sucesso('Armazenagem', ok, txt))
                else:  self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
            threading.Thread(target=_t, daemon=True).start()
        btn.configure(command=_run)



    def _resumo_sucesso(self, modulo, caminho_entrada, txt=None):
        cfg      = RELATORIOS_CONFIG[modulo]
        mes, ano = _mes_ano_referencia()
        mes_ano  = f'{mes:02d}-{ano}'
        # Resolve caminho do arquivo gerado
        if caminho_entrada and os.path.isfile(caminho_entrada):
            caminho_arq = caminho_entrada
        else:
            pasta_cfg    = cfg['pasta']
            subpasta_ano = os.path.join(pasta_cfg, str(ano))
            caminho_arq  = None
            for base in [subpasta_ano, pasta_cfg]:
                if os.path.isdir(base):
                    candidatos = sorted(
                        [f for f in os.listdir(base)
                         if f.endswith('.xlsx') and cfg['arquivo_base'] in f
                         and mes_ano in f],
                        key=lambda f: os.path.getmtime(os.path.join(base, f)),
                        reverse=True
                    )
                    if candidatos:
                        caminho_arq = os.path.join(base, candidatos[0])
                        break

        if not caminho_arq:
            return

        nome_arq = os.path.basename(caminho_arq)
        tamanho  = os.path.getsize(caminho_arq) / 1024

        # Escreve resumo no log
        if txt:
            log_write(txt, f"\n{'─'*52}\n")
            log_write(txt, f"✅ Relatório gerado com sucesso!\n")
            log_write(txt, f"📄 {nome_arq}  ·  {tamanho:.0f} KB\n")
            log_write(txt, f"📁 {os.path.dirname(caminho_arq)}\n\n")

            # Botão embutido no log (widget dentro do Text)
            cor = cfg.get('cor', COR_ACCENT)
            btn_abrir = tk.Button(
                txt,
                text=f'📂  Abrir {nome_arq}',
                font=('Segoe UI', 9, 'bold'),
                bg=cor, fg='white',
                relief='flat', padx=14, pady=6,
                cursor='hand2',
                command=lambda p=caminho_arq: _abrir_excel(p)
            )
            txt.configure(state='normal')
            txt.window_create(tk.END, window=btn_abrir)
            txt.insert(tk.END, '\n')
            txt.configure(state='disabled')
            txt.see(tk.END)

        # Notificação Windows (silenciosa, sem pop-up)
        try:
            ico = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'central_relatorios.ico')
            _notificar_windows(
                f'Relatório de {modulo} gerado!',
                f'{nome_arq}  ·  {tamanho:.0f} KB',
                icone_path=ico if os.path.exists(ico) else None
            )
        except Exception:
            pass


    # ----------------------------------------------------------
    # ESTOQUE
    # ----------------------------------------------------------
    def _pagina_estoque(self):
        self._cabecalho('📋  Estoque',
                        'Volume ocupado · Produtos ociosos por cliente', COR_ACCENT4)
        self._separador()
        c = self.frm_conteudo

        def _dica(txt):
            tk.Label(c, text=f'ℹ  {txt}', font=('Segoe UI', 8),
                     bg=COR_BG, fg=COR_TEXTO_SUB,
                     wraplength=820, justify='left').pack(anchor='w', padx=32, pady=(0,8))

        # ── Painel do Banco de Dados Interno ────────────────────────────
        frm_db = tk.Frame(c, bg=COR_CARD, highlightbackground=COR_BORDA, highlightthickness=1)
        frm_db.pack(fill='x', padx=32, pady=(8,12))

        frm_db_top = tk.Frame(frm_db, bg=COR_CARD)
        frm_db_top.pack(fill='x', padx=12, pady=(10,6))

        tk.Label(frm_db_top, text='🗄️  Banco de Dados Interno de Estoque',
                 font=('Segoe UI', 9, 'bold'), bg=COR_CARD, fg=COR_TEXTO).pack(side='left')

        frm_db_info = tk.Frame(frm_db, bg=COR_CARD)
        frm_db_info.pack(fill='x', padx=12, pady=(0,8))

        lbl_db_status = tk.Label(frm_db_info, text='Carregando...', font=('Segoe UI', 8),
                                  bg=COR_CARD, fg=COR_TEXTO_SUB)
        lbl_db_status.pack(side='left')

        def _atualizar_status_db():
            total, ultima, clientes = _db_estoque_info()
            if total == 0:
                lbl_db_status.config(
                    text='⚠️  DB vazio — faça a Carga Inicial antes de gerar o relatório.',
                    fg=COR_AVISO)
            else:
                lbl_db_status.config(
                    text=f'✅  {total} SKUs cadastrados · {len(clientes)} clientes · Última atualização: {ultima}',
                    fg=COR_ACCENT3)

        _atualizar_status_db()

        frm_db_btns = tk.Frame(frm_db, bg=COR_CARD)
        frm_db_btns.pack(fill='x', padx=12, pady=(0,10))

        def _carga_inicial():
            arq = filedialog.askopenfilename(
                title='Selecionar Relatório de Estoque (carga inicial)',
                filetypes=[('Excel','*.xlsx *.xls')])
            if not arq:
                return
            log_db = []
            def _log_carga(m): log_db.append(m)
            db = _carregar_estoque_xlsx(arq, _log_carga)
            if db:
                _salvar_db_estoque(db)
                total = sum(len(v) for v in db.values())
                _atualizar_status_db()
                # Salva o caminho para reuso automático no Fat. Armazenagem
                est = _carregar_estado()
                est['ultimo_arquivo_estoque'] = arq
                _salvar_estado(est)
                messagebox.showinfo('Carga Inicial',
                    f'✅ {total} SKUs cadastrados em {len(db)} clientes!\n\n' + ''.join(log_db))
            else:
                messagebox.showerror('Erro', 'Nenhum dado encontrado no arquivo.')

        def _atualizar_db():
            arq = filedialog.askopenfilename(
                title='Selecionar Movimentação de Estoque (atualização)',
                filetypes=[('Excel','*.xlsx *.xls')])
            if not arq:
                return
            log_db = []
            def _log_atu(m): log_db.append(m)
            ok = _atualizar_db_com_movimentacao(arq, _log_atu)
            _atualizar_status_db()
            if ok:
                # Salva o caminho para reuso no Fat. Armazenagem
                est = _carregar_estado()
                est['ultimo_arquivo_mov'] = arq
                _salvar_estado(est)
                messagebox.showinfo('DB Atualizado', ''.join(log_db))
            else:
                messagebox.showerror('Erro', ''.join(log_db))

        tk.Button(frm_db_btns, text='📥  Carga Inicial',
                  font=('Segoe UI', 8, 'bold'), bg='#1e3a5f', fg='white',
                  relief='flat', padx=10, pady=5, cursor='hand2',
                  command=_carga_inicial).pack(side='left', padx=(0,8))
        tk.Button(frm_db_btns, text='🔄  Atualizar com Movimentação',
                  font=('Segoe UI', 8, 'bold'), bg='#1e4d3a', fg='white',
                  relief='flat', padx=10, pady=5, cursor='hand2',
                  command=_atualizar_db).pack(side='left')

        def _campo_arquivo(titulo, dica):
            tk.Label(c, text=titulo, font=('Segoe UI', 9, 'bold'),
                     bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', padx=32, pady=(8,0))
            _dica(dica)
            frm = tk.Frame(c, bg=COR_BG)
            frm.pack(fill='x', padx=32, pady=(0,4))
            ent = estilo_entry(frm, width=62)
            ent.pack(side='left', ipady=5)
            def _pick(e=ent):
                p = filedialog.askopenfilename(
                    initialdir=_pasta_inicial(PASTA_ESTOQUE),
                    filetypes=[('Excel','*.xlsx *.xls')])
                if p: e.delete(0,tk.END); e.insert(0,p)
            tk.Button(frm, text='📁', font=('Segoe UI',11), bg=COR_CARD2,
                      fg=COR_TEXTO, relief='flat', padx=8, pady=4,
                      cursor='hand2', command=_pick).pack(side='left', padx=(6,0))
            btn_x = tk.Button(frm, text='✕', font=('Segoe UI',10), bg=COR_CARD2,
                              fg=COR_TEXTO_SUB, relief='flat', padx=8, pady=4,
                              cursor='hand2', command=lambda e=ent: e.delete(0,tk.END))
            btn_x.pack(side='left', padx=(4,0))
            btn_x.bind('<Enter>', lambda e, b=btn_x: b.config(fg=COR_ERRO))
            btn_x.bind('<Leave>', lambda e, b=btn_x: b.config(fg=COR_TEXTO_SUB))
            return ent

        def _campo_pasta(titulo, dica):
            tk.Label(c, text=titulo, font=('Segoe UI', 9, 'bold'),
                     bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', padx=32, pady=(8,0))
            _dica(dica)
            frm = tk.Frame(c, bg=COR_BG)
            frm.pack(fill='x', padx=32, pady=(0,4))
            ent = estilo_entry(frm, width=62)
            ent.pack(side='left', ipady=5)
            def _pick(e=ent):
                p = filedialog.askdirectory(initialdir=_pasta_inicial(PASTA_ESTOQUE))
                if p: e.delete(0,tk.END); e.insert(0,p)
            tk.Button(frm, text='📁', font=('Segoe UI',11), bg=COR_CARD2,
                      fg=COR_TEXTO, relief='flat', padx=8, pady=4,
                      cursor='hand2', command=_pick).pack(side='left', padx=(6,0))
            btn_x = tk.Button(frm, text='✕', font=('Segoe UI',10), bg=COR_CARD2,
                              fg=COR_TEXTO_SUB, relief='flat', padx=8, pady=4,
                              cursor='hand2', command=lambda e=ent: e.delete(0,tk.END))
            btn_x.pack(side='left', padx=(4,0))
            btn_x.bind('<Enter>', lambda e, b=btn_x: b.config(fg=COR_ERRO))
            btn_x.bind('<Leave>', lambda e, b=btn_x: b.config(fg=COR_TEXTO_SUB))
            return ent

        mes_ref = f'{_mes_ano_referencia()[0]:02d}-{_mes_ano_referencia()[1]}'

        def _campo_arquivo_cfg(titulo, dica, cfg_key):
            tk.Label(c, text=titulo, font=('Segoe UI', 9, 'bold'),
                     bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', padx=32, pady=(8,0))
            _dica(dica)
            frm = tk.Frame(c, bg=COR_BG)
            frm.pack(fill='x', padx=32, pady=(0,4))
            ent = estilo_entry(frm, width=62)
            cfg_val = _carregar_estado().get('config_pastas', {}).get(cfg_key, '')
            if cfg_val: ent.insert(0, cfg_val)
            ent.pack(side='left', ipady=5)
            def _pick(e=ent, k=cfg_key):
                cfg_v = _carregar_estado().get('config_pastas', {}).get(k, '')
                ini   = _pasta_inicial(os.path.dirname(cfg_v)) if cfg_v else ''
                p = filedialog.askopenfilename(
                    initialdir=ini,
                    filetypes=[('Excel', '*.xlsx *.xls')])
                if p: e.delete(0, tk.END); e.insert(0, p)
            tk.Button(frm, text='📁', font=('Segoe UI', 11), bg=COR_CARD2,
                      fg=COR_TEXTO, relief='flat', padx=8, pady=4,
                      cursor='hand2', command=_pick).pack(side='left', padx=(6,0))
            btn_x = tk.Button(frm, text='✕', font=('Segoe UI',10), bg=COR_CARD2,
                              fg=COR_TEXTO_SUB, relief='flat', padx=8, pady=4,
                              cursor='hand2', command=lambda e=ent: e.delete(0,tk.END))
            btn_x.pack(side='left', padx=(4,0))
            btn_x.bind('<Enter>', lambda e, b=btn_x: b.config(fg=COR_ERRO))
            btn_x.bind('<Leave>', lambda e, b=btn_x: b.config(fg=COR_TEXTO_SUB))
            return ent

        _MESES_PT = ['Janeiro','Fevereiro','Março','Abril','Maio','Junho',
                     'Julho','Agosto','Setembro','Outubro','Novembro','Dezembro']
        _m, _a   = _mes_ano_referencia()
        mes_nome = _MESES_PT[_m - 1]
        ano_ref  = _a

        entry_vol = _campo_arquivo_cfg(
            'Pico de Estoque por Cliente',
            f'Extrair do sistema o relatório de pico de estoque filtrando por cada depositante. '
            f'Colar cada resultado em uma aba separada do mesmo arquivo Excel e salvar em: '
            f'RELATORIO PICO DE ESTOQUE \\ {mes_nome} {ano_ref}',
            'arquivo_pico_estoque')



        # Dias ocioso — OptionMenu customizado (tema escuro)
        frm_dias = tk.Frame(c, bg=COR_BG)
        frm_dias.pack(anchor='w', padx=32, pady=(8,4))
        tk.Label(frm_dias, text='Considerar ocioso após:',
                 font=('Segoe UI', 9), bg=COR_BG, fg=COR_TEXTO_SUB).pack(side='left', padx=(0,10))

        estado_d = _carregar_estado()
        _opcoes  = [30, 60, 90, 120, 150, 180, 270, 365]
        _atual   = estado_d.get('dias_ocioso', 60)
        if _atual not in _opcoes:
            _atual = 60
        _var_dias = tk.StringVar(value=f'{_atual} dias')

        opt = tk.OptionMenu(frm_dias, _var_dias, *[f'{o} dias' for o in _opcoes])
        opt.config(font=('Segoe UI', 10, 'bold'),
                   bg=COR_CARD2, fg=COR_TEXTO,
                   activebackground=COR_ACCENT, activeforeground='white',
                   highlightthickness=0, relief='flat',
                   indicatoron=True, bd=0, padx=12, pady=6,
                   cursor='hand2')
        opt['menu'].config(font=('Segoe UI', 10),
                           bg=COR_CARD2, fg=COR_TEXTO,
                           activebackground=COR_ACCENT,
                           activeforeground='white',
                           relief='flat', bd=0)
        opt.pack(side='left')

        # alias para o _run usar
        class _SpinAlias:
            def get(self): return str(int(_var_dias.get().replace(' dias', '')))
        spin_dias = _SpinAlias()

        self._separador()
        var_mes_est = _dropdown_mes(c, cor_label=COR_ACCENT4)

        frm_outer_est = tk.Frame(c, bg=COR_BG)
        frm_outer_est.pack(fill='both', expand=True, padx=32, pady=(0, 16))
        btn = estilo_btn(frm_outer_est, '📋  Gerar Relatório de Estoque', COR_ACCENT4, lambda: None, width=28)
        btn.pack(side='bottom', pady=(8, 0))
        frm_log, txt = criar_log(frm_outer_est)
        frm_log.pack(fill='both', expand=True)

        def _run():
            pasta_v = entry_vol.get().strip()
            total_db, _, _ = _db_estoque_info()
            if total_db == 0:
                messagebox.showwarning('DB Vazio',
                    'O banco de dados interno está vazio.\nFaça a Carga Inicial primeiro.')
                return
            try:
                dias = int(spin_dias.get())
            except Exception:
                messagebox.showwarning('Atenção', 'Informe um número válido de dias.')
                return
            est2 = _carregar_estado()
            est2['dias_ocioso'] = dias
            _salvar_estado(est2)
            txt.configure(state='normal'); txt.delete('1.0', tk.END); txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Gerando...')
            mes_ref = var_mes_est.get()
            def _t():
                ok = run_estoque('', pasta_v, '', dias,
                                 lambda m: self.root.after(0, log_write, txt, m),
                                 mes_ref=mes_ref)
                self.root.after(0, lambda: btn.configure(state='normal', text='📋  Gerar Relatório de Estoque'))
                if ok: self.root.after(0, lambda: self._resumo_sucesso('Estoque', ok, txt))
                else:  self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
            threading.Thread(target=_t, daemon=True).start()
        btn.configure(command=_run)

    # ----------------------------------------------------------
    # PRODUTIVIDADE DE EQUIPE
    # ----------------------------------------------------------
    def _pagina_produtividade(self):
        COR_PROD = COR_ACCENT5
        self._cabecalho('👥  Produtividade de Equipe',
                        'Utilização · Ociosidade · Ranking por etapa  ·  Turno 08h–18h (9h úteis)',
                        COR_PROD)
        self._separador()
        c = self.frm_conteudo

        entry_arq = self._file_row(PASTA_PRODUTIVIDADE,
                                   'Arquivo de controle (planilha de OS):',
                                   config_key='pasta_produtividade')

        tk.Label(c,
                 text='ℹ  A aba de dados é detectada automaticamente.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 6))

        var_mes_prod = _dropdown_mes(c, cor_label=COR_PROD)

        tk.Label(c,
                 text='ℹ  Horários sobrepostos são mesclados automaticamente (tempo real por colaborador/dia).\n'
                      '   Campos "QUEM?" com múltiplos nomes (ex: Lemuel/Thaís) creditam tempo completo para cada um.\n'
                      '   Datas com erros de digitação (ano errado, hora fora do turno) são ignoradas automaticamente.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB,
                 justify='left', wraplength=820).pack(anchor='w', padx=32, pady=(0, 8))

        self._separador()
        txt, btn = self._log_e_btn('👥  Gerar Relatório', COR_PROD)

        def _run():
            arq = entry_arq.get().strip()
            if not arq:
                messagebox.showwarning('Atenção', 'Selecione o arquivo de entrada.')
                return
            txt.configure(state='normal'); txt.delete('1.0', tk.END)
            txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Gerando...')
            mes_ref = var_mes_prod.get()
            # Produtividade usa formato AAAA-MM internamente
            try:
                mm, aaaa = mes_ref.split('-')
                mes_filtro = f'{aaaa}-{mm}'
            except Exception:
                mes_filtro = None

            def _t():
                ok = run_produtividade(arq, mes_filtro,
                                       lambda m: self.root.after(0, log_write, txt, m),
                                       mes_ref=mes_ref)
                self.root.after(0, lambda: btn.configure(
                    state='normal', text='👥  Gerar Relatório'))
                if ok:
                    self.root.after(0, lambda: self._resumo_sucesso('Produtividade', ok, txt))
                else:
                    self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
            threading.Thread(target=_t, daemon=True).start()

        btn.configure(command=_run)

    # ----------------------------------------------------------
    # CAPACIDADE OPERACIONAL
    # ----------------------------------------------------------
    def _pagina_cap_operacional(self):
        self._cabecalho('⚙️  Capacidade Operacional',
                        'Análise de OS por Depositante — SKUs e Complexidade', COR_ACCENT7)
        self._separador()
        c = self.frm_conteudo

        # Instruções
        tk.Label(c,
                 text='ℹ  Como extrair o relatório do Sistema ESL:\n'
                      '   1. Acesse Relatório → Movimentação de Estoque\n'
                      '   2. Selecione a opção Kardex\n'
                      '   3. Filtre pelo período desejado\n'
                      '   4. Exporte como PDF e selecione o arquivo abaixo',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB,
                 justify='left').pack(anchor='w', padx=32, pady=(0, 10))

        # Seleção do arquivo PDF
        entry_pdf = self._file_row(
            PASTA_CAP_OPERACIONAL,
            'PDF do Kardex (ESL):',
            filetypes=[('PDF', '*.pdf'), ('Todos os arquivos', '*.*')]
        )

        # Mês de referência
        var_mes = _dropdown_mes(c, label='Mês de referência:', cor_label=COR_ACCENT7)

        # Parâmetros de complexidade
        tk.Label(c, text='Parâmetros de Complexidade (SKUs/OS):',
                 font=('Segoe UI', 9, 'bold'), bg=COR_BG,
                 fg=COR_ACCENT7).pack(anchor='w', padx=32, pady=(4, 4))

        frm_comp = tk.Frame(c, bg=COR_BG)
        frm_comp.pack(anchor='w', padx=32, pady=(0, 10))

        # Baixa → Média
        tk.Label(frm_comp, text='Média a partir de:',
                 font=('Segoe UI', 9), bg=COR_BG,
                 fg=COR_TEXTO_SUB).grid(row=0, column=0, sticky='w', pady=3)
        spin_media = tk.Spinbox(frm_comp, from_=1, to=20, width=4,
                                font=('Segoe UI', 10, 'bold'),
                                bg=COR_CARD2, fg=COR_TEXTO,
                                buttonbackground=COR_CARD2,
                                relief='flat', bd=2)
        spin_media.delete(0, tk.END); spin_media.insert(0, '3')
        spin_media.grid(row=0, column=1, padx=(8, 4))
        tk.Label(frm_comp, text='SKUs/OS',
                 font=('Segoe UI', 9), bg=COR_BG,
                 fg=COR_TEXTO_SUB).grid(row=0, column=2, sticky='w')

        # Média → Alta
        tk.Label(frm_comp, text='Alta a partir de:',
                 font=('Segoe UI', 9), bg=COR_BG,
                 fg=COR_TEXTO_SUB).grid(row=1, column=0, sticky='w', pady=3)
        spin_alta = tk.Spinbox(frm_comp, from_=1, to=50, width=4,
                               font=('Segoe UI', 10, 'bold'),
                               bg=COR_CARD2, fg=COR_TEXTO,
                               buttonbackground=COR_CARD2,
                               relief='flat', bd=2)
        spin_alta.delete(0, tk.END); spin_alta.insert(0, '5')
        spin_alta.grid(row=1, column=1, padx=(8, 4))
        tk.Label(frm_comp, text='SKUs/OS',
                 font=('Segoe UI', 9), bg=COR_BG,
                 fg=COR_TEXTO_SUB).grid(row=1, column=2, sticky='w')

        self._separador()
        txt, btn = self._log_e_btn('⚙️  Gerar Relatório', COR_ACCENT7)

        def _run():
            caminho = entry_pdf.get().strip()
            if not caminho:
                messagebox.showwarning('Atenção', 'Selecione o arquivo PDF do Kardex.')
                return
            if not caminho.lower().endswith('.pdf'):
                messagebox.showwarning('Atenção', 'O arquivo selecionado deve ser um PDF.')
                return
            mes_ref = var_mes.get()
            try:
                lim_media = float(spin_media.get())
                lim_alta  = float(spin_alta.get())
                if lim_media >= lim_alta:
                    messagebox.showwarning('Atenção',
                        'O limiar de "Alta" deve ser maior que o de "Média".')
                    return
            except ValueError:
                messagebox.showwarning('Atenção', 'Informe valores numéricos para os parâmetros.')
                return

            txt.configure(state='normal'); txt.delete('1.0', tk.END); txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Gerando...')

            def _t():
                ok = run_cap_operacional_pdf(
                    caminho, mes_ref,
                    lambda m: self.root.after(0, log_write, txt, m),
                    limiar_media=lim_media,
                    limiar_alta=lim_alta
                )
                self.root.after(0, lambda: btn.configure(state='normal', text='⚙️  Gerar Relatório'))
                if ok:
                    self.root.after(0, lambda: self._resumo_sucesso('Capacidade Operacional', ok, txt))
                else:
                    self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
            threading.Thread(target=_t, daemon=True).start()

        btn.configure(command=_run)

    # ----------------------------------------------------------
    # RECEBIMENTOS E DEVOLUÇÕES
    # ----------------------------------------------------------
    def _pagina_recebimentos(self):
        self._cabecalho('📥  Recebimentos e Devoluções',
                        'Entradas · Devoluções · Retiradas por depositante', COR_ACCENT8)
        self._separador()
        c = self.frm_conteudo

        tk.Label(c,
                 text='ℹ  Como extrair o relatório do Sistema ESL:\n'
                      '   1. Acesse Movimentação → NF de Entrada\n'
                      '   2. Filtre pelo período desejado\n'
                      '   3. Exporte como XLS, converta para Excel (.xlsx) e selecione abaixo\n'
                      '   Ajustes e Manuseios são ignorados automaticamente.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB,
                 justify='left').pack(anchor='w', padx=32, pady=(0, 10))

        entry_arq = self._file_row(
            PASTA_RECEBIMENTOS,
            'Arquivo de Entradas (Excel):',
            filetypes=[('Excel', '*.xlsx *.xls'), ('Todos os arquivos', '*.*')]
        )

        var_mes = _dropdown_mes(c, label='Mês de referência:', cor_label=COR_ACCENT8)

        self._separador()
        txt, btn = self._log_e_btn('📥  Gerar Relatório', COR_ACCENT8)

        def _run():
            caminho = entry_arq.get().strip()
            if not caminho:
                messagebox.showwarning('Atenção', 'Selecione o arquivo de entradas.')
                return
            mes_ref = var_mes.get()
            txt.configure(state='normal'); txt.delete('1.0', tk.END); txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Processando...')

            def _t():
                ok = run_recebimentos(
                    caminho, mes_ref,
                    lambda m: self.root.after(0, log_write, txt, m)
                )
                self.root.after(0, lambda: btn.configure(state='normal', text='📥  Gerar Relatório'))
                if ok:
                    self.root.after(0, lambda: self._resumo_sucesso('Recebimentos e Devoluções', ok, txt))
                else:
                    self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
            threading.Thread(target=_t, daemon=True).start()

        btn.configure(command=_run)

    # ----------------------------------------------------------
    # FINANCEIRO
    # ----------------------------------------------------------
    def _pagina_financeiro(self):
        c = self.frm_conteudo

        tk.Label(c, text='💰  Financeiro',
                 font=('Segoe UI', 18, 'bold'),
                 bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', padx=32, pady=(28, 4))
        tk.Label(c, text='Selecione um módulo financeiro',
                 font=('Segoe UI', 10), bg=COR_BG,
                 fg=COR_TEXTO_SUB).pack(anchor='w', padx=32, pady=(0, 24))

        frm = tk.Frame(c, bg=COR_BG)
        frm.pack(fill='both', expand=True, padx=32, pady=(0, 32))

        modulos = [
            ('💰', 'Contas a Pagar\ne Receber',
             'Títulos · Status · Resumo por categoria',
             COR_RECEBER, lambda: self._ir_para('contas')),
            ('🏭', 'Faturamento\nArmazenagem',
             'Pico m³ por cliente · SKUs na data do pico',
             COR_FAT_ARM, lambda: self._ir_para('fat_arm')),
        ]

        for i in range(len(modulos)):
            frm.columnconfigure(i, weight=1, uniform='card')
        frm.rowconfigure(0, weight=1)

        for i, (icone, titulo, desc, cor, cmd) in enumerate(modulos):
            self._card_relatorio(frm, icone, titulo, desc, cor, cmd, i)

    def _pagina_financeiro_contas(self):
        self._cabecalho('💰  Financeiro',
                        'Contas a Pagar · Contas a Receber', COR_RECEBER)
        self._separador()
        self._build_financeiro_contas(self.frm_conteudo)

    def _build_financeiro_contas(self, c):
        """Conteúdo da sub-aba Contas a Pagar / Receber."""
        BG_CARD = '#0f1117'

        def _entry_row(parent, label, cor_label):
            tk.Label(parent, text=label,
                     font=('Segoe UI', 8, 'bold'),
                     bg=COR_BG, fg=cor_label).pack(anchor='w')
            frm_file = tk.Frame(parent, bg=COR_BG)
            frm_file.pack(fill='x', pady=(2, 0))
            entry = estilo_entry(frm_file, width=52)
            entry.pack(side='left', ipady=5)

            def _pick():
                estado = _carregar_estado()
                pasta_ini = (estado.get('config_pastas', {}).get('pasta_financeiro')
                             or os.path.expanduser('~'))
                p = filedialog.askopenfilename(
                    initialdir=_pasta_inicial(pasta_ini),
                    filetypes=[('Excel', '*.xlsx *.xls')])
                if p:
                    entry.delete(0, tk.END); entry.insert(0, p)
            def _limpar(): entry.delete(0, tk.END)

            tk.Button(frm_file, text='📁', font=('Segoe UI', 11), bg=COR_CARD2,
                      fg=COR_TEXTO, relief='flat', padx=8, pady=4,
                      cursor='hand2', command=_pick).pack(side='left', padx=(6, 0))
            btn_x = tk.Button(frm_file, text='✕', font=('Segoe UI', 10), bg=COR_CARD2,
                              fg=COR_TEXTO_SUB, relief='flat', padx=8, pady=4,
                              cursor='hand2', command=_limpar)
            btn_x.pack(side='left', padx=(4, 0))
            btn_x.bind('<Enter>', lambda e: btn_x.config(fg=COR_ERRO))
            btn_x.bind('<Leave>', lambda e: btn_x.config(fg=COR_TEXTO_SUB))
            return entry

        frm_inputs = tk.Frame(c, bg=COR_BG)
        frm_inputs.pack(fill='x', padx=32, pady=(12, 0))
        frm_pagar_col = tk.Frame(frm_inputs, bg=COR_BG)
        frm_pagar_col.pack(side='left', padx=(0, 32))
        entry_pagar = _entry_row(frm_pagar_col, '📕  Contas a Pagar (.xlsx)', COR_PAGAR)
        frm_receber_col = tk.Frame(frm_inputs, bg=COR_BG)
        frm_receber_col.pack(side='left')
        entry_receber = _entry_row(frm_receber_col, '📗  Contas a Receber (.xlsx)', COR_RECEBER)

        var_mes = _dropdown_mes(c, label='Mês de referência:', cor_label='#94a3b8')
        tk.Label(c, text=f'ℹ  Saída: {PASTA_FINANCEIRO}\\AAAA\\',
                 font=('Segoe UI', 7), bg=COR_BG, fg=COR_TEXTO_SUB,
                 justify='left').pack(anchor='w', padx=32)
        self._separador()

        frm_kpis = tk.Frame(c, bg=COR_BG)
        frm_kpis.pack(fill='x', padx=32, pady=(0, 4))
        txt, btn = self._log_e_btn('💰  Gerar Relatório Financeiro', COR_RECEBER)

        def _kpi_card(parent, label, valor, cor_borda, icone):
            outer = tk.Frame(parent, bg=cor_borda, padx=1, pady=1)
            outer.pack(side='left', padx=(0, 10), pady=4)
            inner = tk.Frame(outer, bg=BG_CARD, padx=14, pady=10)
            inner.pack()
            frm_top = tk.Frame(inner, bg=BG_CARD)
            frm_top.pack(anchor='w')
            tk.Label(frm_top, text=icone, font=('Consolas', 9),
                     bg=BG_CARD, fg=cor_borda).pack(side='left', padx=(0, 4))
            tk.Label(frm_top, text=label, font=('Consolas', 7),
                     bg=BG_CARD, fg='#6b7280').pack(side='left')
            tk.Frame(inner, bg=cor_borda, height=1).pack(fill='x', pady=(4, 6))
            tk.Label(inner, text=valor, font=('Consolas', 11, 'bold'),
                     bg=BG_CARD, fg=cor_borda).pack(anchor='w')

        def _mostrar_kpis(dados_pagar, dados_receber):
            for w in frm_kpis.winfo_children():
                w.destroy()
            for dados, cor_titulo, titulo, cor_total, cor_ok, cor_warn, cor_err in [
                (dados_pagar,   COR_PAGAR,   '─── 📕 CONTAS A PAGAR ───',
                 COR_PAGAR,   '#22c55e', '#f59e0b', '#ef4444'),
                (dados_receber, COR_RECEBER, '─── 📗 CONTAS A RECEBER ───',
                 COR_RECEBER, '#22c55e', '#f59e0b', '#ef4444'),
            ]:
                if not dados: continue
                tk.Label(frm_kpis, text=titulo, font=('Consolas', 8),
                         bg=COR_BG, fg=cor_titulo).pack(anchor='w', pady=(8, 4))
                frm_linha = tk.Frame(frm_kpis, bg=COR_BG)
                frm_linha.pack(anchor='w')
                _kpi_card(frm_linha, 'TOTAL GERAL',
                          f"R$ {dados['total_geral']:,.2f}",    cor_total, '◈')
                _kpi_card(frm_linha, 'QUITADOS',
                          f"R$ {dados['total_pago']:,.2f}",     cor_ok,    '✓')
                _kpi_card(frm_linha, 'PENDENTES',
                          f"R$ {dados['total_pendente']:,.2f}", cor_warn,  '◷')
                _kpi_card(frm_linha, 'ATRASADOS',
                          f"R$ {dados['total_atrasado']:,.2f}", cor_err,   '⚠')
                _kpi_card(frm_linha, 'QTD ATRASADOS',
                          str(dados['qtd_atrasado']),            cor_err,   '#')

        def _run():
            cp = entry_pagar.get().strip()
            cr = entry_receber.get().strip()
            if not cp and not cr:
                messagebox.showwarning('Atenção',
                    'Selecione ao menos um arquivo (Contas a Pagar ou Contas a Receber).')
                return
            mes_ref = var_mes.get()
            txt.configure(state='normal'); txt.delete('1.0', tk.END)
            txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Processando...')
            def _t():
                try:
                    ok = run_financeiro(
                        cp or None, cr or None, mes_ref,
                        lambda m: self.root.after(0, log_write, txt, m))
                    self.root.after(0, lambda: btn.configure(
                        state='normal', text='💰  Gerar Relatório Financeiro'))
                    if ok:
                        dp = processar_financeiro(cp, 'pagar',   lambda m: None) if cp else None
                        dr = processar_financeiro(cr, 'receber', lambda m: None) if cr else None
                        self.root.after(0, lambda: _mostrar_kpis(dp, dr))
                        self.root.after(0, lambda: self._resumo_sucesso('Financeiro', ok, txt))
                    else:
                        self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
                except Exception as e:
                    self.root.after(0, lambda: btn.configure(
                        state='normal', text='💰  Gerar Relatório Financeiro'))
                    self.root.after(0, lambda err=str(e):
                        log_write(txt, f"❌ Erro inesperado: {err}\n"))
            threading.Thread(target=_t, daemon=True).start()
        btn.configure(command=_run)

    def _pagina_fat_arm(self):
        self._cabecalho('🏭  Faturamento Armazenagem',
                        'Pico de estoque m³ por cliente · SKUs na data do pico', COR_FAT_ARM)
        self._separador()
        c = self.frm_conteudo

        # ── Painéis lado a lado: DB Famílias + DB Configuração ────────────
        frm_paineis = tk.Frame(c, bg=COR_BG)
        frm_paineis.pack(fill='x', padx=32, pady=(8, 4))

        # Painel DB Famílias
        frm_db = tk.Frame(frm_paineis, bg=COR_CARD,
                          highlightbackground=COR_BORDA, highlightthickness=1)
        frm_db.pack(side='left', fill='both', expand=True, padx=(0, 8))

        tk.Label(frm_db, text='🗂️  Banco de Dados de Famílias',
                 font=('Segoe UI', 9, 'bold'), bg=COR_CARD,
                 fg=COR_TEXTO).pack(anchor='w', padx=12, pady=(10, 4))

        lbl_fam_status = tk.Label(frm_db, text='Carregando...',
                                   font=('Segoe UI', 8), bg=COR_CARD, fg=COR_TEXTO_SUB)
        lbl_fam_status.pack(anchor='w', padx=12, pady=(0, 6))

        def _atualizar_status_familias():
            total_skus, total_cli, ultima = _db_familias_info()
            if total_skus == 0:
                lbl_fam_status.config(text='⚠️  DB vazio — carregue o arquivo de SKUs.',
                                      fg=COR_AVISO)
            else:
                lbl_fam_status.config(
                    text=f'✅  {total_skus} SKUs · {total_cli} clientes · {ultima}',
                    fg=COR_ACCENT3)
        _atualizar_status_familias()

        frm_db_btns = tk.Frame(frm_db, bg=COR_CARD)
        frm_db_btns.pack(fill='x', padx=12, pady=(0, 10))

        def _carregar_familias():
            arq = filedialog.askopenfilename(
                title='Selecionar arquivo de Famílias (sistema ou correções)',
                filetypes=[('Excel', '*.xlsx *.xls')])
            if not arq: return
            log_buf = []
            def _log(m): log_buf.append(m)
            try:
                xl = pd.ExcelFile(arq)
                fmt = _detectar_formato_familias(xl)
            except Exception as e:
                messagebox.showerror('Erro', f'Não foi possível abrir o arquivo:\n{e}'); return

            # Detecta parcial
            db_atual = _carregar_db_familias()
            n_abas = len(xl.sheet_names)
            is_parcial = (fmt == 'sistema' and db_atual and
                          n_abas < len(db_atual))

            if fmt == 'sistema' and is_parcial:
                fmt_label = f'Exportação Parcial ({n_abas} cliente(s))'
                aviso = f'✅  Adiciona/atualiza apenas os {n_abas} cliente(s) do arquivo.\n   Os demais {len(db_atual)} clientes já cadastrados são mantidos.'
            elif fmt == 'sistema':
                fmt_label = 'Exportação do Sistema (carga completa)'
                aviso = '⚠️  Isso vai substituir todo o DB de famílias.'
            else:
                fmt_label = 'Planilha de Correções (merge seguro)'
                aviso = '✅  Só serão preenchidos os SKUs que estão SEM FAMÍLIA.'

            if not messagebox.askyesno('Confirmar Importação',
                f'Formato detectado:\n{fmt_label}\n\n{aviso}\n\nDeseja continuar?'):
                return

            db = _carregar_familias_xlsx(arq, _log)
            if db:
                _salvar_db_familias(db); _atualizar_status_familias()
                messagebox.showinfo('✅ DB Atualizado',
                    f'{sum(len(v) for v in db.values())} SKUs em {len(db)} clientes.\n\n'
                    + ''.join(log_buf))
            else:
                messagebox.showerror('Erro', 'Nenhum dado encontrado.\n\n' + ''.join(log_buf))

        tk.Button(frm_db_btns, text='📥  Carregar / Atualizar Famílias',
                  font=('Segoe UI', 8, 'bold'), bg='#4C1D95', fg='white',
                  relief='flat', padx=10, pady=5, cursor='hand2',
                  command=_carregar_familias).pack(side='left')

        # Painel DB Configuração (preços + grupos/famílias)
        frm_cfg = tk.Frame(frm_paineis, bg=COR_CARD,
                           highlightbackground=COR_BORDA, highlightthickness=1)
        frm_cfg.pack(side='left', fill='both', expand=True)

        tk.Label(frm_cfg, text='💰  Configuração (Preços e Agrupamentos)',
                 font=('Segoe UI', 9, 'bold'), bg=COR_CARD,
                 fg=COR_TEXTO).pack(anchor='w', padx=12, pady=(10, 4))

        lbl_cfg_status = tk.Label(frm_cfg, text='Carregando...',
                                   font=('Segoe UI', 8), bg=COR_CARD, fg=COR_TEXTO_SUB)
        lbl_cfg_status.pack(anchor='w', padx=12, pady=(0, 6))

        def _atualizar_status_cfg():
            total, ultima = _db_precos_arm_info()
            if total == 0:
                lbl_cfg_status.config(
                    text='⚠️  DB vazio — carregue o arquivo de configuração.',
                    fg=COR_AVISO)
            else:
                lbl_cfg_status.config(
                    text=f'✅  {total} clientes com preço · Atualizado: {ultima}',
                    fg=COR_ACCENT3)
        _atualizar_status_cfg()

        frm_cfg_btns = tk.Frame(frm_cfg, bg=COR_CARD)
        frm_cfg_btns.pack(fill='x', padx=12, pady=(0, 10))

        def _carregar_cfg():
            arq = filedialog.askopenfilename(
                title='Selecionar arquivo de Configuração (Grupo-Família e Preços)',
                filetypes=[('Excel', '*.xlsx *.xls')])
            if not arq: return
            log_buf = []
            db = _carregar_config_fat_arm_xlsx(arq, lambda m: log_buf.append(m))
            if db:
                _salvar_db_precos_arm(db); _atualizar_status_cfg()
                n = len(db.get('clientes', {}))
                messagebox.showinfo('✅ Configuração Atualizada',
                    f'{n} clientes configurados.\n\n' + ''.join(log_buf))
            else:
                messagebox.showerror('Erro', 'Erro ao carregar.\n\n' + ''.join(log_buf))

        tk.Button(frm_cfg_btns, text='📥  Carregar Configuração',
                  font=('Segoe UI', 8, 'bold'), bg='#065F46', fg='white',
                  relief='flat', padx=10, pady=5, cursor='hand2',
                  command=_carregar_cfg).pack(side='left')

        tk.Label(c,
                 text='ℹ  Selecione o Banco de Dados de Movimentação (Excel, abas por cliente)\n'
                      '   e o arquivo de volumes (mesmo arquivo usado no módulo de Estoque).\n'
                      '   O relatório identifica o dia de maior volume m³ de cada cliente\n'
                      '   e lista o saldo de cada SKU nessa data.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB,
                 justify='left').pack(anchor='w', padx=32, pady=(8, 4))

        def _file_row_fat(parent, label, tipos):
            frm = tk.Frame(parent, bg=COR_BG)
            frm.pack(fill='x', padx=32, pady=(0, 8))
            tk.Label(frm, text=label, font=('Segoe UI', 8, 'bold'),
                     bg=COR_BG, fg=COR_FAT_ARM).pack(anchor='w')
            frm_inp = tk.Frame(frm, bg=COR_BG)
            frm_inp.pack(fill='x')
            entry = estilo_entry(frm_inp, width=60)
            entry.pack(side='left', ipady=5)
            def _pick():
                p = filedialog.askopenfilename(filetypes=tipos)
                if p: entry.delete(0, tk.END); entry.insert(0, p)
            def _limpar(): entry.delete(0, tk.END)
            tk.Button(frm_inp, text='📁', font=('Segoe UI', 11), bg=COR_CARD2,
                      fg=COR_TEXTO, relief='flat', padx=8, pady=4,
                      cursor='hand2', command=_pick).pack(side='left', padx=(6, 0))
            btn_x = tk.Button(frm_inp, text='✕', font=('Segoe UI', 10), bg=COR_CARD2,
                              fg=COR_TEXTO_SUB, relief='flat', padx=8, pady=4,
                              cursor='hand2', command=_limpar)
            btn_x.pack(side='left', padx=(4, 0))
            btn_x.bind('<Enter>', lambda e: btn_x.config(fg=COR_ERRO))
            btn_x.bind('<Leave>', lambda e: btn_x.config(fg=COR_TEXTO_SUB))
            return entry

        entry_mov = _file_row_fat(c, '📊  Banco de Dados de Movimentação (Excel):',
                                     [('Excel', '*.xlsx *.xls'), ('Todos', '*.*')])
        entry_vol = _file_row_fat(c, '📊  Arquivo de Volumes por Cliente (Excel):',
                                     [('Excel', '*.xlsx *.xls'), ('Todos', '*.*')])

        # Pré-preenche com o último arquivo de movimentação usado no Estoque
        ultimo_mov = _carregar_estado().get('ultimo_arquivo_mov', '')
        if ultimo_mov and os.path.isfile(ultimo_mov):
            entry_mov.configure(state='normal')
            entry_mov.delete(0, tk.END)
            entry_mov.insert(0, ultimo_mov)

        var_mes = _dropdown_mes(c, label='Mês de referência:', cor_label=COR_FAT_ARM)
        tk.Label(c, text=f'ℹ  Saída: {PASTA_FINANCEIRO}\\AAAA\\',
                 font=('Segoe UI', 7), bg=COR_BG, fg=COR_TEXTO_SUB).pack(anchor='w', padx=32)

        self._separador()
        txt, btn = self._log_e_btn('Gerar Faturamento Armazenagem', COR_FAT_ARM)

        def _run():
            ck = entry_mov.get().strip()
            cv = entry_vol.get().strip()
            if not ck or not cv:
                messagebox.showwarning('Atenção',
                    'Selecione o Banco de Dados de Movimentação e o arquivo de volumes.')
                return
            # Usa automaticamente o arquivo de estoque salvo no módulo de Estoque
            ce = _carregar_estado().get('ultimo_arquivo_estoque') or None
            mes_ref = var_mes.get()
            txt.configure(state='normal'); txt.delete('1.0', tk.END)
            txt.configure(state='disabled')
            log_write(txt, f"⏱ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}\n{'─'*52}\n")
            btn.configure(state='disabled', text='⏳  Processando...')
            def _t():
                try:
                    ok = run_faturamento_armazenagem(
                        ck, cv, mes_ref,
                        lambda m: self.root.after(0, log_write, txt, m),
                        arquivo_estoque=ce)
                    self.root.after(0, lambda: btn.configure(
                        state='normal', text='Gerar Faturamento Armazenagem'))
                    if ok:
                        self.root.after(0, lambda: self._resumo_sucesso('Financeiro', ok, txt))
                    else:
                        self.root.after(0, messagebox.showerror, 'Erro', '❌ Verifique o log.')
                except Exception as e:
                    self.root.after(0, lambda: btn.configure(
                        state='normal', text='Gerar Faturamento Armazenagem'))
                    self.root.after(0, lambda err=str(e):
                        log_write(txt, f"❌ Erro inesperado: {err}\n"))
            threading.Thread(target=_t, daemon=True).start()
        btn.configure(command=_run)

    # ----------------------------------------------------------
    # DE-PARA EDITOR
    # ----------------------------------------------------------
    def _pagina_de_para(self):
        """Editor de De-Para com virtualização — renderiza só linhas visíveis."""
        self._cabecalho('🔗  De-Para SKU → Cliente',
                        'Mapeie cada código de produto para o cliente correspondente', COR_ACCENT4)
        self._separador()
        c = self.frm_conteudo

        # ── Dados em memória (nunca como widgets) ──────────────────────────
        de_para = _carregar_de_para()
        # Lista mutável: cada item é [cod, cli] — editável diretamente
        dp_data = [[k, v] for k, v in sorted(de_para.items())]
        dp_filtrado = list(dp_data)   # view filtrada
        dp_editando = {}              # idx_visivel -> (cod_var, cli_var, frm)

        ROW_H    = 38   # altura de cada linha em pixels
        PAD_ROWS = 4    # linhas extras acima/abaixo da viewport

        # ── Toolbar ────────────────────────────────────────────────────────
        frm_tb = tk.Frame(c, bg=COR_BG)
        frm_tb.pack(fill='x', padx=32, pady=(0,6))

        lbl_count = tk.Label(frm_tb, font=('Segoe UI', 9), bg=COR_BG, fg=COR_TEXTO_SUB)
        lbl_count.pack(side='left')

        def _atualizar_count():
            lbl_count.config(text=f'{len(dp_data)} SKUs  ·  {len(dp_filtrado)} exibidos')

        # Busca
        frm_busca = tk.Frame(frm_tb, bg=COR_CARD2, highlightbackground=COR_BORDA, highlightthickness=1)
        frm_busca.pack(side='left', padx=(16,0))
        tk.Label(frm_busca, text='🔍', bg=COR_CARD2, fg=COR_TEXTO_SUB,
                 font=('Segoe UI', 9)).pack(side='left', padx=(8,2))
        busca_var = tk.StringVar()
        ent_busca = tk.Entry(frm_busca, textvariable=busca_var,
                             font=('Segoe UI', 9), bg=COR_CARD2, fg=COR_TEXTO,
                             insertbackground=COR_TEXTO, relief='flat', width=24)
        ent_busca.pack(side='left', ipady=4, padx=(0,8))

        def _salvar():
            # Flush widgets visíveis para dp_filtrado (in-place, propaga para dp_data)
            for idx_v, (cv, clv, _) in dp_editando.items():
                if idx_v < len(dp_filtrado):
                    dp_filtrado[idx_v][0] = cv.get().strip()
                    dp_filtrado[idx_v][1] = clv.get().strip()
            # Salva todos os itens válidos de dp_data (preserva duplicatas)
            # Usa o ÚLTIMO cliente para cada código duplicado
            resultado = {}
            for r in dp_data:
                cod, cli = r[0].strip(), r[1].strip()
                if cod and cli:
                    resultado[cod] = cli
            _salvar_de_para(resultado)
            n_salvos = len(resultado)
            n_duplas = len([r for r in dp_data if r[0].strip() and r[1].strip()]) - n_salvos
            msg = f'✅ {n_salvos} SKUs salvos!'
            if n_duplas > 0:
                msg += f'\n⚠️ {n_duplas} entradas duplicadas foram mescladas (mesmo código, clientes diferentes).'
            messagebox.showinfo('Salvo', msg)
            _atualizar_count()

        def _importar():
            path = filedialog.askopenfilename(filetypes=[('Excel','*.xlsx *.xls')])
            if not path: return
            try:
                # dtype=str preserva zeros à esquerda e evita conversão numérica
                df = pd.read_excel(path, dtype=str)
                df.columns = df.columns.str.strip()
                col_c  = next((x for x in df.columns if 'cod' in x.lower() or 'sku' in x.lower()), None)
                col_cl = next((x for x in df.columns if 'cliente' in x.lower() or 'client' in x.lower()), None)
                if not col_c or not col_cl:
                    messagebox.showwarning('Atenção',
                        f'Esperado: Código e Cliente\nDisponível: {list(df.columns)}')
                    return
                dp_data.clear()
                descartados = []
                for i, row in df.iterrows():
                    cod = str(row[col_c]).strip() if pd.notna(row[col_c]) else ''
                    cli = str(row[col_cl]).strip() if pd.notna(row[col_cl]) else ''
                    # Remove .0 de números inteiros lidos como float string
                    if cod.endswith('.0') and cod[:-2].isdigit():
                        cod = cod[:-2]
                    if cod and cli and cod.lower() != 'nan' and cli.lower() != 'nan':
                        dp_data.append([cod, cli])
                    else:
                        descartados.append(f'Linha {i+2}: cod={repr(cod)} cli={repr(cli)}')
                # Detecta duplicatas de código
                codigos = [r[0] for r in dp_data]
                vistos = {}
                duplas = []
                for r in dp_data:
                    cod = r[0]
                    if cod in vistos:
                        duplas.append(cod)
                    vistos[cod] = True

                _filtrar()
                msg = f'✅ {len(dp_data)} linhas importadas!'
                if duplas:
                    msg += f'\n⚠️ {len(duplas)} códigos duplicados encontrados.'
                    msg += f'\nAo salvar, será mantida apenas a última ocorrência de cada código.'
                    msg += f'\nTotal único após salvar: {len(vistos)}'
                    msg += f'\n\nExemplos duplicados: {", ".join(duplas[:5])}'
                else:
                    msg += f'\n{len(dp_data)} SKUs únicos. Clique em Salvar para confirmar.'
                if descartados:
                    msg += f'\n\n{len(descartados)} linhas vazias ignoradas.'
                messagebox.showinfo('Importado', msg)
            except Exception as e:
                messagebox.showerror('Erro', str(e))

        def _adicionar():
            # Flush edits visíveis
            for idx_v, (cv, clv, _) in dp_editando.items():
                if idx_v < len(dp_filtrado):
                    dp_filtrado[idx_v][0] = cv.get().strip()
                    dp_filtrado[idx_v][1] = clv.get().strip()
            novo = ['', '']
            dp_data.insert(0, novo)
            busca_var.set('')   # limpa busca para novo aparecer
            _filtrar()
            # Foca no primeiro campo
            if 0 in dp_editando:
                dp_editando[0][0].set('')
                # foca entry
                try: dp_editando[0][2].winfo_children()[0].focus_set()
                except: pass

        tk.Button(frm_tb, text='➕ Adicionar', font=('Segoe UI', 8, 'bold'),
                  bg=COR_ACCENT4, fg='white', relief='flat', padx=10, pady=5,
                  cursor='hand2', command=_adicionar).pack(side='right', padx=(6,0))
        tk.Button(frm_tb, text='💾 Salvar', font=('Segoe UI', 8, 'bold'),
                  bg=COR_ACCENT3, fg='white', relief='flat', padx=10, pady=5,
                  cursor='hand2', command=_salvar).pack(side='right', padx=(6,0))
        tk.Button(frm_tb, text='📥 Importar Excel', font=('Segoe UI', 8, 'bold'),
                  bg=COR_CARD2, fg=COR_TEXTO, relief='flat', padx=10, pady=5,
                  cursor='hand2', command=_importar).pack(side='right', padx=(6,0))

        # ── Header ─────────────────────────────────────────────────────────
        frm_hdr = tk.Frame(c, bg=COR_CARD2)
        frm_hdr.pack(fill='x', padx=32)
        tk.Label(frm_hdr, text='Código SKU', font=('Segoe UI', 9, 'bold'),
                 bg=COR_CARD2, fg=COR_TEXTO, width=22, anchor='w').pack(side='left', padx=(12,0), pady=5)
        tk.Label(frm_hdr, text='Cliente', font=('Segoe UI', 9, 'bold'),
                 bg=COR_CARD2, fg=COR_TEXTO, anchor='w').pack(side='left', padx=(8,0))

        # ── Canvas virtualizado ────────────────────────────────────────────
        frm_outer = tk.Frame(c, bg=COR_BG)
        frm_outer.pack(fill='both', expand=True, padx=32, pady=(2,12))

        sb_v = ttk.Scrollbar(frm_outer, orient='vertical')
        sb_v.pack(side='right', fill='y')
        canvas = tk.Canvas(frm_outer, bg=COR_BG, highlightthickness=0,
                           yscrollcommand=sb_v.set)
        canvas.pack(side='left', fill='both', expand=True)
        sb_v.config(command=canvas.yview)

        # Frame interno — altura = total de linhas * ROW_H
        frm_inner = tk.Frame(canvas, bg=COR_BG)
        cw = canvas.create_window((0,0), window=frm_inner, anchor='nw')

        def _on_canvas_resize(e):
            canvas.itemconfig(cw, width=e.width)
        canvas.bind('<Configure>', _on_canvas_resize)

        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>',
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), 'units')))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

        # ── Virtualização ──────────────────────────────────────────────────
        rendered_range = [0, 0]   # [first_visible_idx, last_visible_idx]

        def _flush_edits():
            """Lê valores dos widgets visíveis de volta para dp_filtrado."""
            for idx_v, (cv, clv, _) in list(dp_editando.items()):
                if idx_v < len(dp_filtrado):
                    dp_filtrado[idx_v][0] = cv.get().strip()
                    dp_filtrado[idx_v][1] = clv.get().strip()

        def _render_viewport(*_):
            _flush_edits()
            n = len(dp_filtrado)
            if n == 0:
                for w in frm_inner.winfo_children(): w.destroy()
                dp_editando.clear()
                tk.Label(frm_inner, text='Nenhum SKU encontrado.',
                         font=('Segoe UI', 9), bg=COR_BG, fg=COR_TEXTO_SUB).pack(pady=20)
                canvas.configure(scrollregion=(0,0,0,50))
                return

            total_h = n * ROW_H
            canvas.configure(scrollregion=(0, 0, canvas.winfo_width(), total_h))

            # Calcula índices visíveis
            canvas_h = canvas.winfo_height() or 400
            try:
                top_frac = canvas.yview()[0]
            except Exception:
                top_frac = 0
            first = max(0, int(top_frac * n) - PAD_ROWS)
            last  = min(n, first + int(canvas_h / ROW_H) + PAD_ROWS * 2 + 2)

            if rendered_range[0] == first and rendered_range[1] == last:
                return  # nada mudou

            # Remove widgets fora do range — flush antes de destruir
            for idx_v in list(dp_editando.keys()):
                if idx_v < first or idx_v >= last:
                    cv, clv, frm = dp_editando[idx_v]
                    if idx_v < len(dp_filtrado):
                        dp_filtrado[idx_v][0] = cv.get().strip()
                        dp_filtrado[idx_v][1] = clv.get().strip()
                    frm.place_forget()
                    frm.destroy()
                    del dp_editando[idx_v]

            # Cria/atualiza widgets visíveis
            w_canvas = canvas.winfo_width() or 900
            for idx_v in range(first, last):
                if idx_v in dp_editando:
                    continue
                row = dp_filtrado[idx_v]
                bg = COR_CARD if idx_v % 2 == 0 else COR_CARD2

                frm_r = tk.Frame(frm_inner, bg=bg,
                                 highlightbackground=COR_BORDA, highlightthickness=1)
                frm_r.place(x=0, y=idx_v * ROW_H, width=w_canvas, height=ROW_H)

                cod_var = tk.StringVar(value=row[0])
                cli_var = tk.StringVar(value=row[1])

                e_c = tk.Entry(frm_r, textvariable=cod_var, font=('Segoe UI',9),
                               bg=bg, fg=COR_TEXTO, insertbackground=COR_TEXTO,
                               relief='flat', width=22, bd=0)
                e_c.pack(side='left', padx=(10,4), pady=4, ipady=3)

                e_cl = tk.Entry(frm_r, textvariable=cli_var, font=('Segoe UI',9),
                                bg=bg, fg=COR_TEXTO, insertbackground=COR_TEXTO,
                                relief='flat', bd=0)
                e_cl.pack(side='left', padx=(0,4), pady=4, ipady=3, fill='x', expand=True)

                def _del(i=idx_v):
                    _flush_edits()
                    item = dp_filtrado[i]
                    if item in dp_data:
                        dp_data.remove(item)
                    dp_filtrado.pop(i)
                    # Remove todos widgets e re-renderiza
                    for v in list(dp_editando.values()):
                        v[2].destroy()
                    dp_editando.clear()
                    rendered_range[0] = rendered_range[1] = 0
                    _atualizar_count()
                    _render_viewport()

                tk.Button(frm_r, text='✕', font=('Segoe UI', 8),
                          bg=bg, fg=COR_ERRO, relief='flat',
                          padx=4, pady=2, cursor='hand2',
                          command=_del).pack(side='right', padx=6)

                dp_editando[idx_v] = (cod_var, cli_var, frm_r)

            rendered_range[0] = first
            rendered_range[1] = last
            frm_inner.configure(height=total_h)

        def _filtrar(*_):
            _flush_edits()
            termo = busca_var.get().strip().lower()
            dp_filtrado.clear()
            dp_filtrado.extend(
                r for r in dp_data
                if not termo or termo in r[0].lower() or termo in r[1].lower()
            )
            for v in list(dp_editando.values()):
                v[2].destroy()
            dp_editando.clear()
            rendered_range[0] = rendered_range[1] = 0
            _atualizar_count()
            canvas.yview_moveto(0)
            _render_viewport()

        busca_var.trace_add('write', _filtrar)
        canvas.bind('<Configure>', lambda e: [_on_canvas_resize(e), _render_viewport()])
        sb_v.config(command=lambda *a: [canvas.yview(*a), _render_viewport()])
        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>',
            lambda e: [canvas.yview_scroll(int(-1*(e.delta/120)), 'units'), _render_viewport()]))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

        _filtrar()
        _atualizar_count()

    # ----------------------------------------------------------
    # LEMBRETE DE EXTRAÇÃO
    # ----------------------------------------------------------
    def _verificar_lembrete(self):
        estado    = _carregar_estado()
        hoje      = date.today()
        dias_mod  = estado.get('dias_extracao_modulos', {})
        dia_fixo  = estado.get('dia_extracao', 6)
        extracoes = estado.get('extrações', {})
        hoje_str  = hoje.strftime('%d/%m/%Y')
        # Módulos cujo dia de extração é hoje E ainda não foram extraídos hoje
        pendentes = [
            nome for nome in RELATORIOS_CONFIG
            if hoje.day == dias_mod.get(nome, dia_fixo)
            and extracoes.get(nome, '') != hoje_str
        ]
        if not pendentes:
            return
        lista = '\n'.join(f'  • {n}' for n in pendentes)
        messagebox.showinfo(
            '📅 Lembrete de Extração',
            f'Hoje é dia {hoje.day} — há relatórios para gerar!\n\n'
            f'Pendentes:\n{lista}\n\n'
            f'Acesse a barra lateral para gerar cada um.'
        )

    # ----------------------------------------------------------
    # CONFIGURAÇÕES
    # ----------------------------------------------------------
    def _pagina_config(self):
        self._cabecalho('⚙️  Configurações',
                        'Edite as pastas padrão e preferências dos módulos', '#64748b')
        self._separador()
        c = self.frm_conteudo

        # Scroll area
        frm_outer = tk.Frame(c, bg=COR_BG)
        frm_outer.pack(fill='both', expand=True, padx=32, pady=(0,16))
        canvas = tk.Canvas(frm_outer, bg=COR_BG, highlightthickness=0)
        sb_sc  = ttk.Scrollbar(frm_outer, orient='vertical', command=canvas.yview)
        canvas.configure(yscrollcommand=sb_sc.set)
        sb_sc.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)
        frm_s = tk.Frame(canvas, bg=COR_BG)
        cw    = canvas.create_window((0,0), window=frm_s, anchor='nw')
        canvas.bind('<Configure>', lambda e: canvas.itemconfig(cw, width=e.width))
        frm_s.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))
        canvas.bind('<Enter>', lambda e: canvas.bind_all('<MouseWheel>',
            lambda e: canvas.yview_scroll(int(-1*(e.delta/120)), 'units')))
        canvas.bind('<Leave>', lambda e: canvas.unbind_all('<MouseWheel>'))

        estado   = _carregar_estado()
        cfg_edit = estado.get('config_pastas', {})

        entries = {}

        def _secao(titulo, cor):
            tk.Frame(frm_s, bg=cor, height=3).pack(fill='x', pady=(16,0))
            tk.Label(frm_s, text=titulo, font=('Segoe UI', 11, 'bold'),
                     bg=COR_BG, fg=COR_TEXTO).pack(anchor='w', pady=(4,8))

        def _campo(label, chave, valor_padrao):
            tk.Label(frm_s, text=label, font=('Segoe UI', 8),
                     bg=COR_BG, fg=COR_TEXTO_SUB).pack(anchor='w')
            frm_e = tk.Frame(frm_s, bg=COR_BG)
            frm_e.pack(fill='x', pady=(2,8))
            ent = estilo_entry(frm_e, width=70)
            ent.insert(0, cfg_edit.get(chave, valor_padrao))
            ent.pack(side='left', ipady=4, fill='x', expand=True)
            def _browse():
                p = filedialog.askdirectory()
                if p: ent.delete(0, tk.END); ent.insert(0, p)
            tk.Button(frm_e, text='📁', font=('Segoe UI', 10), bg=COR_CARD2,
                      fg=COR_TEXTO, relief='flat', padx=6, pady=3,
                      cursor='hand2', command=_browse).pack(side='left', padx=(6,0))
            entries[chave] = ent

        _secao('📦  Pedidos e Recebimentos', COR_ACCENT)
        _campo('Pasta padrão', 'pasta_pedidos', PASTA_PEDIDOS)

        _secao('🚚  Fretes', COR_ACCENT2)
        _campo('Pasta padrão', 'pasta_fretes', PASTA_FRETES)

        _secao('🏭  Armazenagem', COR_ACCENT3)
        _campo('Pasta padrão', 'pasta_armazenagem', PASTA_ARMAZENAGEM)

        _secao('📋  Estoque', COR_ACCENT4)
        _campo('Pasta padrão', 'pasta_estoque', PASTA_ESTOQUE)

        _secao('👥  Produtividade de Equipe', COR_ACCENT5)
        _campo('Pasta padrão', 'pasta_produtividade', PASTA_PRODUTIVIDADE)

        _secao('⚙️  Capacidade Operacional', COR_ACCENT7)
        _campo('Pasta padrão', 'pasta_cap_operacional', PASTA_CAP_OPERACIONAL)

        _secao('📥  Recebimentos e Devoluções', COR_ACCENT8)
        _campo('Pasta padrão', 'pasta_recebimentos', PASTA_RECEBIMENTOS)

        _secao('💰  Financeiro', COR_RECEBER)
        _campo('Pasta padrão', 'pasta_financeiro', PASTA_FINANCEIRO)

        _secao('📁  Consolidado Histórico', COR_ACCENT6)
        _campo('Pasta de saída', 'pasta_consolidados', PASTA_CONSOLIDADOS)

        def _salvar_config():
            vals = {k: e.get().strip() for k, e in entries.items()}
            est  = _carregar_estado()
            est.setdefault('config_pastas', {}).update(vals)
            # Garante que pastas de saída nunca sejam sobrescritas pelo formulário
            est['config_pastas'].update({
                'pasta_pedidos':          PASTA_PEDIDOS,
                'pasta_fretes':           PASTA_FRETES,
                'pasta_armazenagem':      PASTA_ARMAZENAGEM,
                'pasta_estoque':          PASTA_ESTOQUE,
                'pasta_produtividade':    PASTA_PRODUTIVIDADE,
                'pasta_cap_operacional':  PASTA_CAP_OPERACIONAL,
                'pasta_recebimentos':      PASTA_RECEBIMENTOS,
                'pasta_financeiro':        PASTA_FINANCEIRO,
            'pasta_financeiro':        PASTA_FINANCEIRO,
            'pasta_recebimentos':      PASTA_RECEBIMENTOS,
            'pasta_financeiro':        PASTA_FINANCEIRO,
                'pasta_consolidados':     PASTA_CONSOLIDADOS,
            })
            _salvar_estado(est)
            messagebox.showinfo('Salvo', '✅ Configurações salvas com sucesso!')

        tk.Button(frm_s, text='💾  Salvar Configurações',
                  font=('Segoe UI', 11, 'bold'),
                  bg=COR_ACCENT, fg='white', relief='flat',
                  padx=24, pady=10, cursor='hand2',
                  command=_salvar_config).pack(pady=(20,8))
        tk.Label(frm_s,
                 text='As alterações de pasta serão aplicadas na próxima vez que abrir o seletor de arquivo.',
                 font=('Segoe UI', 8), bg=COR_BG, fg=COR_TEXTO_SUB).pack()

# ============================================================
# ENTRY POINT
# ============================================================
if __name__ == '__main__':
    root = tk.Tk()
    App(root)
    root.mainloop()